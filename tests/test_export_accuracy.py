from __future__ import annotations

import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

import cv2
import fitz
import numpy as np
from openpyxl import load_workbook

from omr_software import (
    MARK_TYPE_OPTION,
    MarkItem,
    OMRSoftware,
    QSettings,
    QtWidgets,
)


class ExcelAccuracyTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.settings_folder = tempfile.TemporaryDirectory()
        QSettings.setDefaultFormat(QSettings.IniFormat)
        QSettings.setPath(
            QSettings.IniFormat,
            QSettings.UserScope,
            cls.settings_folder.name,
        )
        cls.app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([])

    @classmethod
    def tearDownClass(cls):
        cls.settings_folder.cleanup()

    def setUp(self):
        self.window = OMRSoftware()
        self.window.first_page_key = False
        self.window.student_order = []
        self.window.extra_students = []
        self.window.student_absence = {1: True}
        self.window.answer_key = {1: "A", 2: ""}
        self.window.topic_map = {1: "Topic", 2: "Topic"}
        self.window.results = {
            0: {"options": {1: "", 2: ""}, "text": {}},
            1: {"options": {1: "A", 2: ""}, "text": {}},
        }
        self.window.check_include_summary.setChecked(True)
        self.window.check_include_topics.setChecked(True)

    def tearDown(self):
        self.window.close()
        self.app.processEvents()

    def test_blank_answers_and_absent_pages_are_not_counted_correct(self):
        with tempfile.TemporaryDirectory() as folder:
            output = Path(folder) / "results.xlsx"
            self.window._export_excel_internal(str(output))
            workbook = load_workbook(output, data_only=False)

        results = workbook["OMR Results"]
        stats_row = next(
            row for row in range(1, results.max_row + 1)
            if results.cell(row=row, column=1).value == "% Correct"
        )

        # With no text columns, B is Absent and questions start at C.
        self.assertEqual(results.cell(row=stats_row, column=3).value, 0)
        self.assertIsNone(results.cell(row=stats_row, column=4).value)
        score_formula = results.cell(row=3, column=5).value
        self.assertIn('<>""', score_formula)

        topic_analysis = workbook["Topic Analysis"]
        self.assertEqual(topic_analysis.cell(row=2, column=3).value, 0)
        self.assertEqual(topic_analysis.cell(row=2, column=4).value, 0)

    def test_template_load_is_atomic_and_applies_page_offset(self):
        self.window.page_offsets = {0: (10.0, 20.0)}
        valid = {
            "schema_version": 2,
            "option_marks": [
                {
                    "question": 1,
                    "label": "Q1",
                    "x": 5,
                    "y": 6,
                    "width": 100,
                    "height": 30,
                    "options_count": 4,
                }
            ],
        }
        self.window._load_template_data(valid)
        original = self.window.view.option_marks[0]
        self.assertAlmostEqual(original.pos().x(), 15.0)
        self.assertAlmostEqual(original.pos().y(), 26.0)

        invalid = {
            "option_marks": [
                valid["option_marks"][0],
                {
                    "question": 2,
                    "label": "Q2",
                    "x": 5,
                    "y": 50,
                    "width": 0,
                    "height": 30,
                    "options_count": 4,
                },
            ]
        }
        with self.assertRaises(ValueError):
            self.window._load_template_data(invalid)
        self.assertEqual(self.window.view.option_marks, [original])

        with self.assertRaisesRegex(ValueError, "no option or text regions"):
            self.window._load_template_data(
                {
                    "schema_version": 2,
                    "text_marks": [],
                    "option_marks": [],
                    "align_marks": [],
                }
            )
        self.assertEqual(self.window.view.option_marks, [original])

    def test_empty_template_cannot_be_exported(self):
        self.window.clear_all_marks()
        with (
            patch("omr_software.QMessageBox.warning") as warning,
            patch("omr_software.QFileDialog.getSaveFileName") as save_dialog,
        ):
            self.window.export_template()

        warning.assert_called_once()
        save_dialog.assert_not_called()

    def test_pdf_render_crop_and_omr_pipeline(self):
        page_image = np.full((300, 600, 3), 248, dtype=np.uint8)
        answer_region = np.full((32, 160, 3), 248, dtype=np.uint8)
        for index in range(4):
            center = (index * 40 + 20, 16)
            cv2.circle(answer_region, center, 8, (115, 115, 115), 1)
        cv2.circle(answer_region, (60, 16), 6, (35, 35, 35), -1)
        page_image[100:132, 100:260] = answer_region
        ok, encoded = cv2.imencode(
            ".png", cv2.cvtColor(page_image, cv2.COLOR_RGB2BGR)
        )
        self.assertTrue(ok)

        with tempfile.TemporaryDirectory() as folder:
            pdf_path = Path(folder) / "synthetic.pdf"
            document = fitz.open()
            page = document.new_page(width=300, height=150)
            page.insert_image(page.rect, stream=encoded.tobytes())
            document.save(pdf_path)
            document.close()

            self.window.check_auto_deskew.setChecked(False)
            self.window.check_auto_align.setChecked(False)
            self.window.check_recognize_text.setChecked(False)
            self.window._open_pdf_path(str(pdf_path))
            mark = MarkItem(
                0,
                0,
                160,
                32,
                MARK_TYPE_OPTION,
                1,
                "Q1",
                4,
                view_ref=self.window.view,
            )
            mark.setPos(100, 100)
            self.window.view.option_marks.append(mark)
            self.window.scene.addItem(mark)

            result = self.window._recognize_page(
                0, preserve_overrides=False
            )
            self.window.pdf_document.close()
            self.window.pdf_document = None
        self.assertEqual(result["options"][1], "B")


if __name__ == "__main__":
    unittest.main()
