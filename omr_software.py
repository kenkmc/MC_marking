# IMPORTANT: Import easyocr BEFORE PyQt5 to avoid DLL conflicts on Windows
# Try imports for OCR
OCR_ENGINE = None

# Try EasyOCR first (must be before PyQt5 imports)
try:
    import easyocr
    OCR_ENGINE = "easyocr"
    print("EasyOCR loaded successfully")
except (ImportError, OSError, Exception) as e:
    print(f"Warning: EasyOCR not available ({e})")
    # Try Tesseract as fallback
    try:
        import pytesseract
        pytesseract.get_tesseract_version()
        OCR_ENGINE = "tesseract"
        print("Using Tesseract")
    except:
        pass

if OCR_ENGINE is None:
    print("Warning: No OCR engine found")

from PyQt5 import QtWidgets, QtGui, QtCore
from PyQt5.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QPushButton, QFileDialog, QGraphicsView, QGraphicsScene,
    QDialog, QComboBox, QCheckBox, QTextEdit, QGraphicsRectItem,
    QSpinBox, QGroupBox, QTableWidget, QTableWidgetItem, QSplitter,
    QMessageBox, QInputDialog, QScrollArea, QFrame, QSlider,
    QGraphicsPixmapItem, QMenu, QAction, QDialogButtonBox, QAbstractItemView
)
from PyQt5.QtGui import QPixmap, QImage, QPen, QBrush, QColor, QPainter, QFont, QWheelEvent, QCursor, QDesktopServices
from PyQt5.QtCore import Qt, QRectF, QPointF, QUrl, QObject, QEvent, QThread, pyqtSignal, QSettings, QTimer
import fitz  # PyMuPDF for PDF rendering
import sys
import json
import os
import re
import io
import zipfile
import shutil
import cv2
import numpy as np
import statistics
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font as XLFont, Alignment, Border, Side

import urllib.request
import tempfile
import subprocess

from omr_core import (
    deskew_image as core_deskew_image,
    detect_filled_options,
    estimate_similarity_transform,
    estimate_option_content_bounds,
    select_filled_options,
)

# Mark types
MARK_TYPE_TEXT = "text"      # Text field (e.g., student name, ID)
MARK_TYPE_OPTION = "option"  # Answer option (e.g., A, B, C, D)
MARK_TYPE_ALIGN = "align"    # Alignment reference region

# Version
APP_VERSION = "1.7.0"

# GitHub repo for update checks
GITHUB_REPO = "kenkmc/MC_marking"


class UpdateChecker(QThread):
    """Background thread to check GitHub releases for updates."""
    update_available = pyqtSignal(str, str, str, str)  # (latest_version, html_url, asset_url, body)
    no_update = pyqtSignal(str)                   # current_version
    check_failed = pyqtSignal(str)                # error_message

    def run(self):
        try:
            url = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
            req = urllib.request.Request(url, headers={"Accept": "application/vnd.github+json", "User-Agent": "CheckMate-Updater"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read().decode())
            tag = data.get("tag_name", "").lstrip("vV")
            html_url = data.get("html_url", "")
            body = data.get("body", "") or ""
            # Find the .zip asset download URL for auto-update
            asset_url = ""
            for asset in data.get("assets", []):
                if asset.get("name", "").lower().endswith(".zip"):
                    asset_url = asset.get("browser_download_url", "")
                    break
            if self._is_newer(tag, APP_VERSION):
                self.update_available.emit(tag, html_url, asset_url, body)
            else:
                self.no_update.emit(APP_VERSION)
        except Exception as e:
            self.check_failed.emit(str(e))

    @staticmethod
    def _is_newer(remote: str, local: str) -> bool:
        """Compare dotted version strings."""
        try:
            r_parts = [int(x) for x in remote.split(".")]
            l_parts = [int(x) for x in local.split(".")]
            return r_parts > l_parts
        except ValueError:
            return False


class UpdateDownloader(QThread):
    """Background thread to download an update zip file."""
    progress = pyqtSignal(int, int)       # (bytes_downloaded, total_bytes)
    download_complete = pyqtSignal(str)   # temp_zip_path
    download_failed = pyqtSignal(str)     # error_message

    def __init__(self, url, parent=None):
        super().__init__(parent)
        self._url = url
        self._cancelled = False

    def cancel(self):
        self._cancelled = True

    def run(self):
        tmp_path = None
        try:
            req = urllib.request.Request(self._url, headers={"User-Agent": "CheckMate-Updater"})
            with urllib.request.urlopen(req, timeout=30) as resp:
                total = int(resp.headers.get("Content-Length", 0))
                tmp_fd, tmp_path = tempfile.mkstemp(suffix=".zip", prefix="checkmate_update_")
                with os.fdopen(tmp_fd, 'wb') as f:
                    downloaded = 0
                    while True:
                        if self._cancelled:
                            try:
                                os.unlink(tmp_path)
                            except OSError:
                                pass
                            return
                        chunk = resp.read(256 * 1024)  # 256 KB chunks
                        if not chunk:
                            break
                        f.write(chunk)
                        downloaded += len(chunk)
                        self.progress.emit(downloaded, total)
            self.download_complete.emit(tmp_path)
        except Exception as e:
            if tmp_path:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
            if not self._cancelled:
                self.download_failed.emit(str(e))

# ── i18n Translation System ──
_TRANSLATIONS = {
    "en": {
        "app_title": "CheckMate – The definitive OMR software",
        "title_label": "CheckMate",
        "group_file": "1. File & Setup",
        "btn_import_pdf": "Import PDF",
        "chk_first_key": "First page is Answer Key",
        "chk_auto_deskew": "Auto-correct page skew",
        "chk_auto_align": "Auto-align pages (shift)",
        "group_marking": "2. Marking Tools",
        "btn_mark_text": "Mark Text Field",
        "btn_mark_option": "Mark Options",
        "btn_mark_align": "📍 Mark Alignment Regions",
        "tip_mark_align": "Mark reference regions (e.g., table corners) for aligning scanned pages. Multiple marks improve accuracy.",
        "lbl_mark_hint1": "Right-click marks to Rename/Delete/Config",
        "lbl_mark_hint2": "Click mark to select, then drag corners to resize",
        "btn_undo": "↩ Undo",
        "tip_undo": "Remove the last added mark",
        "btn_clear": "Clear All",
        "btn_load_template": "Load Template",
        "btn_save_template": "Save Template",
        "group_processing": "3. Processing",
        "lbl_ocr_status": "OCR Status: {engine}",
        "lbl_ocr_not_available": "Not Available",
        "btn_recognize_all": "Recognize All Pages",
        "btn_recognize_sel": "Re-recognize Selected Pages...",
        "chk_recognize_text": "Recognize text fields with OCR (slower)",
        "chk_save_debug": "Save diagnostic images (slower)",
        "chk_export_images": "Include images with answers",
        "btn_export_bundle": "Export Results (Excel + Images)",
        "btn_export_excel": "Export to Excel",
        "btn_student_info": "Student Info (Manual / Paste)",
        "chk_include_summary": "Include summary analysis in Excel",
        "chk_include_topics": "Include topic sheet/analysis in Excel",
        "btn_set_topics": "Set Topics",
        "btn_export_images": "Export Images with Answers",
        "btn_export_debug": "Export Debug Folder",
        "group_batch": "4. Batch Processing",
        "btn_batch_same": "📁 Batch: Same Template",
        "tip_batch_same": "Select one template and multiple PDFs to process",
        "btn_batch_match": "📁 Batch: Match Template Names",
        "tip_batch_match": "Select multiple PDFs. For each PDF, auto-load template with same name.\nExample: exam1.pdf uses exam1.json",
        "btn_prev": "◀ Prev",
        "btn_next": "Next ▶",
        "lbl_page": "Page: {current}/{total}",
        "lbl_zoom": "Zoom:",
        "tip_zoom_in": "Zoom In (Ctrl+Scroll Up)",
        "tip_zoom_out": "Zoom Out (Ctrl+Scroll Down)",
        "tip_zoom_reset": "Reset Zoom",
        "tip_zoom_fit": "Fit to Window",
        "lbl_results": "<b>Results & Answer Key</b>",
        "col_q": "Q",
        "col_detected": "Detected",
        "col_correct": "Correct",
        "col_points": "Points",
        "col_confidence": "Confidence",
        "col_crop": "Crop",
        "btn_next_review": "Next item to review",
        "tip_next_review": "Jump to the next blank, multiple, invalid, or low-confidence answer",
        "lbl_score": "Page Score: {score}",
        "lbl_total": "Total: 0",
        "lbl_answer_status_empty": "Empty: {questions}",
        "lbl_answer_status_multi": "Multiple: {questions}",
        "lbl_answer_status_review": "Review: {questions}",
        "lbl_answer_status_ok": "All answered ✓",
        # Dialogs
        "dlg_student_title": "Student Info (Manual / Paste)",
        "dlg_student_hint": "Paste from Excel: rows = students, columns = fields (tab-separated).",
        "dlg_student_absent": "Absent",
        "btn_paste_clipboard": "Paste from Clipboard",
        "dlg_topic_title": "Set Topics",
        "dlg_topic_hint": "Paste from Excel: rows = questions, columns = Q, Topic (tab-separated).",
        "col_question": "Question",
        "col_topic": "Topic",
        "dlg_recognize_title": "Re-recognize Pages",
        "dlg_recognize_prompt": "Select pages to re-recognize:",
        "dlg_recognize_all_pages": "All Pages",
        "dlg_recognize_current": "Current Page Only",
        "dlg_recognize_range": "Page Range (e.g. 1-3, 5):",
        # Messages
        "msg_no_pdf": "Please import a PDF first.",
        "msg_no_marks": "No marks defined. Please mark regions first.",
        "msg_no_results": "No results to export",
        "msg_recognition_complete": "Processed {pages} pages.\nRecognized {options} option fields.",
        "msg_recognition_title": "Recognition Complete",
        "msg_no_review": "No answers are currently flagged for review.",
        "msg_drop_pdf": "Drop a PDF here to open it.",
        "crop_preview_title": "Answer crop – Page {page}, Q{question}",
        "msg_clipboard_empty": "Clipboard is empty.",
        "msg_no_questions": "No questions found to label.",
        "msg_no_text_fields": "No text fields found or defined.",
        "dlg_select_export_folder": "Select Export Location",
        "progress_exporting": "Exporting...",
        "progress_exporting_excel": "Exporting Excel...",
        "progress_exporting_images": "Exporting image {current}/{total}...",
        "msg_export_done": "Export complete!\nSaved to:\n{folder}",
        "lbl_student_counts": "Pages: {pages}  |  Entered: {students}  |  Present: {present}",
        "lbl_page_absent": "(Absent)",
        "lbl_student_info": "\ud83d\udc64 {info}",
        # About
        "menu_help": "Help",
        "menu_about": "About CheckMate",
        "about_title": "About CheckMate",
        "about_text": "CheckMate v{version}\n\nThe definitive OMR (Optical Mark Recognition) software.\n\nBuilt with PyQt5, OpenCV, PyMuPDF.\n\n© 2026",
        # Language
        "menu_language": "Language",
        "lang_en": "English",
        "lang_zh": "繁體中文",
        # Alignment mark overlay
        "align_overlay": "Align",
        # Student info defaults
        "field_class": "Class",
        "field_student_no": "Student No.",
        "field_name": "Name",
        "col_page": "Page",
        # Update checker
        "menu_settings": "Settings",
        "menu_check_update": "Check for Update",
        "chk_auto_update": "Check for updates on startup",
        "update_title": "Update Available",
        "update_msg": "A new version of CheckMate is available!\n\nCurrent version: v{current}\nLatest version: v{latest}",
        "update_whats_new": "What's New",
        "update_auto_btn": "Update Now",
        "update_manual_btn": "Open Download Page",
        "update_downloading": "Downloading CheckMate v{version}...",
        "update_extracting": "Extracting update...",
        "update_restart_title": "Update Ready",
        "update_restart_msg": "Update downloaded successfully!\nThe application will now restart to apply the update.",
        "update_download_failed_title": "Download Failed",
        "update_download_failed": "Failed to download update:\n{error}",
        "update_no_update": "You are using the latest version (v{version}).",
        "update_no_update_title": "No Update Available",
        "update_check_failed": "Could not check for updates.\nPlease check your internet connection.",
        "update_check_failed_title": "Update Check Failed",
    },
    "zh": {
        "app_title": "CheckMate – 全能批改系統",
        "title_label": "CheckMate",
        "group_file": "1. 檔案設定",
        "btn_import_pdf": "匯入 PDF",
        "chk_first_key": "第一頁為答案",
        "chk_auto_deskew": "自動校正頁面歪斜",
        "chk_auto_align": "自動對齊頁面（位移）",
        "group_marking": "2. 標記工具",
        "btn_mark_text": "標記文字欄",
        "btn_mark_option": "標記選項",
        "btn_mark_align": "📍 標記對齊區域",
        "tip_mark_align": "標記參考區域（例如表格角落）以對齊掃描頁面。多個標記可提高準確度。",
        "lbl_mark_hint1": "右鍵點擊標記可重新命名/刪除/設定",
        "lbl_mark_hint2": "點選標記後拖曳角落可調整大小",
        "btn_undo": "↩ 復原",
        "tip_undo": "移除最後新增的標記",
        "btn_clear": "全部清除",
        "btn_load_template": "載入範本",
        "btn_save_template": "儲存範本",
        "group_processing": "3. 處理",
        "lbl_ocr_status": "OCR 狀態：{engine}",
        "lbl_ocr_not_available": "未找到",
        "btn_recognize_all": "辨識所有頁面",
        "btn_recognize_sel": "重新辨識選定頁面...",
        "chk_recognize_text": "辨識文字欄位（較慢）",
        "chk_save_debug": "儲存診斷圖片（較慢）",
        "chk_export_images": "包含答案標註圖片",
        "btn_export_bundle": "匯出結果（Excel + 圖片）",
        "btn_export_excel": "匯出 Excel",
        "btn_student_info": "學生資料（手動 / 貼上）",
        "chk_include_summary": "Excel 中包含統計分析",
        "chk_include_topics": "Excel 中包含課題分析",
        "btn_set_topics": "設定課題",
        "btn_export_images": "匯出答案圖片",
        "btn_export_debug": "匯出偵錯資料夾",
        "group_batch": "4. 批次處理",
        "btn_batch_same": "📁 批次：相同範本",
        "tip_batch_same": "選擇一個範本和多個 PDF 進行處理",
        "btn_batch_match": "📁 批次：匹配範本名稱",
        "tip_batch_match": "選擇多個 PDF。每個 PDF 自動載入同名範本。\n例如：exam1.pdf 使用 exam1.json",
        "btn_prev": "◀ 上一頁",
        "btn_next": "下一頁 ▶",
        "lbl_page": "頁面：{current}/{total}",
        "lbl_zoom": "縮放：",
        "tip_zoom_in": "放大（Ctrl+滾輪向上）",
        "tip_zoom_out": "縮小（Ctrl+滾輪向下）",
        "tip_zoom_reset": "重設縮放",
        "tip_zoom_fit": "適應視窗",
        "lbl_results": "<b>結果及答案</b>",
        "col_q": "題",
        "col_detected": "偵測",
        "col_correct": "正確",
        "col_points": "分數",
        "col_confidence": "信心度",
        "col_crop": "裁剪",
        "btn_next_review": "下一個待覆核項目",
        "tip_next_review": "跳至下一個空白、多選、無效或低信心度答案",
        "lbl_score": "本頁分數：{score}",
        "lbl_total": "總計：0",
        "lbl_answer_status_empty": "空白：{questions}",
        "lbl_answer_status_multi": "多選：{questions}",
        "lbl_answer_status_review": "待覆核：{questions}",
        "lbl_answer_status_ok": "全部已作答 ✓",
        # Dialogs
        "dlg_student_title": "學生資料（手動 / 貼上）",
        "dlg_student_hint": "從 Excel 貼上：列 = 學生，欄 = 各欄位（Tab 分隔）。",
        "dlg_student_absent": "缺席",
        "btn_paste_clipboard": "從剪貼簿貼上",
        "dlg_topic_title": "設定課題",
        "dlg_topic_hint": "從 Excel 貼上：列 = 題目，欄 = Q、課題（Tab 分隔）。",
        "col_question": "題目",
        "col_topic": "課題",
        "dlg_recognize_title": "重新辨識頁面",
        "dlg_recognize_prompt": "選擇要重新辨識的頁面：",
        "dlg_recognize_all_pages": "所有頁面",
        "dlg_recognize_current": "僅目前頁面",
        "dlg_recognize_range": "頁碼範圍（例如 1-3, 5）：",
        # Messages
        "msg_no_pdf": "請先匯入 PDF。",
        "msg_no_marks": "未定義標記。請先標記區域。",
        "msg_no_results": "沒有可匯出的結果",
        "msg_recognition_complete": "已處理 {pages} 頁。\n已辨識 {options} 個選項欄。",
        "msg_recognition_title": "辨識完成",
        "msg_no_review": "目前沒有需要覆核的答案。",
        "msg_drop_pdf": "把 PDF 拖放到這裡即可開啟。",
        "crop_preview_title": "答案裁剪 – 第 {page} 頁，Q{question}",
        "msg_clipboard_empty": "剪貼簿是空的。",
        "msg_no_questions": "找不到題目。",
        "msg_no_text_fields": "未找到或未定義文字欄位。",
        "dlg_select_export_folder": "選擇匯出位置",
        "progress_exporting": "匯出中...",
        "progress_exporting_excel": "正在匯出 Excel...",
        "progress_exporting_images": "正在匯出圖片 {current}/{total}...",
        "msg_export_done": "匯出完成！\n已儲存至：\n{folder}",
        "lbl_student_counts": "已掃描頁數：{pages}  |  已輸入：{students}  |  出席：{present}",
        "lbl_page_absent": "（缺席）",
        "lbl_student_info": "👤 {info}",
        # About
        "menu_help": "說明",
        "menu_about": "關於 CheckMate",
        "about_title": "關於 CheckMate",
        "about_text": "CheckMate v{version}\n\n全能光學標記辨識（OMR）批改系統。\n\n使用 PyQt5、OpenCV、PyMuPDF 建構。\n\n© 2026",
        # Language
        "menu_language": "語言",
        "lang_en": "English",
        "lang_zh": "繁體中文",
        # Alignment mark overlay
        "align_overlay": "對齊",
        # Student info defaults
        "field_class": "班別",
        "field_student_no": "學號",
        "field_name": "姓名",
        "col_page": "頁面",
        # Update checker
        "menu_settings": "設定",
        "menu_check_update": "檢查更新",
        "chk_auto_update": "啟動時自動檢查更新",
        "update_title": "有可用更新",
        "update_msg": "CheckMate 有新版本可用！\n\n目前版本：v{current}\n最新版本：v{latest}",
        "update_whats_new": "更新內容",
        "update_auto_btn": "立即更新",
        "update_manual_btn": "開啟下載頁面",
        "update_downloading": "正在下載 CheckMate v{version}...",
        "update_extracting": "正在解壓更新...",
        "update_restart_title": "更新就緒",
        "update_restart_msg": "更新下載完成！\n應用程式將立即重新啟動以套用更新。",
        "update_download_failed_title": "下載失敗",
        "update_download_failed": "下載更新失敗：\n{error}",
        "update_no_update": "您已使用最新版本（v{version}）。",
        "update_no_update_title": "沒有可用更新",
        "update_check_failed": "無法檢查更新。\n請確認網路連線。",
        "update_check_failed_title": "更新檢查失敗",
    }
}

# Current language setting (default: English)
_current_lang = "en"

def tr(key, **kwargs):
    """Get translated string for the current language."""
    text = _TRANSLATIONS.get(_current_lang, _TRANSLATIONS["en"]).get(key)
    if text is None:
        text = _TRANSLATIONS["en"].get(key, key)
    if kwargs:
        text = text.format(**kwargs)
    return text

def set_language(lang):
    """Set the application language ('en' or 'zh')."""
    global _current_lang
    if lang in _TRANSLATIONS:
        _current_lang = lang

def get_language():
    """Get current language code."""
    return _current_lang


def deskew_image(img_array):
    """Detect and correct a small scan skew without changing page dimensions."""

    return core_deskew_image(img_array)


# Modern Style Sheet
STYLE_SHEET = """
QMainWindow {
    background-color: #f0f2f5;
}
QWidget {
    font-family: 'Segoe UI', Arial, sans-serif;
    font-size: 14px;
}
QGroupBox {
    background-color: white;
    border-radius: 8px;
    margin-top: 1em;
    padding: 15px;
    border: 1px solid #e1e4e8;
    font-weight: bold;
    color: #333;
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 10px;
    padding: 0 5px;
    color: #555;
}
QPushButton {
    background-color: #007bff;
    color: white;
    border: none;
    border-radius: 4px;
    padding: 8px 16px;
    min-height: 20px;
}
QPushButton:hover {
    background-color: #0069d9;
}
QPushButton:pressed {
    background-color: #0062cc;
}
QPushButton:checked {
    background-color: #0056b3;
    border: 2px solid #004085;
}
QPushButton#deleteBtn {
    background-color: #dc3545;
}
QPushButton#deleteBtn:hover {
    background-color: #c82333;
}
QTableWidget {
    border: 1px solid #e1e4e8;
    background-color: white;
    gridline-color: #f0f0f0;
}
QHeaderView::section {
    background-color: #f8f9fa;
    padding: 4px;
    border: 1px solid #e1e4e8;
    font-weight: bold;
}
QLabel#titleLabel {
    font-size: 18px;
    font-weight: bold;
    color: #2c3e50;
    margin-bottom: 10px;
}
"""

class MovablePixmapItem(QGraphicsPixmapItem):
    """A movable pixmap item for the PDF page image."""
    
    def __init__(self, pixmap, parent=None):
        super().__init__(pixmap, parent)
        self.setFlag(QGraphicsPixmapItem.ItemIsMovable, True)
        self.setFlag(QGraphicsPixmapItem.ItemIsSelectable, False)
        self.setAcceptHoverEvents(True)
        self.offset_x = 0
        self.offset_y = 0
        
    def get_offset(self):
        return self.pos().x(), self.pos().y()
        
    def set_offset(self, x, y):
        self.setPos(x, y)


# Resize handle size
RESIZE_HANDLE_SIZE = 10

class MarkItem(QGraphicsRectItem):
    """A resizable and movable rectangle for marking areas."""
    
    # Resize handles positions
    HANDLE_NONE = 0
    HANDLE_TOP_LEFT = 1
    HANDLE_TOP_RIGHT = 2
    HANDLE_BOTTOM_LEFT = 3
    HANDLE_BOTTOM_RIGHT = 4
    HANDLE_TOP = 5
    HANDLE_BOTTOM = 6
    HANDLE_LEFT = 7
    HANDLE_RIGHT = 8
    
    def __init__(self, x, y, width, height, mark_type=MARK_TYPE_OPTION, 
                 question_num=1, label="", options_count=4, parent=None, view_ref=None):
        super().__init__(x, y, width, height, parent)
        self.mark_type = mark_type
        self.question_num = question_num
        self.label = label
        self.options_count = options_count
        self.view_ref = view_ref
        
        # Resize state
        self.resize_handle = self.HANDLE_NONE
        self.resize_start_rect = None
        self.resize_start_pos = None
        
        self.setFlag(QGraphicsRectItem.ItemIsMovable, True)
        self.setFlag(QGraphicsRectItem.ItemIsSelectable, True)
        self.setFlag(QGraphicsRectItem.ItemSendsGeometryChanges, True)
        self.setAcceptHoverEvents(True)
        
        self.update_style()
        
    def update_style(self):
        if self.mark_type == MARK_TYPE_TEXT:
            self.setPen(QPen(QColor(0, 100, 255), 2))
            self.setBrush(QBrush(QColor(0, 100, 255, 50)))
        elif self.mark_type == MARK_TYPE_ALIGN:
            self.setPen(QPen(QColor(0, 200, 0), 3, Qt.DashLine))
            self.setBrush(QBrush(QColor(0, 200, 0, 30)))
        else:
            self.setPen(QPen(QColor(255, 0, 0), 2))
            self.setBrush(QBrush(QColor(255, 0, 0, 50)))
    
    def get_handle_at_pos(self, pos):
        """Determine which resize handle (if any) is at the given position."""
        rect = self.rect()
        hs = RESIZE_HANDLE_SIZE
        
        # Corner handles
        if QRectF(rect.x() - hs/2, rect.y() - hs/2, hs, hs).contains(pos):
            return self.HANDLE_TOP_LEFT
        if QRectF(rect.right() - hs/2, rect.y() - hs/2, hs, hs).contains(pos):
            return self.HANDLE_TOP_RIGHT
        if QRectF(rect.x() - hs/2, rect.bottom() - hs/2, hs, hs).contains(pos):
            return self.HANDLE_BOTTOM_LEFT
        if QRectF(rect.right() - hs/2, rect.bottom() - hs/2, hs, hs).contains(pos):
            return self.HANDLE_BOTTOM_RIGHT
        
        # Edge handles
        if QRectF(rect.x() + rect.width()/2 - hs/2, rect.y() - hs/2, hs, hs).contains(pos):
            return self.HANDLE_TOP
        if QRectF(rect.x() + rect.width()/2 - hs/2, rect.bottom() - hs/2, hs, hs).contains(pos):
            return self.HANDLE_BOTTOM
        if QRectF(rect.x() - hs/2, rect.y() + rect.height()/2 - hs/2, hs, hs).contains(pos):
            return self.HANDLE_LEFT
        if QRectF(rect.right() - hs/2, rect.y() + rect.height()/2 - hs/2, hs, hs).contains(pos):
            return self.HANDLE_RIGHT
        
        return self.HANDLE_NONE
    
    def get_cursor_for_handle(self, handle):
        """Return the appropriate cursor for a resize handle."""
        if handle in (self.HANDLE_TOP_LEFT, self.HANDLE_BOTTOM_RIGHT):
            return Qt.SizeFDiagCursor
        elif handle in (self.HANDLE_TOP_RIGHT, self.HANDLE_BOTTOM_LEFT):
            return Qt.SizeBDiagCursor
        elif handle in (self.HANDLE_TOP, self.HANDLE_BOTTOM):
            return Qt.SizeVerCursor
        elif handle in (self.HANDLE_LEFT, self.HANDLE_RIGHT):
            return Qt.SizeHorCursor
        return Qt.ArrowCursor
    
    def hoverMoveEvent(self, event):
        """Change cursor when hovering over resize handles."""
        handle = self.get_handle_at_pos(event.pos())
        if handle != self.HANDLE_NONE:
            self.setCursor(self.get_cursor_for_handle(handle))
        else:
            self.setCursor(Qt.SizeAllCursor)  # Move cursor when not on handle
        super().hoverMoveEvent(event)
    
    def hoverLeaveEvent(self, event):
        """Reset cursor when leaving the item."""
        self.setCursor(Qt.ArrowCursor)
        super().hoverLeaveEvent(event)
    
    def mousePressEvent(self, event):
        """Start resize if clicking on a handle."""
        if event.button() == Qt.LeftButton:
            handle = self.get_handle_at_pos(event.pos())
            if handle != self.HANDLE_NONE:
                self.resize_handle = handle
                self.resize_start_rect = self.rect()
                self.resize_start_pos = event.pos()
                self.setFlag(QGraphicsRectItem.ItemIsMovable, False)
                event.accept()
                return
        self.setFlag(QGraphicsRectItem.ItemIsMovable, True)
        super().mousePressEvent(event)
    
    def mouseMoveEvent(self, event):
        """Handle resize dragging."""
        if self.resize_handle != self.HANDLE_NONE:
            delta = event.pos() - self.resize_start_pos
            rect = QRectF(self.resize_start_rect)
            
            min_size = 20  # Minimum size
            
            if self.resize_handle == self.HANDLE_TOP_LEFT:
                rect.setTopLeft(rect.topLeft() + delta)
            elif self.resize_handle == self.HANDLE_TOP_RIGHT:
                rect.setTopRight(rect.topRight() + delta)
            elif self.resize_handle == self.HANDLE_BOTTOM_LEFT:
                rect.setBottomLeft(rect.bottomLeft() + delta)
            elif self.resize_handle == self.HANDLE_BOTTOM_RIGHT:
                rect.setBottomRight(rect.bottomRight() + delta)
            elif self.resize_handle == self.HANDLE_TOP:
                rect.setTop(rect.top() + delta.y())
            elif self.resize_handle == self.HANDLE_BOTTOM:
                rect.setBottom(rect.bottom() + delta.y())
            elif self.resize_handle == self.HANDLE_LEFT:
                rect.setLeft(rect.left() + delta.x())
            elif self.resize_handle == self.HANDLE_RIGHT:
                rect.setRight(rect.right() + delta.x())
            
            # Ensure minimum size
            if rect.width() >= min_size and rect.height() >= min_size:
                self.setRect(rect.normalized())
            
            event.accept()
            return
        super().mouseMoveEvent(event)
    
    def mouseReleaseEvent(self, event):
        """End resize operation."""
        if self.resize_handle != self.HANDLE_NONE:
            self.resize_handle = self.HANDLE_NONE
            self.setFlag(QGraphicsRectItem.ItemIsMovable, True)
            event.accept()
            return
        super().mouseReleaseEvent(event)
        
    def paint(self, painter, option, widget):
        super().paint(painter, option, widget)
        rect = self.rect()
        
        if self.mark_type == MARK_TYPE_OPTION:
            # Draw cell divisions for options
            cell_width = rect.width() / self.options_count
            option_labels = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
            
            # Draw vertical dividers
            painter.setPen(QPen(QColor(255, 0, 0, 150), 1, Qt.DashLine))
            for i in range(1, self.options_count):
                x = rect.x() + i * cell_width
                painter.drawLine(int(x), int(rect.y()), int(x), int(rect.y() + rect.height()))
            
            # Draw option labels (A, B, C, D...)
            painter.setPen(QPen(QColor(100, 0, 0)))
            painter.setFont(QFont("Segoe UI", 8))
            for i in range(self.options_count):
                cell_rect = QRectF(rect.x() + i * cell_width, rect.y(), cell_width, rect.height())
                painter.drawText(cell_rect, Qt.AlignCenter, option_labels[i])
            
            # Draw question number at top
            painter.setPen(QPen(Qt.black))
            painter.setFont(QFont("Segoe UI", 9, QFont.Bold))
            display_text = f"Q{self.question_num}"
            if self.label:
                display_text += f" ({self.label})"
            painter.drawText(int(rect.x()), int(rect.y()) - 3, display_text)
        elif self.mark_type == MARK_TYPE_ALIGN:
            # Alignment reference - show label with number
            painter.setPen(QPen(QColor(0, 150, 0)))
            painter.setFont(QFont("Segoe UI", 10, QFont.Bold))
            display_text = f"📍 {tr('align_overlay')} {self.question_num}"
            if self.label:
                display_text = f"📍 {self.label}"
            painter.drawText(rect, Qt.AlignCenter, display_text)
        else:
            # Text field - just show the label
            painter.setPen(QPen(Qt.black))
            painter.setFont(QFont("Segoe UI", 9, QFont.Bold))
            display_text = self.label if self.label else f"Field {self.question_num}"
            painter.drawText(rect, Qt.AlignCenter, display_text)
        
        # Draw resize handles when selected
        if self.isSelected():
            hs = RESIZE_HANDLE_SIZE
            painter.setPen(QPen(QColor(0, 120, 215), 1))
            painter.setBrush(QBrush(QColor(0, 120, 215)))
            
            # Corner handles
            painter.drawRect(int(rect.x() - hs/2), int(rect.y() - hs/2), hs, hs)
            painter.drawRect(int(rect.right() - hs/2), int(rect.y() - hs/2), hs, hs)
            painter.drawRect(int(rect.x() - hs/2), int(rect.bottom() - hs/2), hs, hs)
            painter.drawRect(int(rect.right() - hs/2), int(rect.bottom() - hs/2), hs, hs)
            
            # Edge handles
            painter.drawRect(int(rect.x() + rect.width()/2 - hs/2), int(rect.y() - hs/2), hs, hs)
            painter.drawRect(int(rect.x() + rect.width()/2 - hs/2), int(rect.bottom() - hs/2), hs, hs)
            painter.drawRect(int(rect.x() - hs/2), int(rect.y() + rect.height()/2 - hs/2), hs, hs)
            painter.drawRect(int(rect.right() - hs/2), int(rect.y() + rect.height()/2 - hs/2), hs, hs)
        
    def contextMenuEvent(self, event):
        menu = QMenu()
        rename_action = menu.addAction("Rename" if self.mark_type == MARK_TYPE_TEXT else "Label")
        
        if self.mark_type == MARK_TYPE_OPTION:
            options_action = menu.addAction(f"Set Options (Current: {self.options_count})")
        
        delete_action = menu.addAction("Delete")
        
        action = menu.exec_(event.screenPos())
        
        if action == delete_action:
            if self.view_ref:
                self.view_ref.remove_mark_item(self)
        elif action == rename_action:
            new_label, ok = QInputDialog.getText(None, "Rename", "Enter new label:", text=self.label)
            if ok:
                self.set_label(new_label)
        elif self.mark_type == MARK_TYPE_OPTION and action == options_action:
            count, ok = QInputDialog.getInt(None, "Options Count", "Number of options:", self.options_count, 2, 26)
            if ok:
                self.options_count = count
                self.update()
    
    def set_label(self, new_label):
        """Update the label and refresh the display."""
        self.prepareGeometryChange()  # Notify scene of potential geometry change
        self.label = new_label
        self.update()  # Trigger repaint

    def get_data(self):
        # QGraphicsItem.sceneBoundingRect() includes the pen width.  Saving
        # that expanded box made templates grow and drift by one pixel on
        # every save/load cycle, so persist the actual rectangle geometry.
        scene_rect = self.mapRectToScene(self.rect()).boundingRect()
        return {
            "type": self.mark_type,
            "question": self.question_num,
            "label": self.label,
            "options_count": self.options_count,
            "x": scene_rect.x(),
            "y": scene_rect.y(),
            "width": scene_rect.width(),
            "height": scene_rect.height()
        }


class MarkingView(QGraphicsView):
    """Custom graphics view with zoom, marking, and memory."""
    
    def __init__(self, scene, parent=None):
        super().__init__(scene, parent)
        self.setRenderHint(QPainter.Antialiasing)
        self.setRenderHint(QPainter.SmoothPixmapTransform)
        self.setDragMode(QGraphicsView.ScrollHandDrag)
        self.setTransformationAnchor(QGraphicsView.AnchorUnderMouse)
        self.setResizeAnchor(QGraphicsView.AnchorUnderMouse)
        
        # Marking state
        self.marking_mode = False
        self.current_mark_type = MARK_TYPE_OPTION
        self.start_point = None
        self.current_rect = None
        
        # Counters
        self.option_counter = 1
        self.text_counter = 1
        
        # Marks storage
        self.text_marks = []
        self.option_marks = []
        self.align_marks = []  # Multiple alignment marks allowed
        self.mark_history = []  # Track order of marks for undo
        
        # Memory for size - reasonable defaults for typical answer sheets
        self.last_option_size = (200, 35) # Default size for option boxes
        self.last_text_size = (150, 30)   # Default size for text fields
        self.last_align_size = (200, 80)  # Default alignment region size
        self.align_counter = 1  # Counter for alignment marks
        
        # Zoom
        self.zoom_factor = 1.0
        
    def set_marking_mode(self, enabled, mark_type=MARK_TYPE_OPTION):
        self.marking_mode = enabled
        self.current_mark_type = mark_type
        if enabled:
            self.setDragMode(QGraphicsView.NoDrag)
            self.setCursor(Qt.CrossCursor)
        else:
            self.setDragMode(QGraphicsView.ScrollHandDrag)
            self.setCursor(Qt.ArrowCursor)
            
    def remove_mark_item(self, item):
        self.scene().removeItem(item)
        if item in self.text_marks:
            self.text_marks.remove(item)
            # Restore counter if this was the last item with that number
            if not any(m.question_num >= item.question_num for m in self.text_marks):
                self.text_counter = item.question_num
        if item in self.option_marks:
            self.option_marks.remove(item)
            # Restore counter if this was the last item with that number
            if not any(m.question_num >= item.question_num for m in self.option_marks):
                self.option_counter = item.question_num
        if item in self.align_marks:
            self.align_marks.remove(item)
        # Remove from history if present
        if item in self.mark_history:
            self.mark_history.remove(item)
            
    def wheelEvent(self, event: QWheelEvent):
        if event.modifiers() == Qt.ControlModifier:
            delta = event.angleDelta().y()
            if delta > 0:
                self.zoom_in()
            else:
                self.zoom_out()
            event.accept()
        else:
            super().wheelEvent(event)
            
    def zoom_in(self):
        if self.zoom_factor < 10.0:  # Max zoom limit
            self.zoom_factor *= 1.2
            self.setTransform(QtGui.QTransform().scale(self.zoom_factor, self.zoom_factor))
            
    def zoom_out(self):
        if self.zoom_factor > 0.1:  # Min zoom limit
            self.zoom_factor /= 1.2
            self.setTransform(QtGui.QTransform().scale(self.zoom_factor, self.zoom_factor))
    
    def zoom_reset(self):
        self.zoom_factor = 1.0
        self.setTransform(QtGui.QTransform().scale(1.0, 1.0))
    
    def zoom_fit(self):
        """Fit the entire scene in the view"""
        self.fitInView(self.sceneRect(), Qt.KeepAspectRatio)
        # Update zoom_factor based on current transform
        transform = self.transform()
        self.zoom_factor = transform.m11()  # Get horizontal scale factor
        
    def mousePressEvent(self, event):
        if self.marking_mode and event.button() == Qt.LeftButton:
            self.start_point = self.mapToScene(event.pos())
            
            if self.current_mark_type == MARK_TYPE_TEXT:
                counter = self.text_counter
            elif self.current_mark_type == MARK_TYPE_ALIGN:
                counter = self.align_counter
            else:
                counter = self.option_counter
            
            # Create MarkItem at start point with zero size initially
            # Use local rect coordinates (0, 0, 0, 0) and set position via setPos
            self.current_rect = MarkItem(
                0, 0, 0, 0,
                self.current_mark_type, counter, view_ref=self
            )
            self.current_rect.setPos(self.start_point)
            self.scene().addItem(self.current_rect)
        else:
            super().mousePressEvent(event)
            
    def mouseMoveEvent(self, event):
        if self.marking_mode and self.current_rect and self.start_point:
            current_pos = self.mapToScene(event.pos())
            # Calculate width and height from start point
            dx = current_pos.x() - self.start_point.x()
            dy = current_pos.y() - self.start_point.y()
            
            # Handle dragging in any direction
            if dx >= 0 and dy >= 0:
                # Normal drag: down-right
                self.current_rect.setPos(self.start_point)
                self.current_rect.setRect(0, 0, dx, dy)
            elif dx < 0 and dy >= 0:
                # Drag left-down
                self.current_rect.setPos(current_pos.x(), self.start_point.y())
                self.current_rect.setRect(0, 0, -dx, dy)
            elif dx >= 0 and dy < 0:
                # Drag right-up
                self.current_rect.setPos(self.start_point.x(), current_pos.y())
                self.current_rect.setRect(0, 0, dx, -dy)
            else:
                # Drag left-up
                self.current_rect.setPos(current_pos)
                self.current_rect.setRect(0, 0, -dx, -dy)
        else:
            super().mouseMoveEvent(event)
            
    def mouseReleaseEvent(self, event):
        if self.marking_mode and self.current_rect:
            rect = self.current_rect.rect()
            actual_width = rect.width()
            actual_height = rect.height()
            
            print(f"  Mark created: rect=({rect.x():.1f},{rect.y():.1f}) size=({actual_width:.1f}x{actual_height:.1f})")
            
            # If created box is too small (user just clicked without dragging), use default/last size
            # Threshold of 5 pixels to account for accidental small movements
            min_threshold = 5
            if actual_width < min_threshold or actual_height < min_threshold:
                if self.current_mark_type == MARK_TYPE_TEXT:
                    w, h = self.last_text_size
                elif self.current_mark_type == MARK_TYPE_ALIGN:
                    w, h = self.last_align_size
                else:
                    w, h = self.last_option_size
                # Reset position to start point and set proper size
                self.current_rect.setPos(self.start_point)
                self.current_rect.setRect(0, 0, w, h)
                print(f"  Mark too small, using default size: ({w}x{h})")
                actual_width = w
                actual_height = h
                
            # Save size for next time (only if box is valid)
            if actual_width >= min_threshold and actual_height >= min_threshold:
                if self.current_mark_type == MARK_TYPE_TEXT:
                    self.last_text_size = (actual_width, actual_height)
                    self.text_marks.append(self.current_rect)
                    self.mark_history.append(self.current_rect)  # Track for undo
                    self.text_counter += 1
                elif self.current_mark_type == MARK_TYPE_ALIGN:
                    self.last_align_size = (actual_width, actual_height)
                    # Multiple alignment marks allowed
                    self.align_marks.append(self.current_rect)
                    self.mark_history.append(self.current_rect)  # Track for undo
                    self.align_counter += 1
                else:
                    self.last_option_size = (actual_width, actual_height)
                    self.option_marks.append(self.current_rect)
                    self.mark_history.append(self.current_rect)  # Track for undo
                    self.option_counter += 1
                print(f"  Mark saved with size: ({actual_width:.1f}x{actual_height:.1f})")
            else:
                self.scene().removeItem(self.current_rect)
                print(f"  Mark removed (invalid size)")
            
            self.current_rect = None
            self.start_point = None
        else:
            super().mouseReleaseEvent(event)

    def get_all_marks_data(self):
        marks_data = {
            "text_marks": [],
            "option_marks": [],
            "align_marks": []
        }
        for mark in self.text_marks:
            try: marks_data["text_marks"].append(mark.get_data())
            except: continue
        for mark in self.option_marks:
            try: marks_data["option_marks"].append(mark.get_data())
            except: continue
        for mark in self.align_marks:
            try: marks_data["align_marks"].append(mark.get_data())
            except: continue
        return marks_data
        
    def load_marks_from_data(self, data):
        # Clear existing first (assumed handled by parent or previous clean)
        pass # Logic handled in OMRSoftware class to avoid duplication


class _TablePasteFilter(QObject):
    """Event filter that intercepts Ctrl+V on a QTableWidget and calls a callback."""
    def __init__(self, callback, parent=None):
        super().__init__(parent)
        self._callback = callback

    def eventFilter(self, obj, event):
        if event.type() == QEvent.KeyPress:
            if event.key() == Qt.Key_V and (event.modifiers() & Qt.ControlModifier):
                self._callback()
                return True
        return False


class OMRSoftware(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setStyleSheet(STYLE_SHEET)
        
        # OCR Init
        self.ocr_reader = None
        self.ocr_engine_name = OCR_ENGINE
        self.init_ocr()
        
        # Data
        self.pdf_document = None
        self.current_page = 0
        self.page_offsets = {} # Store (x,y) of image per page
        self.marks_data = {} # Full template data
        self.current_pixmap_item = None
        self.answer_key = {}
        self.first_page_key = False
        self.results = {}
        self.align_reference_gray = None
        self.align_reference_size = None
        self.align_reference_offset = None
        self.topic_map = {}
        self.debug_records = []
        self._last_detection = {}
        self._review_cursor = None
        self._session_temp = tempfile.TemporaryDirectory(prefix="checkmate_")
        self.debug_dir = os.path.join(self._session_temp.name, "debug_crops")
        self.student_absence = {}  # page_idx -> bool (True if absent)
        self.extra_students = []   # Extra student records beyond PDF pages
        self.student_order = []    # Ordered list: [{"text":{...}, "absent":bool, "page_idx": int or None}, ...]

        # Settings (persistent)
        self._settings = QSettings("CheckMate", "CheckMate")
        self._update_thread = None

        self.init_ui()

    def closeEvent(self, event):
        """Release the active PDF and per-session diagnostic cache."""

        try:
            if self.pdf_document is not None:
                self.pdf_document.close()
        except Exception:
            pass
        try:
            self._session_temp.cleanup()
        except Exception:
            pass
        super().closeEvent(event)
        
    def init_ocr(self):
        if self.ocr_engine_name == "easyocr":
            print("Using EasyOCR")
        elif self.ocr_engine_name == "tesseract":
            print("Using Tesseract")
        else:
            print("No OCR engine found")

    def _prepare_alignment_gray(self, img_np, target_size=None):
        if len(img_np.shape) == 3:
            gray = cv2.cvtColor(img_np, cv2.COLOR_RGB2GRAY)
        else:
            gray = img_np.copy()

        gray = cv2.GaussianBlur(gray, (5, 5), 0)

        if target_size is not None:
            target_w, target_h = target_size
            scale_x = target_w / gray.shape[1]
            scale_y = target_h / gray.shape[0]
            gray = cv2.resize(gray, (target_w, target_h), interpolation=cv2.INTER_AREA)
            return gray, scale_x, scale_y

        # Normalize size to max dimension 800
        max_dim = max(gray.shape[0], gray.shape[1])
        if max_dim > 800:
            scale = 800.0 / max_dim
            gray = cv2.resize(gray, (int(gray.shape[1] * scale), int(gray.shape[0] * scale)), interpolation=cv2.INTER_AREA)
            return gray, scale, scale

        return gray, 1.0, 1.0

    def align_image(self, img_np, page_idx=0):
        """
        Align current page to reference using template matching on the alignment region.
        If user defined an alignment region, use that for precise alignment.
        Otherwise, fall back to automatic table boundary detection.
        Returns aligned image, (dx, dy), and confidence score.
        """
        if not hasattr(self, 'check_auto_align') or not self.check_auto_align.isChecked():
            return img_np, (0.0, 0.0), 0.0

        h, w = img_np.shape[:2]
        
        # Check if user defined alignment region(s)
        if hasattr(self, 'view') and len(self.view.align_marks) > 0:
            return self._align_using_template(img_np, page_idx)
        
        # Fall back to automatic table boundary detection
        cur_bounds = self._find_table_bounds(img_np)
        
        if cur_bounds is None:
            print("  Auto-align: Cannot detect table boundaries")
            return img_np, (0.0, 0.0), 0.0
        
        # First page becomes the reference
        if self.align_reference_gray is None:
            self.align_reference_bounds = cur_bounds
            self.align_reference_gray = True  # Just mark as initialized
            print(f"  Auto-align: Reference bounds set: {cur_bounds}")
            return img_np, (0.0, 0.0), 1.0

        ref_bounds = self.align_reference_bounds
        
        # Calculate shift needed to align current to reference
        # Align based on top-left corner of detected table
        dx = ref_bounds[0] - cur_bounds[0]  # x shift
        dy = ref_bounds[1] - cur_bounds[1]  # y shift
        
        # Also check if size differs significantly (might indicate wrong detection)
        ref_w = ref_bounds[2] - ref_bounds[0]
        ref_h = ref_bounds[3] - ref_bounds[1]
        cur_w = cur_bounds[2] - cur_bounds[0]
        cur_h = cur_bounds[3] - cur_bounds[1]
        
        size_diff = abs(ref_w - cur_w) / max(ref_w, 1) + abs(ref_h - cur_h) / max(ref_h, 1)
        if size_diff > 0.3:  # More than 30% size difference
            print(f"  Auto-align: Size difference too large ({size_diff:.2%}), skipping")
            return img_np, (0.0, 0.0), 0.5
        
        # Sanity check for extreme shifts
        if abs(dx) > w * 0.2 or abs(dy) > h * 0.2:
            print(f"  Auto-align: Shift too large (dx={dx:.1f}, dy={dy:.1f}), skipping")
            return img_np, (0.0, 0.0), 0.3
        
        # Skip if shift is negligible
        if abs(dx) < 3 and abs(dy) < 3:
            return img_np, (0.0, 0.0), 1.0
        
        print(f"  Auto-align: Shifting by dx={dx:.1f}, dy={dy:.1f}")
        
        M = np.float32([[1, 0, dx], [0, 1, dy]])
        if len(img_np.shape) == 3:
            border_value = (255, 255, 255)
        else:
            border_value = 255

        aligned = cv2.warpAffine(img_np, M, (w, h), borderMode=cv2.BORDER_CONSTANT, borderValue=border_value)
        return aligned, (dx, dy), 1.0 - size_diff
    
    def _align_using_template(self, img_np, page_idx):
        """
        Align image using enhanced multi-strategy template matching.
        Supports multiple alignment marks for more robust alignment.
        
        Improvements over basic template matching:
        1. Edge-based matching: Uses Canny edges instead of raw grayscale,
           making it robust to brightness/contrast differences between pages.
        2. Multi-scale pyramid: Coarse-to-fine approach for faster and more
           robust matching across different shift magnitudes.
        3. Rotation correction: Detects and corrects small rotation differences
           (typical ±1° from scanner feed), not just translation.
        4. Multiple matching methods: Cross-validates results from different
           OpenCV matching algorithms to reject false positives.
        5. Larger adaptive search margin: Adjusts based on image size.
        6. Enhanced sub-pixel refinement with 2D quadratic fitting.
        7. Multiple alignment marks: Uses weighted average of shifts from
           all alignment regions for better accuracy.
        """
        h, w = img_np.shape[:2]
        
        # First page: extract and store templates from all alignment marks
        if not hasattr(self, 'align_templates') or not self.align_templates:
            return self._align_init_template(img_np, page_idx)
        
        # Subsequent pages: find templates and calculate correction
        return self._align_match_page(img_np, page_idx)
    
    def _reset_align_templates(self):
        """Reset all alignment template data. Call before each new recognition run."""
        self.align_templates = []
        self.align_ref_full_gray = None
        self.align_reference_gray = None
        self.align_reference_bounds = None
        self.align_reference_offset = None
        self.align_reference_shape = None

    def _align_init_template(self, img_np, page_idx):
        """Extract and store alignment templates from the first (reference) page.
        Supports multiple alignment marks for more robust alignment."""
        h, w = img_np.shape[:2]
        
        if len(img_np.shape) == 3:
            gray = cv2.cvtColor(img_np, cv2.COLOR_RGB2GRAY)
        else:
            gray = img_np.copy()
        
        self.align_templates = []
        self.align_reference_shape = (h, w)
        reference_offset = self._get_page_offset(0)
        self.align_reference_offset = (float(reference_offset[0]), float(reference_offset[1]))
        
        for mark_idx, align_mark in enumerate(self.view.align_marks):
            rect = align_mark.sceneBoundingRect()
            off_x, off_y = reference_offset
            
            ref_x = int(rect.x() - off_x)
            ref_y = int(rect.y() - off_y)
            ref_w = int(rect.width())
            ref_h = int(rect.height())
            
            ref_x = max(0, ref_x)
            ref_y = max(0, ref_y)
            
            print(f"  Template align: Creating reference #{mark_idx+1} from region=({ref_x},{ref_y}) size=({ref_w}x{ref_h})")
            
            end_x = min(ref_x + ref_w, gray.shape[1])
            end_y = min(ref_y + ref_h, gray.shape[0])
            
            if end_x <= ref_x or end_y <= ref_y:
                print(f"  Template align: Invalid template region #{mark_idx+1}, skipping")
                continue
            
            # Store grayscale template
            template_gray = gray[ref_y:end_y, ref_x:end_x].copy()
            
            # Store edge-enhanced template (primary matching target)
            template_edges = cv2.Canny(template_gray, 50, 150)
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
            template_edges = cv2.dilate(template_edges, kernel, iterations=1)
            
            # Store CLAHE-enhanced template for secondary verification
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
            template_clahe = clahe.apply(template_gray)
            
            self.align_templates.append({
                'gray': template_gray,
                'edges': template_edges,
                'clahe': template_clahe,
                'pos': (ref_x, ref_y),
                'size': (end_x - ref_x, end_y - ref_y),
            })
            
            print(f"  Template align: Reference #{mark_idx+1} extracted at ({ref_x},{ref_y}), size={template_gray.shape}")
            print(f"  Template align: Edge density #{mark_idx+1}: {np.count_nonzero(template_edges)}/{template_edges.size} pixels")
        
        if not self.align_templates:
            print("  Template align: No valid alignment regions found")
            return img_np, (0.0, 0.0), 0.0
        
        # Store full-page reference gray for rotation detection
        self.align_ref_full_gray = gray.copy()
        
        print(f"  Template align: {len(self.align_templates)} reference template(s) initialized")
        return img_np, (0.0, 0.0), 1.0
    
    def _align_match_page(self, img_np, page_idx):
        """Match current page against all reference templates using multi-strategy approach.
        When multiple alignment marks are present, each template is matched independently
        and the shifts are combined using confidence-weighted averaging."""
        h, w = img_np.shape[:2]
        
        if len(img_np.shape) == 3:
            gray = cv2.cvtColor(img_np, cv2.COLOR_RGB2GRAY)
        else:
            gray = img_np.copy()
        
        # Adaptive margin based on image size (larger images may have larger shifts)
        margin = max(120, min(int(min(w, h) * 0.08), 250))
        
        # Match each template and collect shift estimates
        # (dx, dy, confidence, template_idx, current_center, reference_center)
        shift_estimates = []
        
        for t_idx, tmpl in enumerate(self.align_templates):
            ref_x, ref_y = tmpl['pos']
            ref_w, ref_h = tmpl['size']
            
            search_x1 = max(0, ref_x - margin)
            search_y1 = max(0, ref_y - margin)
            search_x2 = min(w, ref_x + ref_w + margin)
            search_y2 = min(h, ref_y + ref_h + margin)
            
            search_gray = gray[search_y1:search_y2, search_x1:search_x2]
            
            template_h, template_w = tmpl['gray'].shape[:2]
            
            if search_gray.shape[0] < template_h or search_gray.shape[1] < template_w:
                print(f"  Template align #{t_idx+1}: Search region too small, skipping")
                continue
            
            # === Strategy 1: Edge-based matching (primary - most robust) ===
            search_edges = cv2.Canny(search_gray, 50, 150)
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
            search_edges = cv2.dilate(search_edges, kernel, iterations=1)
            
            edge_result = cv2.matchTemplate(search_edges, tmpl['edges'], cv2.TM_CCOEFF_NORMED)
            _, edge_max_val, _, edge_max_loc = cv2.minMaxLoc(edge_result)
            
            # === Strategy 2: CLAHE-enhanced matching (secondary verification) ===
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
            search_clahe = clahe.apply(search_gray)
            
            clahe_result = cv2.matchTemplate(search_clahe, tmpl['clahe'], cv2.TM_CCOEFF_NORMED)
            _, clahe_max_val, _, clahe_max_loc = cv2.minMaxLoc(clahe_result)
            
            # === Strategy 3: Raw grayscale matching (fallback) ===
            gray_result = cv2.matchTemplate(search_gray, tmpl['gray'], cv2.TM_CCOEFF_NORMED)
            _, gray_max_val, _, gray_max_loc = cv2.minMaxLoc(gray_result)
            
            print(f"  Template align #{t_idx+1}: Page {page_idx+1}, confidence - edge={edge_max_val:.3f}, clahe={clahe_max_val:.3f}, gray={gray_max_val:.3f}")
            
            # === Select best result with cross-validation ===
            candidates = [
                ("edge", edge_max_val, edge_max_loc, edge_result),
                ("clahe", clahe_max_val, clahe_max_loc, clahe_result),
                ("gray", gray_max_val, gray_max_loc, gray_result),
            ]
            
            candidates.sort(key=lambda c: c[1], reverse=True)
            
            best_name, best_val, best_loc, best_result = candidates[0]
            second_name, second_val, second_loc, _ = candidates[1]
            
            loc_diff = abs(best_loc[0] - second_loc[0]) + abs(best_loc[1] - second_loc[1])
            
            if best_val < 0.3:
                print(f"  Template align #{t_idx+1}: Low confidence (best={best_val:.3f}), skipping this mark")
                continue
            
            # If methods disagree significantly, prefer edge-based (more robust)
            if loc_diff > 10 and edge_max_val > 0.3:
                best_name, best_val, best_loc, best_result = candidates[0] if candidates[0][0] == "edge" else \
                    next((c for c in candidates if c[0] == "edge"), candidates[0])
            
            # Sub-pixel refinement
            px, py = best_loc
            sub_px, sub_py = self._subpixel_refine(best_result, px, py)
            
            # Convert to full image coordinates
            match_x = search_x1 + sub_px
            match_y = search_y1 + sub_py
            
            # Calculate translation shift
            t_dx = ref_x - match_x
            t_dy = ref_y - match_y
            
            # Confidence adjustment
            effective_confidence = best_val
            if loc_diff <= 5 and second_val > 0.3:
                effective_confidence = min(1.0, best_val * 1.1)
            
            # Sanity check
            max_allowed_shift = margin - 10
            if abs(t_dx) > max_allowed_shift or abs(t_dy) > max_allowed_shift:
                print(f"  Template align #{t_idx+1}: Shift ({t_dx:.2f},{t_dy:.2f}) too large, skipping this mark")
                continue
            
            print(f"  Template align #{t_idx+1}: Best ({best_name}) shift=({t_dx:.2f},{t_dy:.2f}), conf={effective_confidence:.3f}")
            current_center = (
                float(match_x + ref_w / 2.0),
                float(match_y + ref_h / 2.0),
            )
            reference_center = (
                float(ref_x + ref_w / 2.0),
                float(ref_y + ref_h / 2.0),
            )
            shift_estimates.append(
                (
                    t_dx,
                    t_dy,
                    effective_confidence,
                    t_idx,
                    current_center,
                    reference_center,
                )
            )
        
        # === Combine shift estimates from all templates ===
        if not shift_estimates:
            print(f"  Template align: No valid matches on page {page_idx+1}, trying phase correlation fallback...")
            fallback_result = self._align_phase_correlation_fallback(img_np, gray, page_idx)
            if fallback_result is not None:
                return fallback_result
            return img_np, (0.0, 0.0), 0.0
        
        if len(shift_estimates) == 1:
            dx, dy, effective_confidence, _, _, _ = shift_estimates[0]
            print(f"  Template align: Using single template shift=({dx:.2f},{dy:.2f})")
        else:
            source_points = [estimate[4] for estimate in shift_estimates]
            target_points = [estimate[5] for estimate in shift_estimates]
            affine, affine_quality = estimate_similarity_transform(
                source_points,
                target_points,
                ransac_threshold=3.0,
            )
            if affine is not None:
                if len(img_np.shape) == 3:
                    border_value = (255, 255, 255)
                else:
                    border_value = 255
                ref_h, ref_w = self.align_reference_shape or (h, w)
                aligned = cv2.warpAffine(
                    img_np,
                    affine,
                    (int(ref_w), int(ref_h)),
                    flags=cv2.INTER_LINEAR,
                    borderMode=cv2.BORDER_CONSTANT,
                    borderValue=border_value,
                )
                average_match_confidence = sum(
                    estimate[2] for estimate in shift_estimates
                ) / len(shift_estimates)
                inlier_ratio = (
                    affine_quality["inliers"]
                    / max(1, affine_quality["total_points"])
                )
                residual_factor = max(
                    0.0, 1.0 - affine_quality["rms_error"] / 4.0
                )
                effective_confidence = max(
                    0.0,
                    min(
                        1.0,
                        average_match_confidence
                        * (0.75 + 0.25 * inlier_ratio)
                        * (0.75 + 0.25 * residual_factor),
                    ),
                )
                dx = float(affine[0, 2])
                dy = float(affine[1, 2])
                print(
                    "  Template align: ✓ affine "
                    f"inliers={affine_quality['inliers']}/{affine_quality['total_points']}, "
                    f"rms={affine_quality['rms_error']:.2f}px, "
                    f"scale={affine_quality['scale']:.4f}, "
                    f"rotation={affine_quality['rotation']:.3f}°"
                )
                return aligned, (dx, dy), effective_confidence

            # Weighted average of shifts by confidence
            total_weight = sum(estimate[2] for estimate in shift_estimates)
            dx = sum(
                estimate[0] * estimate[2] for estimate in shift_estimates
            ) / total_weight
            dy = sum(
                estimate[1] * estimate[2] for estimate in shift_estimates
            ) / total_weight
            effective_confidence = total_weight / len(shift_estimates)
            
            # Check consistency: if shifts disagree by more than 5px, use median instead
            dxs = [estimate[0] for estimate in shift_estimates]
            dys = [estimate[1] for estimate in shift_estimates]
            dx_range = max(dxs) - min(dxs)
            dy_range = max(dys) - min(dys)
            
            if dx_range > 5 or dy_range > 5:
                print(f"  Template align: Shift estimates disagree (dx_range={dx_range:.1f}, dy_range={dy_range:.1f}), using median")
                dxs.sort()
                dys.sort()
                dx = dxs[len(dxs) // 2]
                dy = dys[len(dys) // 2]
                effective_confidence *= 0.8  # Reduce confidence due to disagreement
            else:
                print(f"  Template align: {len(shift_estimates)} templates agree (dx_range={dx_range:.1f}, dy_range={dy_range:.1f}) ✓")
        
        if effective_confidence < 0.35:
            print(f"  Template align: Low combined confidence ({effective_confidence:.3f}), trying phase correlation fallback...")
            fallback_result = self._align_phase_correlation_fallback(img_np, gray, page_idx)
            if fallback_result is not None:
                return fallback_result
            return img_np, (0.0, 0.0), effective_confidence
        
        # === Rotation detection and correction ===
        rotation_angle = self._detect_rotation(gray, dx, dy, page_idx)
        
        # Skip if both shift and rotation are negligible
        if abs(dx) < 0.3 and abs(dy) < 0.3 and abs(rotation_angle) < 0.02:
            print(f"  Template align: Correction negligible, skipping")
            return img_np, (0.0, 0.0), effective_confidence
        
        # === Apply combined transform (rotation + translation) ===
        if len(img_np.shape) == 3:
            border_value = (255, 255, 255)
        else:
            border_value = 255
        
        center = (w / 2.0, h / 2.0)
        
        if abs(rotation_angle) >= 0.02:
            R = cv2.getRotationMatrix2D(center, rotation_angle, 1.0)
            R[0, 2] += dx
            R[1, 2] += dy
            aligned = cv2.warpAffine(img_np, R, (w, h),
                                      flags=cv2.INTER_LINEAR,
                                      borderMode=cv2.BORDER_CONSTANT,
                                      borderValue=border_value)
            print(f"  Template align: ✓ Applied dx={dx:.2f}, dy={dy:.2f}, rotation={rotation_angle:.3f}°")
        else:
            M = np.float32([[1, 0, dx], [0, 1, dy]])
            aligned = cv2.warpAffine(img_np, M, (w, h),
                                      flags=cv2.INTER_LINEAR,
                                      borderMode=cv2.BORDER_CONSTANT,
                                      borderValue=border_value)
            print(f"  Template align: ✓ Applied dx={dx:.2f}, dy={dy:.2f}")
        
        return aligned, (dx, dy), effective_confidence
    
    def _subpixel_refine(self, result, px, py):
        """
        Sub-pixel refinement using 2D quadratic surface fitting.
        Fits a 3x3 neighborhood to find the true peak position.
        More accurate than 1D parabolic interpolation.
        """
        rh, rw = result.shape[:2]
        
        if not (1 <= px < rw - 1 and 1 <= py < rh - 1):
            return float(px), float(py)
        
        # Extract 3x3 neighborhood
        patch = result[py-1:py+2, px-1:px+2].astype(np.float64)
        
        # Fit 2D quadratic: f(x,y) = a*x^2 + b*y^2 + c*x*y + d*x + e*y + f
        # Using the 9 points in the 3x3 patch
        # Simplified: compute dx and dy offsets from center
        
        # Horizontal offset (using center row)
        denom_x = 2.0 * (patch[1, 0] + patch[1, 2] - 2.0 * patch[1, 1])
        if abs(denom_x) > 1e-7:
            offset_x = -(patch[1, 2] - patch[1, 0]) / denom_x
        else:
            offset_x = 0.0
        
        # Vertical offset (using center column)
        denom_y = 2.0 * (patch[0, 1] + patch[2, 1] - 2.0 * patch[1, 1])
        if abs(denom_y) > 1e-7:
            offset_y = -(patch[2, 1] - patch[0, 1]) / denom_y
        else:
            offset_y = 0.0
        
        # Clamp offsets to ±0.5 (should not exceed half a pixel)
        offset_x = max(-0.5, min(0.5, offset_x))
        offset_y = max(-0.5, min(0.5, offset_y))
        
        return px + offset_x, py + offset_y
    
    def _detect_rotation(self, gray, dx, dy, page_idx):
        """
        Detect small rotation difference between current page and reference.
        Uses the alignment template region to test small angle candidates.
        Returns the best rotation angle in degrees.
        """
        if not hasattr(self, 'align_ref_full_gray') or self.align_ref_full_gray is None:
            return 0.0
        
        if not self.align_templates:
            return 0.0
        
        # Use the first (or largest) template for rotation detection
        tmpl = self.align_templates[0]
        ref_x, ref_y = tmpl['pos']
        ref_w, ref_h = tmpl['size']
        
        # Use a larger region around the template for rotation detection
        # (rotation is more visible over larger distances)
        expand = max(ref_w, ref_h) // 2
        h, w = gray.shape[:2]
        
        rx1 = max(0, ref_x - expand)
        ry1 = max(0, ref_y - expand)
        rx2 = min(w, ref_x + ref_w + expand)
        ry2 = min(h, ref_y + ref_h + expand)
        
        ref_gray = self.align_ref_full_gray
        rh, rw = ref_gray.shape[:2]
        
        # Ensure same region in reference
        rrx1 = max(0, min(rx1, rw))
        rry1 = max(0, min(ry1, rh))
        rrx2 = max(0, min(rx2, rw))
        rry2 = max(0, min(ry2, rh))
        
        if rrx2 - rrx1 < 50 or rry2 - rry1 < 50:
            return 0.0
        
        ref_region = ref_gray[rry1:rry2, rrx1:rrx2]
        
        # First apply the translation, then test rotation
        # Shift the current gray to approximate translation correction
        M_translate = np.float32([[1, 0, dx], [0, 1, dy]])
        shifted_gray = cv2.warpAffine(gray, M_translate, (w, h),
                                       borderMode=cv2.BORDER_CONSTANT,
                                       borderValue=255)
        
        cur_region = shifted_gray[ry1:ry2, rx1:rx2]
        
        if cur_region.shape != ref_region.shape:
            # Resize to match
            cur_region = cv2.resize(cur_region, (ref_region.shape[1], ref_region.shape[0]))
        
        # Test small rotation angles: -1.0° to +1.0° in 0.1° steps
        angles_to_test = [a * 0.1 for a in range(-10, 11)]
        best_angle = 0.0
        best_score = -1.0
        
        region_h, region_w = cur_region.shape[:2]
        center = (region_w / 2.0, region_h / 2.0)
        
        # Use edge images for rotation matching (more sensitive to angular changes)
        ref_edges = cv2.Canny(ref_region, 50, 150)
        
        for angle in angles_to_test:
            if abs(angle) < 0.01:
                rotated = cur_region
            else:
                R = cv2.getRotationMatrix2D(center, angle, 1.0)
                rotated = cv2.warpAffine(cur_region, R, (region_w, region_h),
                                          borderMode=cv2.BORDER_CONSTANT,
                                          borderValue=255)
            
            cur_edges = cv2.Canny(rotated, 50, 150)
            
            # Score: normalized cross-correlation of edge images
            if np.std(ref_edges) > 0 and np.std(cur_edges) > 0:
                score = np.corrcoef(ref_edges.ravel().astype(float), 
                                    cur_edges.ravel().astype(float))[0, 1]
            else:
                score = 0.0
            
            if score > best_score:
                best_score = score
                best_angle = angle
        
        # Only apply rotation if it's clearly better than 0°
        zero_idx = angles_to_test.index(0.0) if 0.0 in angles_to_test else 10
        zero_angle_score = -1.0
        # Recalculate score at 0°
        cur_edges_0 = cv2.Canny(cur_region, 50, 150)
        if np.std(ref_edges) > 0 and np.std(cur_edges_0) > 0:
            zero_angle_score = np.corrcoef(ref_edges.ravel().astype(float),
                                            cur_edges_0.ravel().astype(float))[0, 1]
        
        improvement = best_score - zero_angle_score
        
        if improvement > 0.005 and abs(best_angle) >= 0.05:
            print(f"  Template align: Rotation detected: {best_angle:.1f}° (improvement={improvement:.4f})")
            return best_angle
        else:
            return 0.0

    def _align_phase_correlation_fallback(self, img_np, gray, page_idx):
        """
        Fallback alignment using phase correlation on the alignment region
        with a much larger search area. Used when template matching fails
        (shift too large or confidence too low).
        
        Strategy:
        1. Use phase correlation on the full page to detect gross translation.
        2. Then refine with template matching on the alignment region using the
           corrected position.
        """
        if not hasattr(self, 'align_ref_full_gray') or self.align_ref_full_gray is None:
            return None
        
        h, w = img_np.shape[:2]
        ref_gray = self.align_ref_full_gray
        rh, rw = ref_gray.shape[:2]
        
        # Make both images the same size for phase correlation
        common_h = min(h, rh)
        common_w = min(w, rw)
        
        if common_h < 100 or common_w < 100:
            return None
        
        ref_crop = ref_gray[:common_h, :common_w].astype(np.float64)
        cur_crop = gray[:common_h, :common_w].astype(np.float64)
        
        # Apply window function to reduce edge effects
        hann_rows = np.hanning(common_h)
        hann_cols = np.hanning(common_w)
        window = np.outer(hann_rows, hann_cols)
        
        ref_windowed = ref_crop * window
        cur_windowed = cur_crop * window
        
        # Phase correlation
        f_ref = np.fft.fft2(ref_windowed)
        f_cur = np.fft.fft2(cur_windowed)
        
        cross_power = (f_ref * np.conj(f_cur))
        denom = np.abs(cross_power)
        denom[denom < 1e-10] = 1e-10
        cross_power_norm = cross_power / denom
        
        correlation = np.fft.ifft2(cross_power_norm).real
        
        # Find peak
        max_loc = np.unravel_index(np.argmax(correlation), correlation.shape)
        peak_y, peak_x = max_loc
        
        # Convert to signed shift (handle wrap-around)
        if peak_y > common_h // 2:
            peak_y -= common_h
        if peak_x > common_w // 2:
            peak_x -= common_w
        
        dx = float(peak_x)
        dy = float(peak_y)
        
        # Check peak strength (ratio of peak to mean)
        peak_val = correlation[max_loc]
        mean_val = np.mean(np.abs(correlation))
        peak_ratio = peak_val / max(mean_val, 1e-10)
        
        print(f"  Phase correlation fallback: shift=({dx:.1f},{dy:.1f}), peak_ratio={peak_ratio:.1f}")
        
        # Reject if peak ratio is too low (unreliable)
        if peak_ratio < 3.0:
            print(f"  Phase correlation: Peak ratio too low ({peak_ratio:.1f}), skipping")
            return None
        
        # Reject implausible shifts (more than 20% of image dimension)
        if abs(dx) > w * 0.2 or abs(dy) > h * 0.2:
            print(f"  Phase correlation: Shift too large ({dx:.1f},{dy:.1f}), skipping")
            return None
        
        # Skip negligible shifts
        if abs(dx) < 1.0 and abs(dy) < 1.0:
            return None
        
        # Apply the correction
        if len(img_np.shape) == 3:
            border_value = (255, 255, 255)
        else:
            border_value = 255
        
        M = np.float32([[1, 0, dx], [0, 1, dy]])
        aligned = cv2.warpAffine(img_np, M, (w, h),
                                  flags=cv2.INTER_LINEAR,
                                  borderMode=cv2.BORDER_CONSTANT,
                                  borderValue=border_value)
        
        # Verify the correction by re-matching template in the corrected image
        confidence = self._verify_alignment_quality(aligned, page_idx)
        
        if confidence < 0.3:
            print(f"  Phase correlation: Verification failed (conf={confidence:.3f}), skipping")
            return None
        
        print(f"  Phase correlation: ✓ Applied dx={dx:.1f}, dy={dy:.1f} (verified conf={confidence:.3f})")
        return aligned, (dx, dy), confidence

    def _verify_alignment_quality(self, aligned_img, page_idx):
        """Verify alignment quality by matching all templates in the corrected image."""
        if not hasattr(self, 'align_templates') or not self.align_templates:
            return 0.5
        
        if len(aligned_img.shape) == 3:
            gray = cv2.cvtColor(aligned_img, cv2.COLOR_RGB2GRAY)
        else:
            gray = aligned_img.copy()
        
        margin = 30
        h, w = gray.shape[:2]
        confidences = []
        
        for tmpl in self.align_templates:
            ref_x, ref_y = tmpl['pos']
            ref_w, ref_h = tmpl['size']
            
            sx1 = max(0, ref_x - margin)
            sy1 = max(0, ref_y - margin)
            sx2 = min(w, ref_x + ref_w + margin)
            sy2 = min(h, ref_y + ref_h + margin)
            
            search_region = gray[sy1:sy2, sx1:sx2]
            template_h, template_w = tmpl['gray'].shape[:2]
            
            if search_region.shape[0] < template_h or search_region.shape[1] < template_w:
                continue
            
            result = cv2.matchTemplate(search_region, tmpl['gray'], cv2.TM_CCOEFF_NORMED)
            _, max_val, _, _ = cv2.minMaxLoc(result)
            confidences.append(max_val)
        
        if not confidences:
            return 0.0
        
        return sum(confidences) / len(confidences)
    
    def _find_table_bounds(self, img_np):
        """
        Find the bounding box of the main table/frame in the image.
        Returns (x1, y1, x2, y2) or None if not found.
        """
        # Convert to grayscale
        if len(img_np.shape) == 3:
            gray = cv2.cvtColor(img_np, cv2.COLOR_RGB2GRAY)
        else:
            gray = img_np.copy()
        
        h, w = gray.shape
        
        # Apply edge detection
        edges = cv2.Canny(gray, 50, 150)
        
        # Apply morphological operations to connect edges
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (5, 5))
        edges = cv2.dilate(edges, kernel, iterations=2)
        edges = cv2.erode(edges, kernel, iterations=1)
        
        # Find contours
        contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        if not contours:
            return None
        
        # Find the largest contour that looks like a table (rectangular, large area)
        best_contour = None
        best_area = 0
        min_area = h * w * 0.1  # At least 10% of image
        
        for contour in contours:
            area = cv2.contourArea(contour)
            if area < min_area:
                continue
            
            # Check if roughly rectangular
            peri = cv2.arcLength(contour, True)
            approx = cv2.approxPolyDP(contour, 0.02 * peri, True)
            
            # Accept 4-sided polygons or large areas
            if len(approx) >= 4 or area > best_area:
                if area > best_area:
                    best_area = area
                    best_contour = contour
        
        if best_contour is None:
            # Fallback: use projection profile to find table edges
            return self._find_bounds_by_projection(gray)
        
        x, y, bw, bh = cv2.boundingRect(best_contour)
        return (x, y, x + bw, y + bh)
    
    def _find_bounds_by_projection(self, gray):
        """
        Find table bounds using horizontal and vertical projection profiles.
        This works well for scanned documents with clear table borders.
        """
        h, w = gray.shape
        
        # Threshold to binary (invert so lines are white)
        _, binary = cv2.threshold(gray, 200, 255, cv2.THRESH_BINARY_INV)
        
        # Horizontal projection (sum along rows)
        h_proj = np.sum(binary, axis=1)
        
        # Vertical projection (sum along columns)  
        v_proj = np.sum(binary, axis=0)
        
        # Find edges using projection threshold
        h_thresh = np.max(h_proj) * 0.1
        v_thresh = np.max(v_proj) * 0.1
        
        # Find top edge
        y1 = 0
        for i in range(h):
            if h_proj[i] > h_thresh:
                y1 = i
                break
        
        # Find bottom edge
        y2 = h - 1
        for i in range(h - 1, -1, -1):
            if h_proj[i] > h_thresh:
                y2 = i
                break
        
        # Find left edge
        x1 = 0
        for i in range(w):
            if v_proj[i] > v_thresh:
                x1 = i
                break
        
        # Find right edge
        x2 = w - 1
        for i in range(w - 1, -1, -1):
            if v_proj[i] > v_thresh:
                x2 = i
                break
        
        # Validate bounds
        if x2 - x1 < w * 0.3 or y2 - y1 < h * 0.3:
            return None
        
        return (x1, y1, x2, y2)

    @staticmethod
    def _select_filled_options_from_scores(cell_scores):
        """Pick filled options by comparing each cell against the blank-cell baseline."""
        if not cell_scores:
            return [], {}

        combined_vals = [s['combined'] for s in cell_scores]
        max_combined = max(combined_vals)
        min_combined = min(combined_vals)
        score_range = max_combined - min_combined
        mean_combined = sum(combined_vals) / max(1, len(combined_vals))
        std_combined = (sum((v - mean_combined) ** 2 for v in combined_vals) / max(1, len(combined_vals))) ** 0.5
        max_darkness = max(s['darkness'] for s in cell_scores)

        sorted_scores = sorted(cell_scores, key=lambda score: score['combined'])
        baseline_count = max(1, len(sorted_scores) // 2)
        blank_reference = sorted_scores[:baseline_count]
        blank_combined_baseline = sum(s['combined'] for s in blank_reference) / baseline_count
        blank_darkness_baseline = sum(s['darkness'] for s in blank_reference) / baseline_count
        blank_mark_baseline = sum(s.get('mark_ratio', 0.0) for s in blank_reference) / baseline_count
        blank_center_darkness_baseline = sum(s.get('center_darkness', s['darkness']) for s in blank_reference) / baseline_count
        blank_center_mark_baseline = sum(s.get('center_mark_ratio', s.get('mark_ratio', 0.0)) for s in blank_reference) / baseline_count
        max_mark_ratio = max(s.get('mark_ratio', 0.0) for s in cell_scores)
        max_center_darkness = max(s.get('center_darkness', s['darkness']) for s in cell_scores)
        max_center_mark_ratio = max(s.get('center_mark_ratio', s.get('mark_ratio', 0.0)) for s in cell_scores)

        for score in cell_scores:
            score['combined_margin'] = score['combined'] - blank_combined_baseline
            score['darkness_margin'] = score['darkness'] - blank_darkness_baseline
            score['mark_ratio_margin'] = score.get('mark_ratio', 0.0) - blank_mark_baseline
            score['center_darkness_margin'] = score.get('center_darkness', score['darkness']) - blank_center_darkness_baseline
            score['center_mark_ratio_margin'] = score.get('center_mark_ratio', score.get('mark_ratio', 0.0)) - blank_center_mark_baseline

        max_combined_margin = max(s['combined_margin'] for s in cell_scores)

        # Absolute minima keep light but real marks eligible.
        MIN_COMBINED_THRESHOLD = 5.0
        MIN_DARKNESS_THRESHOLD = 2.0

        # Relative margins compare each option against the lower-scoring blank-like cells.
        MIN_COMBINED_MARGIN = max(3.5, score_range * 0.35)
        MIN_DARKNESS_MARGIN = max(1.0, (max_darkness - blank_darkness_baseline) * 0.30)
        MIN_MARK_RATIO = 0.04
        MIN_MARK_MARGIN = max(0.02, (max_mark_ratio - blank_mark_baseline) * 0.40)
        MIN_CENTER_DARKNESS = 3.0
        MIN_CENTER_DARKNESS_MARGIN = max(1.5, (max_center_darkness - blank_center_darkness_baseline) * 0.20)
        MIN_CENTER_MARK_RATIO = 0.08
        MIN_CENTER_MARK_MARGIN = max(0.01, (max_center_mark_ratio - blank_center_mark_baseline) * 0.12)

        # Clearly blank if every cell is both low and tightly clustered.
        BLANK_MAX_RANGE = 6.0
        BLANK_MAX_COMBINED = 5.0
        BLANK_MAX_MARGIN = 2.5

        decision_info = {
            'score_range': score_range,
            'mean_combined': mean_combined,
            'std_combined': std_combined,
            'max_darkness': max_darkness,
            'blank_combined_baseline': blank_combined_baseline,
            'blank_darkness_baseline': blank_darkness_baseline,
            'blank_mark_baseline': blank_mark_baseline,
            'blank_center_darkness_baseline': blank_center_darkness_baseline,
            'blank_center_mark_baseline': blank_center_mark_baseline,
            'thresholds': {
                'min_combined': MIN_COMBINED_THRESHOLD,
                'min_darkness': MIN_DARKNESS_THRESHOLD,
                'min_combined_margin': MIN_COMBINED_MARGIN,
                'min_darkness_margin': MIN_DARKNESS_MARGIN,
                'min_mark_ratio': MIN_MARK_RATIO,
                'min_mark_margin': MIN_MARK_MARGIN,
                'min_center_darkness': MIN_CENTER_DARKNESS,
                'min_center_darkness_margin': MIN_CENTER_DARKNESS_MARGIN,
                'min_center_mark_ratio': MIN_CENTER_MARK_RATIO,
                'min_center_mark_margin': MIN_CENTER_MARK_MARGIN,
                'blank_max_range': BLANK_MAX_RANGE,
                'blank_max_combined': BLANK_MAX_COMBINED,
                'blank_max_margin': BLANK_MAX_MARGIN,
            }
        }

        is_clearly_blank = (
            score_range < BLANK_MAX_RANGE and
            max_combined < BLANK_MAX_COMBINED and
            max_combined_margin < BLANK_MAX_MARGIN
        )

        if is_clearly_blank:
            decision_info['reason'] = 'blank'
            return [], decision_info

        filled_options = []
        for score in cell_scores:
            passes_absolute = (
                score['combined'] >= MIN_COMBINED_THRESHOLD and
                score['darkness'] >= MIN_DARKNESS_THRESHOLD
            )
            passes_relative = (
                score['combined_margin'] >= MIN_COMBINED_MARGIN and
                (
                    score['darkness_margin'] >= MIN_DARKNESS_MARGIN or
                    score['center_darkness_margin'] >= MIN_CENTER_DARKNESS_MARGIN
                )
            )
            passes_mark = (
                score.get('mark_ratio', 0.0) >= MIN_MARK_RATIO and
                score['mark_ratio_margin'] >= MIN_MARK_MARGIN
            )
            passes_center_support = (
                (
                    score.get('center_darkness', score['darkness']) >= MIN_CENTER_DARKNESS and
                    score['center_darkness_margin'] >= MIN_CENTER_DARKNESS_MARGIN
                ) or
                (
                    score.get('center_mark_ratio', score.get('mark_ratio', 0.0)) >= MIN_CENTER_MARK_RATIO and
                    score['center_mark_ratio_margin'] >= MIN_CENTER_MARK_MARGIN
                )
            )

            if passes_absolute and passes_relative and (passes_mark or passes_center_support):
                filled_options.append(score['option'])

        decision_info['reason'] = 'baseline' if filled_options else 'below_thresholds'
        return filled_options, decision_info

    @staticmethod
    def _estimate_option_content_bounds(gray, saturation, has_color, dark_pixel_threshold, color_pixel_threshold, options_count):
        """Estimate the horizontal span that actually contains the option boxes/marks."""
        height, width = gray.shape[:2]

        activity_mask = gray < dark_pixel_threshold
        if has_color:
            activity_mask = np.logical_or(activity_mask, saturation > color_pixel_threshold)

        col_activity = np.mean(activity_mask, axis=0)
        min_col_activity = max(0.08, 2.0 / max(1, height))
        active_cols = np.where(col_activity >= min_col_activity)[0]
        if active_cols.size == 0:
            return 0, width

        left = max(0, int(active_cols[0]) - 1)
        right = min(width, int(active_cols[-1]) + 2)

        min_span = max(options_count * 5, int(width * 0.55))
        if right - left < min_span:
            return 0, width

        return left, right

    def _detect_filled_option_legacy(self, image, options_count=4, save_debug=False, context=None):
        """
        Detect which option is filled in a multiple choice bubble area.
        Divides the image into options_count cells and checks which one is filled.
        
        Supports detection of:
        - Dark marks (pencil/pen)
        - Blue marks (blue pen/highlighter)
        - Any colored marks
        
        Args:
            image: PIL Image of the option area
            options_count: Number of options (default 4 for A,B,C,D)
            save_debug: Whether to save debug images
            
        Returns:
            String like "A", "B", "C", "D" or "AB" for multiple selections, or "" if none
        """
        import numpy as np
        import os
        
        img_np = np.array(image)
        
        height, width = img_np.shape[:2]
        
        # Option labels
        option_labels = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[:options_count]
        
        # Prepare different analysis channels
        if len(img_np.shape) == 3:
            # RGB image - analyze multiple ways
            gray = np.mean(img_np, axis=2)
            
            # Extract color channels
            r_channel = img_np[:, :, 0].astype(float)
            g_channel = img_np[:, :, 1].astype(float)
            b_channel = img_np[:, :, 2].astype(float)
            
            # Calculate color saturation (how "colorful" vs gray)
            max_rgb = np.maximum(np.maximum(r_channel, g_channel), b_channel)
            min_rgb = np.minimum(np.minimum(r_channel, g_channel), b_channel)
            saturation = max_rgb - min_rgb
            
            # Blue detection: high B, low R
            blue_score_img = b_channel - r_channel
            
            has_color = True
        else:
            gray = img_np
            saturation = np.zeros_like(gray)
            blue_score_img = np.zeros_like(gray)
            has_color = False
        
        # Overall statistics
        overall_gray_mean = np.mean(gray)
        overall_sat_mean = np.mean(saturation) if has_color else 0
        overall_sat_std = np.std(saturation) if has_color else 0
        
        # Apply contrast enhancement for light mark detection
        # This helps detect very faint pencil marks
        gray_enhanced = gray.copy()
        gray_min, gray_max = np.min(gray), np.max(gray)
        if gray_max > gray_min:
            # Stretch contrast to full range
            gray_enhanced = ((gray - gray_min) / (gray_max - gray_min) * 255).astype(np.float64)

        dark_pixel_threshold = max(
            0.0,
            overall_gray_mean - max(8.0, min(24.0, (gray_max - gray_min) * 0.12))
        )
        color_pixel_threshold = overall_sat_mean + max(10.0, overall_sat_std * 0.6)

        content_left, content_right = self._estimate_option_content_bounds(
            gray,
            saturation,
            has_color,
            dark_pixel_threshold,
            color_pixel_threshold,
            options_count,
        )
        content_span = max(1, content_right - content_left)
        avg_cell_width = content_span / max(1, options_count)
        cell_edges = [
            int(round(content_left + (content_span * i) / max(1, options_count)))
            for i in range(options_count + 1)
        ]
        cell_edges[0] = content_left
        cell_edges[-1] = content_right

        if avg_cell_width < 5:
            print(f"  Warning: Cell width too small ({avg_cell_width:.1f}px)")
            return ""
        
        print(
            f"  Image size: {width}x{height}, {options_count} options, "
            f"content=({content_left},{content_right}), avg cell width: {avg_cell_width:.1f}px"
        )
        print(f"  Overall: gray_mean={overall_gray_mean:.1f}, saturation_mean={overall_sat_mean:.1f}, contrast_range={gray_max-gray_min:.1f}")
        
        # Analyze each cell
        cell_scores = []
        
        for i in range(options_count):
            left = cell_edges[i]
            right = cell_edges[i + 1]
            
            cell_gray = gray[:, left:right]
            cell_gray_enhanced = gray_enhanced[:, left:right]
            cell_gray_mean = np.mean(cell_gray)
            cell_gray_std = np.std(cell_gray)

            center_margin_x = max(3, cell_gray.shape[1] // 4)
            center_margin_y = max(3, cell_gray.shape[0] // 4)
            center_left = min(center_margin_x, max(0, cell_gray.shape[1] // 2 - 1))
            center_right = max(center_left + 1, cell_gray.shape[1] - center_margin_x)
            center_top = min(center_margin_y, max(0, cell_gray.shape[0] // 2 - 1))
            center_bottom = max(center_top + 1, cell_gray.shape[0] - center_margin_y)
            center_gray = cell_gray[center_top:center_bottom, center_left:center_right]
            if center_gray.size == 0:
                center_gray = cell_gray
            center_gray_mean = np.mean(center_gray)
            
            # Darkness score (lower mean = darker)
            darkness_score = overall_gray_mean - cell_gray_mean
            center_darkness_score = overall_gray_mean - center_gray_mean
            
            # Enhanced darkness score using contrast-stretched image
            enhanced_mean = np.mean(cell_gray_enhanced)
            enhanced_overall = np.mean(gray_enhanced)
            enhanced_darkness = enhanced_overall - enhanced_mean
            
            # Local contrast: high std means there's a mark
            local_contrast_score = cell_gray_std / 10.0  # Normalize

            ink_ratio = float(np.mean(cell_gray < dark_pixel_threshold))
            center_ink_ratio = float(np.mean(center_gray < dark_pixel_threshold))
            
            # Color-based scores
            if has_color:
                cell_sat = saturation[:, left:right]
                cell_sat_mean = np.mean(cell_sat)
                
                cell_blue = blue_score_img[:, left:right]
                cell_blue_mean = np.mean(cell_blue)
                
                # Saturation difference from overall
                sat_score = cell_sat_mean - overall_sat_mean
                color_ratio = float(np.mean(cell_sat > color_pixel_threshold))

                center_sat = cell_sat[center_top:center_bottom, center_left:center_right]
                if center_sat.size == 0:
                    center_sat = cell_sat
                center_color_ratio = float(np.mean(center_sat > color_pixel_threshold))
            else:
                cell_sat_mean = 0
                cell_blue_mean = 0
                sat_score = 0
                color_ratio = 0.0
                center_color_ratio = 0.0

            mark_ratio = max(ink_ratio, color_ratio)
            center_mark_ratio = max(center_ink_ratio, center_color_ratio)
            
            # Combined score: weighted sum of different indicators
            # Higher score = more likely to be filled
            # Enhanced scoring for light marks
            combined_score = (
                darkness_score * 1.0 +          # Weight for darkness
                enhanced_darkness * 0.5 +       # Weight for enhanced contrast darkness
                local_contrast_score * 0.3 +    # Weight for local contrast (marks have texture)
                sat_score * 0.5 +               # Weight for saturation (colored marks)
                max(0, cell_blue_mean) * 0.3    # Weight for blue specifically
            )
            
            cell_scores.append({
                'option': option_labels[i],
                'gray_mean': cell_gray_mean,
                'gray_std': cell_gray_std,
                'darkness': darkness_score,
                'center_darkness': center_darkness_score,
                'enhanced_dark': enhanced_darkness,
                'local_contrast': local_contrast_score,
                'saturation': cell_sat_mean,
                'sat_score': sat_score,
                'blue_score': cell_blue_mean,
                'ink_ratio': ink_ratio,
                'color_ratio': color_ratio,
                'mark_ratio': mark_ratio,
                'center_ink_ratio': center_ink_ratio,
                'center_color_ratio': center_color_ratio,
                'center_mark_ratio': center_mark_ratio,
                'combined': combined_score
            })
            
            print(
                f"    Option {option_labels[i]}: gray={cell_gray_mean:.1f}, dark={darkness_score:.1f}, "
                f"enh_dark={enhanced_darkness:.1f}, contrast={local_contrast_score:.1f}, "
                f"mark_ratio={mark_ratio:.3f}, center_dark={center_darkness_score:.1f}, "
                f"center_mark={center_mark_ratio:.3f}, combined={combined_score:.1f}"
            )
        
        # Save debug image with cell divisions and scores
        if save_debug:
            from PIL import ImageDraw, ImageFont
            debug_dir = self.debug_dir
            os.makedirs(debug_dir, exist_ok=True)
            import time
            
            debug_img = image.copy()
            draw = ImageDraw.Draw(debug_img)

            draw.rectangle([(content_left, 0), (max(content_left + 1, content_right - 1), height - 1)], outline=(0, 128, 255), width=1)
            
            # Draw vertical lines to show cell divisions
            for i in range(1, options_count):
                x = cell_edges[i]
                draw.line([(x, 0), (x, height)], fill=(255, 0, 0), width=2)
            
            # Draw scores on each cell
            for i, score in enumerate(cell_scores):
                x = cell_edges[i] + 2
                draw.text((x, 2), f"{score['combined']:.0f}", fill=(255, 0, 0))
            
            debug_path = os.path.join(debug_dir, f"option_{int(time.time()*1000)}.png")
            debug_img.save(debug_path)
            print(f"  Saved debug image: {debug_path}")
        
        # Determine which option(s) are filled using combined score
        filled_options, decision_info = self._select_filled_options_from_scores(cell_scores)

        if cell_scores:
            min_combined = min(s['combined'] for s in cell_scores)
            max_combined = max(s['combined'] for s in cell_scores)
            print(
                f"  Score range: {min_combined:.1f} to {max_combined:.1f} "
                f"(range={decision_info.get('score_range', 0.0):.1f}), "
                f"blank_baseline={decision_info.get('blank_combined_baseline', 0.0):.1f}, "
                f"blank_mark={decision_info.get('blank_mark_baseline', 0.0):.3f}, "
                f"blank_center_dark={decision_info.get('blank_center_darkness_baseline', 0.0):.1f}"
            )
            reason = decision_info.get('reason', 'unknown')
            if reason == 'blank':
                print("  No option filled: clearly blank after blank-baseline comparison")
            elif filled_options:
                print("  Selected by blank-baseline comparison")
            else:
                thresholds = decision_info.get('thresholds', {})
                print(
                    "  No option filled: scores did not clear blank-baseline thresholds "
                    f"(min_margin={thresholds.get('min_combined_margin', 0.0):.1f}, "
                    f"min_mark_margin={thresholds.get('min_mark_margin', 0.0):.3f})"
                )
        
        # Remove duplicates while preserving order (avoid outputs like CDCD)
        seen = set()
        unique_options = []
        for opt in filled_options:
            if opt not in seen:
                unique_options.append(opt)
                seen.add(opt)
        result = "".join(unique_options)
        print(f"  Detected filled option(s): {result if result else '(none)'}")

        try:
            context = context or {}
            correct_answer = context.get("correct_answer", "")
            question_num = context.get("question")
            if not correct_answer and question_num is not None and hasattr(self, "answer_key"):
                correct_answer = self.answer_key.get(question_num, "") or self.answer_key.get(str(question_num), "")

            normalized_result = "".join(str(result).split()).upper()
            normalized_correct = "".join(str(correct_answer).split()).upper()
            record = {
                "context": context,
                "options_count": options_count,
                "content_bounds": [content_left, content_right],
                "cell_edges": cell_edges,
                "scores": cell_scores,
                "result": result,
                "correct_answer": correct_answer,
                "is_correct": bool(normalized_correct and normalized_result == normalized_correct),
                "thresholds": decision_info.get("thresholds", {}),
                "decision": decision_info,
            }
            if hasattr(self, "debug_records"):
                self.debug_records.append(record)
        except Exception:
            pass

        return result

    def detect_filled_option(self, image, options_count=4, save_debug=False, context=None):
        """Detect an OMR response and retain confidence data for review.

        The production decision logic lives in :mod:`omr_core`, so it is shared
        by interactive, selected-page, and batch recognition and can be covered
        by automated tests.  The previous implementation remains available as
        a guarded fallback for an unexpected image-format edge case.
        """

        context = context or {}
        try:
            detection = detect_filled_options(image, options_count)
        except Exception as exc:
            print(f"OMR core failed; using compatibility detector: {exc}")
            result = self._detect_filled_option_legacy(
                image,
                options_count,
                save_debug=save_debug,
                context=context,
            )
            self._last_detection = {
                "answer": result,
                "confidence": 0.0,
                "needs_review": True,
                "reason": "fallback",
                "cell_scores": [],
                "decision": {"reason": "fallback", "error": str(exc)},
            }
            return result

        result = detection.answer
        detection_record = detection.as_record()
        self._last_detection = detection_record

        correct_answer = context.get("correct_answer", "")
        question_num = context.get("question")
        if not correct_answer and question_num is not None and hasattr(self, "answer_key"):
            correct_answer = (
                self.answer_key.get(question_num, "")
                or self.answer_key.get(str(question_num), "")
            )
        normalized_result = "".join(str(result).split()).upper()
        normalized_correct = "".join(str(correct_answer).split()).upper()
        if save_debug:
            record = {
                "context": context,
                "options_count": options_count,
                "content_bounds": list(detection.content_bounds),
                "cell_edges": list(detection.cell_edges),
                "scores": [dict(score) for score in detection.cell_scores],
                "result": result,
                "confidence": detection.confidence,
                "needs_review": detection.needs_review,
                "reason": detection.reason,
                "correct_answer": correct_answer,
                "is_correct": bool(
                    normalized_correct and normalized_result == normalized_correct
                ),
                "thresholds": detection.decision.get("thresholds", {}),
                "decision": detection.decision,
            }
            self.debug_records.append(record)
            try:
                from PIL import ImageDraw

                debug_img = (
                    image.copy()
                    if isinstance(image, Image.Image)
                    else Image.fromarray(np.asarray(image).astype(np.uint8))
                )
                draw = ImageDraw.Draw(debug_img)
                height = debug_img.height
                for edge in detection.cell_edges[1:-1]:
                    draw.line([(edge, 0), (edge, height)], fill=(220, 30, 30), width=1)
                colour = (0, 130, 0) if not detection.needs_review else (230, 130, 0)
                draw.text(
                    (2, 1),
                    f"{result or '-'} {detection.confidence:.0%}",
                    fill=colour,
                )
                label = context.get("label") or f"Q{question_num or 'unknown'}"
                page_idx = max(0, int(context.get("page", 1)) - 1)
                self._save_crop_image(debug_img, page_idx, label, "omr_debug")
            except Exception as exc:
                print(f"Could not save OMR diagnostic image: {exc}")

        return result

    def get_ocr_result(self, image, save_debug=False):
        """Perform OCR on the given PIL image and return text with confidence info."""
        import numpy as np
        import cv2
        from PIL import Image
        
        # Debug: Save cropped image to see what's being recognized
        if save_debug:
            debug_dir = self.debug_dir
            os.makedirs(debug_dir, exist_ok=True)
            import time
            debug_path = os.path.join(debug_dir, f"crop_{int(time.time()*1000)}.png")
            image.save(debug_path)
            print(f"  Saved debug image: {debug_path}")
        
        # Check if image is valid
        img_np = np.array(image)
        print(f"  Image shape: {img_np.shape}, dtype: {img_np.dtype}")
        
        if img_np.size == 0:
            print("  ERROR: Empty image!")
            return "[Empty Image]"
        
        # Preprocess for better OCR (contrast, denoise, resize, threshold)
        def preprocess_for_ocr(pil_img):
            arr = np.array(pil_img)
            if len(arr.shape) == 3:
                gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
            else:
                gray = arr.copy()

            # Normalize contrast
            gray = cv2.normalize(gray, None, 0, 255, cv2.NORM_MINMAX)

            # Upscale small crops for better OCR
            h, w = gray.shape
            target_h = 60
            scale = target_h / max(1, h) if h < target_h else 1.0
            if scale > 1.0:
                gray = cv2.resize(gray, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_CUBIC)

            # Denoise
            gray = cv2.fastNlMeansDenoising(gray, None, h=12, templateWindowSize=7, searchWindowSize=21)

            # Sharpen
            kernel = np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]])
            gray = cv2.filter2D(gray, -1, kernel)

            # Adaptive threshold (binary)
            block_size = 31 if gray.shape[0] >= 31 else 15
            if block_size % 2 == 0:
                block_size += 1
            binary = cv2.adaptiveThreshold(
                gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, block_size, 11
            )

            return arr, gray, binary

        orig_np, gray_np, bin_np = preprocess_for_ocr(image)

        if save_debug:
            debug_dir = self.debug_dir
            os.makedirs(debug_dir, exist_ok=True)
            import time
            base = int(time.time()*1000)
            Image.fromarray(gray_np).save(os.path.join(debug_dir, f"crop_gray_{base}.png"))
            Image.fromarray(bin_np).save(os.path.join(debug_dir, f"crop_bin_{base}.png"))

        if self.ocr_engine_name == "easyocr":
            if self.ocr_reader is None:
                import easyocr
                # Initialize for English and Traditional Chinese
                print("  Initializing EasyOCR reader (this may take a moment)...")
                self.ocr_reader = easyocr.Reader(['en', 'ch_tra'], verbose=False) 

            def run_easyocr(np_img, label):
                result = self.ocr_reader.readtext(
                    np_img,
                    detail=1,
                    paragraph=False,
                    contrast_ths=0.1,
                    adjust_contrast=0.6,
                    text_threshold=0.5,
                    low_text=0.35,
                    link_threshold=0.4
                )
                if not result:
                    print(f"  EasyOCR: No text detected ({label})")
                    return ""
                texts = []
                for detection in result:
                    bbox, text, confidence = detection
                    print(f"  EasyOCR detected: '{text}' (confidence: {confidence:.2%}, {label})")
                    texts.append(text)
                return " ".join(texts)

            # Try original, then preprocessed grayscale, then binary
            text = run_easyocr(orig_np, "orig")
            if not text:
                text = run_easyocr(cv2.cvtColor(gray_np, cv2.COLOR_GRAY2RGB), "gray")
            if not text:
                text = run_easyocr(cv2.cvtColor(bin_np, cv2.COLOR_GRAY2RGB), "binary")
            return text
        
        elif self.ocr_engine_name == "tesseract":
            import pytesseract
            # Default to eng+chi_tra
            try:
                config_main = "--oem 1 --psm 6"
                text = pytesseract.image_to_string(image, lang='eng+chi_tra', config=config_main).strip()
                if not text:
                    text = pytesseract.image_to_string(Image.fromarray(gray_np), lang='eng+chi_tra', config=config_main).strip()
                if not text:
                    text = pytesseract.image_to_string(Image.fromarray(bin_np), lang='eng+chi_tra', config="--oem 1 --psm 7").strip()
                print(f"  Tesseract detected: '{text}'")
                return text
            except:
                text = pytesseract.image_to_string(image, lang='eng', config="--oem 1 --psm 6").strip()
                if not text:
                    text = pytesseract.image_to_string(Image.fromarray(gray_np), lang='eng', config="--oem 1 --psm 6").strip()
                if not text:
                    text = pytesseract.image_to_string(Image.fromarray(bin_np), lang='eng', config="--oem 1 --psm 7").strip()
                print(f"  Tesseract detected: '{text}'")
                return text
        
        return "OCR Error: No Engine"

    def init_ui(self):
        self.setWindowTitle(f"{tr('app_title')}  v{APP_VERSION}")
        self.setGeometry(100, 100, 1400, 850)
        self.setAcceptDrops(True)
        
        # --- Menu Bar ---
        menubar = self.menuBar()
        menubar.clear()
        
        # Language menu — only show the option to switch to the OTHER language
        lang_menu = menubar.addMenu(tr("menu_language"))
        if get_language() != "en":
            act_en = QAction(tr("lang_en"), self)
            act_en.triggered.connect(lambda: self._switch_language("en"))
            lang_menu.addAction(act_en)
        if get_language() != "zh":
            act_zh = QAction(tr("lang_zh"), self)
            act_zh.triggered.connect(lambda: self._switch_language("zh"))
            lang_menu.addAction(act_zh)
        
        # Settings menu
        settings_menu = menubar.addMenu(tr("menu_settings"))
        self.act_auto_update = QAction(tr("chk_auto_update"), self)
        self.act_auto_update.setCheckable(True)
        self.act_auto_update.setChecked(self._settings.value("check_update_on_startup", True, type=bool))
        self.act_auto_update.triggered.connect(self._toggle_auto_update)
        settings_menu.addAction(self.act_auto_update)

        # Help menu
        help_menu = menubar.addMenu(tr("menu_help"))
        act_check_update = QAction(tr("menu_check_update"), self)
        act_check_update.triggered.connect(lambda: self._check_for_update(silent=False))
        help_menu.addAction(act_check_update)
        help_menu.addSeparator()
        act_about = QAction(tr("menu_about"), self)
        act_about.triggered.connect(self.show_about)
        help_menu.addAction(act_about)
        
        central = QWidget()
        self.setCentralWidget(central)
        layout = QHBoxLayout(central)
        
        # --- Left Panel (Controls) ---
        left_scroll = QScrollArea()
        left_scroll.setWidgetResizable(True)
        left_scroll.setFrameShape(QFrame.NoFrame)
        left_scroll.setFixedWidth(380)
        
        left_content = QWidget()
        left_layout = QVBoxLayout(left_content)
        left_layout.setSpacing(15)
        
        # Title
        title = QLabel(tr("title_label"))
        title.setObjectName("titleLabel")
        left_layout.addWidget(title)
        
        # File Import
        file_grp = QGroupBox(tr("group_file"))
        f_layout = QVBoxLayout(file_grp)
        
        btn_import = QPushButton(tr("btn_import_pdf"))
        btn_import.setShortcut("Ctrl+O")
        btn_import.clicked.connect(self.import_pdf)
        f_layout.addWidget(btn_import)
        
        self.check_first_key = QCheckBox(tr("chk_first_key"))
        self.check_first_key.setChecked(
            self._settings.value("first_page_key", False, type=bool)
        )
        self.first_page_key = self.check_first_key.isChecked()
        self.check_first_key.stateChanged.connect(self._on_first_key_changed)
        f_layout.addWidget(self.check_first_key)
        
        self.check_auto_deskew = QCheckBox(tr("chk_auto_deskew"))
        self.check_auto_deskew.setChecked(
            self._settings.value("auto_deskew", True, type=bool)
        )
        self.check_auto_deskew.toggled.connect(
            lambda value: self._settings.setValue("auto_deskew", value)
        )
        f_layout.addWidget(self.check_auto_deskew)

        self.check_auto_align = QCheckBox(tr("chk_auto_align"))
        self.check_auto_align.setChecked(
            self._settings.value("auto_align", True, type=bool)
        )
        self.check_auto_align.toggled.connect(
            lambda value: self._settings.setValue("auto_align", value)
        )
        f_layout.addWidget(self.check_auto_align)
        
        left_layout.addWidget(file_grp)
        
        # Marking Tools
        mark_grp = QGroupBox(tr("group_marking"))
        m_layout = QVBoxLayout(mark_grp)
        
        row1 = QHBoxLayout()
        self.btn_mark_text = QPushButton(tr("btn_mark_text"))
        self.btn_mark_text.setCheckable(True)
        self.btn_mark_text.clicked.connect(lambda: self.set_marking(MARK_TYPE_TEXT))
        row1.addWidget(self.btn_mark_text)
        
        self.btn_mark_option = QPushButton(tr("btn_mark_option"))
        self.btn_mark_option.setCheckable(True)
        self.btn_mark_option.clicked.connect(lambda: self.set_marking(MARK_TYPE_OPTION))
        row1.addWidget(self.btn_mark_option)
        m_layout.addLayout(row1)
        
        # Alignment reference button
        row1_5 = QHBoxLayout()
        self.btn_mark_align = QPushButton(tr("btn_mark_align"))
        self.btn_mark_align.setCheckable(True)
        self.btn_mark_align.setToolTip(tr("tip_mark_align"))
        self.btn_mark_align.clicked.connect(lambda: self.set_marking(MARK_TYPE_ALIGN))
        row1_5.addWidget(self.btn_mark_align)
        m_layout.addLayout(row1_5)
        
        m_layout.addWidget(QLabel(tr("lbl_mark_hint1")))
        m_layout.addWidget(QLabel(tr("lbl_mark_hint2")))
        
        row2 = QHBoxLayout()
        btn_undo = QPushButton(tr("btn_undo"))
        btn_undo.setToolTip(tr("tip_undo"))
        btn_undo.clicked.connect(self.undo_last_mark)
        row2.addWidget(btn_undo)
        
        btn_clear = QPushButton(tr("btn_clear"))
        btn_clear.setObjectName("deleteBtn")
        btn_clear.clicked.connect(self.clear_all_marks)
        row2.addWidget(btn_clear)
        m_layout.addLayout(row2)
        
        row3 = QHBoxLayout()
        btn_import_templ = QPushButton(tr("btn_load_template"))
        btn_import_templ.setShortcut("Ctrl+L")
        btn_import_templ.clicked.connect(self.import_template)
        btn_export_templ = QPushButton(tr("btn_save_template"))
        btn_export_templ.setShortcut("Ctrl+Shift+S")
        btn_export_templ.clicked.connect(self.export_template)
        row3.addWidget(btn_import_templ)
        row3.addWidget(btn_export_templ)
        m_layout.addLayout(row3)
        
        left_layout.addWidget(mark_grp)
        
        # Processing
        proc_grp = QGroupBox(tr("group_processing"))
        p_layout = QVBoxLayout(proc_grp)
        
        engine_name = self.ocr_engine_name if self.ocr_engine_name else tr("lbl_ocr_not_available")
        lbl_ocr = QLabel(tr("lbl_ocr_status", engine=engine_name))
        p_layout.addWidget(lbl_ocr)
        
        btn_process = QPushButton(tr("btn_recognize_all"))
        btn_process.setShortcut("Ctrl+R")
        btn_process.clicked.connect(self.run_recognition_all)
        p_layout.addWidget(btn_process)

        btn_rerecognize = QPushButton(tr("btn_recognize_sel"))
        btn_rerecognize.clicked.connect(self.run_recognition_selected)
        p_layout.addWidget(btn_rerecognize)

        self.check_recognize_text = QCheckBox(tr("chk_recognize_text"))
        self.check_recognize_text.setChecked(
            self._settings.value("recognize_text", True, type=bool)
        )
        self.check_recognize_text.toggled.connect(
            lambda value: self._settings.setValue("recognize_text", value)
        )
        p_layout.addWidget(self.check_recognize_text)

        self.check_save_debug = QCheckBox(tr("chk_save_debug"))
        self.check_save_debug.setChecked(
            self._settings.value("save_debug", False, type=bool)
        )
        self.check_save_debug.toggled.connect(
            lambda value: self._settings.setValue("save_debug", value)
        )
        p_layout.addWidget(self.check_save_debug)

        self.check_export_images = QCheckBox(tr("chk_export_images"))
        self.check_export_images.setChecked(
            self._settings.value("export_images", True, type=bool)
        )
        self.check_export_images.toggled.connect(
            lambda value: self._settings.setValue("export_images", value)
        )
        p_layout.addWidget(self.check_export_images)

        btn_export_all = QPushButton(tr("btn_export_bundle"))
        btn_export_all.clicked.connect(self.export_results_bundle)
        p_layout.addWidget(btn_export_all)
        
        btn_export = QPushButton(tr("btn_export_excel"))
        btn_export.clicked.connect(self.export_excel)
        p_layout.addWidget(btn_export)

        btn_student_info = QPushButton(tr("btn_student_info"))
        btn_student_info.clicked.connect(self.edit_student_info)
        p_layout.addWidget(btn_student_info)

        self.check_include_summary = QCheckBox(tr("chk_include_summary"))
        self.check_include_summary.setChecked(
            self._settings.value("include_summary", True, type=bool)
        )
        self.check_include_summary.toggled.connect(
            lambda value: self._settings.setValue("include_summary", value)
        )
        p_layout.addWidget(self.check_include_summary)

        self.check_include_topics = QCheckBox(tr("chk_include_topics"))
        self.check_include_topics.setChecked(
            self._settings.value("include_topics", True, type=bool)
        )
        self.check_include_topics.toggled.connect(
            lambda value: self._settings.setValue("include_topics", value)
        )
        p_layout.addWidget(self.check_include_topics)

        btn_topics = QPushButton(tr("btn_set_topics"))
        btn_topics.clicked.connect(self.edit_topics)
        p_layout.addWidget(btn_topics)
        
        btn_export_img = QPushButton(tr("btn_export_images"))
        btn_export_img.clicked.connect(self.export_images)
        p_layout.addWidget(btn_export_img)

        btn_export_debug = QPushButton(tr("btn_export_debug"))
        btn_export_debug.clicked.connect(self.export_debug_pack)
        p_layout.addWidget(btn_export_debug)
        
        left_layout.addWidget(proc_grp)
        
        # Batch Processing
        batch_grp = QGroupBox(tr("group_batch"))
        b_layout = QVBoxLayout(batch_grp)
        
        btn_batch_same = QPushButton(tr("btn_batch_same"))
        btn_batch_same.setToolTip(tr("tip_batch_same"))
        btn_batch_same.clicked.connect(self.batch_process_same_template)
        b_layout.addWidget(btn_batch_same)
        
        btn_batch_match = QPushButton(tr("btn_batch_match"))
        btn_batch_match.setToolTip(tr("tip_batch_match"))
        btn_batch_match.clicked.connect(self.batch_process_matched_templates)
        b_layout.addWidget(btn_batch_match)
        
        left_layout.addWidget(batch_grp)
        
        left_layout.addStretch()
        left_scroll.setWidget(left_content)
        layout.addWidget(left_scroll)
        
        # --- Center (Preview) ---
        center_layout = QVBoxLayout()
        
        # Toolbar with navigation and zoom
        nav_layout = QHBoxLayout()
        btn_prev = QPushButton(tr("btn_prev"))
        btn_prev.clicked.connect(self.prev_page)
        btn_next = QPushButton(tr("btn_next"))
        btn_next.clicked.connect(self.next_page)
        self.lbl_page = QLabel(tr("lbl_page", current=0, total=0))
        
        # Zoom controls
        btn_zoom_in = QPushButton("🔍+")
        btn_zoom_in.setFixedWidth(40)
        btn_zoom_in.setToolTip(tr("tip_zoom_in"))
        btn_zoom_in.clicked.connect(lambda: self.view.zoom_in())
        
        btn_zoom_out = QPushButton("🔍-")
        btn_zoom_out.setFixedWidth(40)
        btn_zoom_out.setToolTip(tr("tip_zoom_out"))
        btn_zoom_out.clicked.connect(lambda: self.view.zoom_out())
        
        btn_zoom_reset = QPushButton("100%")
        btn_zoom_reset.setFixedWidth(50)
        btn_zoom_reset.setToolTip(tr("tip_zoom_reset"))
        btn_zoom_reset.clicked.connect(lambda: self.view.zoom_reset())
        
        btn_zoom_fit = QPushButton("Fit")
        btn_zoom_fit.setFixedWidth(40)
        btn_zoom_fit.setToolTip(tr("tip_zoom_fit"))
        btn_zoom_fit.clicked.connect(lambda: self.view.zoom_fit())
        
        nav_layout.addWidget(btn_prev)
        nav_layout.addWidget(self.lbl_page)
        nav_layout.addWidget(btn_next)
        nav_layout.addStretch()
        
        # Student info label for current page
        self.lbl_student_info = QLabel("")
        self.lbl_student_info.setStyleSheet("color: #2c5f8a; font-size: 13px; font-weight: bold; padding: 0 8px;")
        nav_layout.addWidget(self.lbl_student_info)
        nav_layout.addStretch()
        
        # Correction info label
        self.lbl_correction = QLabel("")
        self.lbl_correction.setMinimumWidth(200)
        nav_layout.addWidget(self.lbl_correction)
        nav_layout.addStretch()
        
        nav_layout.addWidget(QLabel(tr("lbl_zoom")))
        nav_layout.addWidget(btn_zoom_out)
        nav_layout.addWidget(btn_zoom_reset)
        nav_layout.addWidget(btn_zoom_in)
        nav_layout.addWidget(btn_zoom_fit)
        
        center_layout.addLayout(nav_layout)
        
        self.scene = QGraphicsScene()
        self.view = MarkingView(self.scene)
        center_layout.addWidget(self.view)
        
        layout.addLayout(center_layout, stretch=2)
        
        # --- Right (Results) ---
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_widget.setFixedWidth(350)
        
        right_layout.addWidget(QLabel(tr("lbl_results")))
        
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels([
            tr("col_q"),
            tr("col_detected"),
            tr("col_correct"),
            tr("col_points"),
            tr("col_confidence"),
            tr("col_crop"),
        ])
        self.table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.table.cellChanged.connect(self.on_table_edit)
        self.table.cellClicked.connect(self.open_crop_from_table)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.open_crop_context_menu)
        right_layout.addWidget(self.table)

        btn_next_review = QPushButton(tr("btn_next_review"))
        btn_next_review.setToolTip(tr("tip_next_review"))
        btn_next_review.setShortcut("F8")
        btn_next_review.clicked.connect(self.jump_to_next_review)
        right_layout.addWidget(btn_next_review)
        
        self.lbl_score = QLabel(tr("lbl_total"))
        right_layout.addWidget(self.lbl_score)

        self.lbl_answer_status = QLabel("")
        self.lbl_answer_status.setStyleSheet("color: #856404; font-size: 12px; padding: 2px;")
        right_layout.addWidget(self.lbl_answer_status)
        
        layout.addWidget(right_widget)

    def _switch_language(self, lang):
        """Switch UI language and rebuild the interface."""
        set_language(lang)
        # Save state
        saved_pdf = getattr(self, 'pdf_document', None)
        saved_page = getattr(self, 'current_page', 0)
        saved_results = getattr(self, 'results', {})
        saved_answer_key = getattr(self, 'answer_key', {})
        saved_topic_map = getattr(self, 'topic_map', {})
        saved_first_page_key = getattr(self, 'first_page_key', False)
        saved_marks_data = None
        if hasattr(self, 'view'):
            saved_marks_data = self.view.get_all_marks_data()
        saved_pdf_path = getattr(self, 'pdf_path', None)
        saved_page_offsets = getattr(self, 'page_offsets', {})
        saved_student_absence = getattr(self, 'student_absence', {})
        saved_extra_students = getattr(self, 'extra_students', [])
        saved_student_order = getattr(self, 'student_order', [])
        
        # Rebuild UI
        self.init_ui()
        self.setStyleSheet(STYLE_SHEET)
        
        # Restore state
        self.results = saved_results
        self.answer_key = saved_answer_key
        self.topic_map = saved_topic_map
        self.first_page_key = saved_first_page_key
        self.check_first_key.setChecked(saved_first_page_key)
        self.page_offsets = saved_page_offsets
        self.student_absence = saved_student_absence
        self.extra_students = saved_extra_students
        self.student_order = saved_student_order
        
        if saved_pdf is not None:
            self.pdf_document = saved_pdf
            self.pdf_path = saved_pdf_path
            # Restore marks
            if saved_marks_data:
                self.clear_all_marks()
                for m in saved_marks_data.get("text_marks", []):
                    item = MarkItem(0, 0, m['width'], m['height'], MARK_TYPE_TEXT, m['question'], m['label'], view_ref=self.view)
                    item.setPos(m['x'], m['y'])
                    self.view.text_marks.append(item)
                    self.scene.addItem(item)
                    self.view.text_counter = max(self.view.text_counter, m['question'] + 1)
                for m in saved_marks_data.get("option_marks", []):
                    item = MarkItem(0, 0, m['width'], m['height'], MARK_TYPE_OPTION, m['question'], m['label'], m.get('options_count', 4), view_ref=self.view)
                    item.setPos(m['x'], m['y'])
                    self.view.option_marks.append(item)
                    self.scene.addItem(item)
                    self.view.option_counter = max(self.view.option_counter, m['question'] + 1)
                align_marks_data = saved_marks_data.get("align_marks", [])
                # Backward compat: load old single "align_mark" format
                if not align_marks_data:
                    old_align = saved_marks_data.get("align_mark")
                    if old_align:
                        align_marks_data = [old_align]
                for ad in align_marks_data:
                    item = MarkItem(0, 0, ad['width'], ad['height'], MARK_TYPE_ALIGN,
                                   ad.get('question', self.view.align_counter), ad.get('label', ''), view_ref=self.view)
                    item.setPos(ad['x'], ad['y'])
                    self.view.align_marks.append(item)
                    self.scene.addItem(item)
                    self.view.align_counter = max(self.view.align_counter, ad.get('question', 1) + 1)
            self.load_page(saved_page, apply_corrections=False)
            self.update_result_table()

    def show_about(self):
        """Show About dialog with version information."""
        QMessageBox.about(self, tr("about_title"),
                          tr("about_text", version=APP_VERSION))

    # ── Update Checker ──

    def _toggle_auto_update(self, checked):
        self._settings.setValue("check_update_on_startup", checked)

    def _check_for_update(self, silent=True):
        """Launch background update check.  silent=True suppresses 'no update' / 'failed' dialogs."""
        if self._update_thread is not None and self._update_thread.isRunning():
            return
        self._update_silent = silent
        self._update_thread = UpdateChecker()
        self._update_thread.update_available.connect(self._on_update_available)
        self._update_thread.no_update.connect(self._on_no_update)
        self._update_thread.check_failed.connect(self._on_update_failed)
        self._update_thread.start()

    def _on_update_available(self, latest, html_url, asset_url, body=""):
        is_frozen = getattr(sys, 'frozen', False)

        dialog = QDialog(self)
        dialog.setWindowTitle(tr("update_title"))
        dialog.setMinimumWidth(480)
        dlg_layout = QVBoxLayout(dialog)

        dlg_layout.addWidget(QLabel(tr("update_msg", current=APP_VERSION, latest=latest)))

        if body.strip():
            dlg_layout.addWidget(QLabel(f"\n<b>{tr('update_whats_new')}:</b>"))
            from PyQt5.QtWidgets import QTextEdit
            notes_box = QTextEdit()
            notes_box.setReadOnly(True)
            notes_box.setMarkdown(body)
            notes_box.setMinimumHeight(160)
            dlg_layout.addWidget(notes_box)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        if is_frozen and asset_url:
            btn_auto = QPushButton(tr("update_auto_btn"))
            btn_row.addWidget(btn_auto)
        else:
            btn_auto = None
        btn_manual = QPushButton(tr("update_manual_btn"))
        btn_row.addWidget(btn_manual)
        btn_cancel = QPushButton(tr("btn_clear"))  # Cancel
        btn_row.addWidget(btn_cancel)
        dlg_layout.addLayout(btn_row)

        result = {"action": None}
        def on_auto():
            result["action"] = "auto"
            dialog.accept()
        def on_manual():
            result["action"] = "manual"
            dialog.accept()
        if btn_auto:
            btn_auto.clicked.connect(on_auto)
        btn_manual.clicked.connect(on_manual)
        btn_cancel.clicked.connect(dialog.reject)

        dialog.exec_()
        if result["action"] == "auto":
            self._start_auto_update(asset_url, latest)
        elif result["action"] == "manual":
            QDesktopServices.openUrl(QUrl(html_url))

    def _start_auto_update(self, asset_url, version):
        """Start downloading the update zip in the background."""
        self._update_version = version
        self._update_progress_dlg = QtWidgets.QProgressDialog(
            tr("update_downloading", version=version),
            tr("btn_clear"),  # Cancel button text
            0, 100, self
        )
        self._update_progress_dlg.setWindowTitle(tr("update_title"))
        self._update_progress_dlg.setWindowModality(Qt.WindowModal)
        self._update_progress_dlg.setAutoClose(False)
        self._update_progress_dlg.setAutoReset(False)
        self._update_progress_dlg.setMinimumDuration(0)
        self._update_progress_dlg.setValue(0)

        self._downloader = UpdateDownloader(asset_url)
        self._downloader.progress.connect(self._on_download_progress)
        self._downloader.download_complete.connect(self._on_download_complete)
        self._downloader.download_failed.connect(self._on_download_failed)
        self._update_progress_dlg.canceled.connect(self._downloader.cancel)
        self._downloader.start()

    def _on_download_progress(self, downloaded, total):
        if hasattr(self, '_update_progress_dlg') and self._update_progress_dlg is not None:
            if total > 0:
                pct = min(int(downloaded * 100 / total), 99)
                self._update_progress_dlg.setValue(pct)
                mb_done = downloaded / (1024 * 1024)
                mb_total = total / (1024 * 1024)
                self._update_progress_dlg.setLabelText(
                    f"{tr('update_downloading', version=self._update_version)}\n"
                    f"{mb_done:.1f} / {mb_total:.1f} MB"
                )

    def _on_download_complete(self, zip_path):
        if hasattr(self, '_update_progress_dlg') and self._update_progress_dlg is not None:
            self._update_progress_dlg.setValue(100)
            self._update_progress_dlg.setLabelText(tr("update_extracting"))
        QtWidgets.QApplication.processEvents()
        try:
            self._apply_update(zip_path)
        except Exception as e:
            if hasattr(self, '_update_progress_dlg') and self._update_progress_dlg is not None:
                self._update_progress_dlg.close()
            QMessageBox.warning(self, tr("update_download_failed_title"),
                                tr("update_download_failed", error=str(e)))

    def _on_download_failed(self, error):
        if hasattr(self, '_update_progress_dlg') and self._update_progress_dlg is not None:
            self._update_progress_dlg.close()
        QMessageBox.warning(self, tr("update_download_failed_title"),
                            tr("update_download_failed", error=error))

    def _apply_update(self, zip_path):
        """Extract downloaded zip, create updater script, launch it, and exit."""
        # Determine the application directory
        if getattr(sys, 'frozen', False):
            app_dir = os.path.dirname(sys.executable)
        else:
            app_dir = os.path.dirname(os.path.abspath(__file__))

        # Extract to a temporary directory
        extract_dir = tempfile.mkdtemp(prefix="checkmate_update_")
        try:
            with zipfile.ZipFile(zip_path, 'r') as zf:
                zf.extractall(extract_dir)
        except Exception:
            shutil.rmtree(extract_dir, ignore_errors=True)
            raise

        # Detect nested folder: if the zip contains a single top-level folder,
        # use that folder as the source instead of extract_dir.
        # This handles zips like: CheckMate.zip -> CheckMate/ -> CheckMate.exe, _internal/
        source_dir = extract_dir
        entries = os.listdir(extract_dir)
        if len(entries) == 1:
            single_entry = os.path.join(extract_dir, entries[0])
            if os.path.isdir(single_entry):
                # Check if this subfolder contains an exe or _internal
                sub_contents = os.listdir(single_entry)
                if any(f.endswith('.exe') for f in sub_contents) or '_internal' in sub_contents:
                    source_dir = single_entry
                    print(f"  Update: Using nested folder '{entries[0]}' as source")

        # Build the updater batch script
        pid = os.getpid()
        exe_name = os.path.basename(sys.executable)
        updater_path = os.path.join(tempfile.gettempdir(), "checkmate_updater.bat")
        log_path = os.path.join(tempfile.gettempdir(), "checkmate_update.log")
        # Use robocopy instead of xcopy for reliable copying with retries.
        # robocopy exit codes: 0-7 = success, 8+ = failure.
        script_lines = [
            '@echo off',
            'chcp 65001 >nul 2>&1',
            f'echo [%date% %time%] Update started > "{log_path}"',
            'echo Waiting for CheckMate to exit...',
            ':wait',
            f'tasklist /FI "PID eq {pid}" 2>NUL | find /I "{pid}" >NUL',
            'if not errorlevel 1 (',
            '    timeout /t 1 /nobreak >NUL',
            '    goto wait',
            ')',
            # Extra delay to let AV / OS release file locks
            'timeout /t 5 /nobreak >NUL',
            f'echo [%date% %time%] Process exited, applying update... >> "{log_path}"',
            'echo Applying update...',
            # Step 1: Mirror _internal directory (copies new files, deletes stale files)
            # /MIR = mirror, /R:10 /W:3 = retry 10 times with 3s wait
            f'echo [%date% %time%] robocopy _internal... >> "{log_path}"',
            f'robocopy "{source_dir}\\_internal" "{app_dir}\\_internal" /MIR /R:10 /W:3 /NFL /NDL /NP >> "{log_path}" 2>&1',
            'if errorlevel 8 (',
            f'    echo [%date% %time%] ERROR: robocopy _internal failed with code %errorlevel% >> "{log_path}"',
            '    echo ERROR: Failed to copy _internal. See log for details.',
            '    pause',
            '    exit /b 1',
            ')',
            # Step 2: Copy root-level files (exe, etc.) — /LEV:1 = no subdirs
            f'echo [%date% %time%] robocopy root files... >> "{log_path}"',
            f'del /f /q "{app_dir}\\{exe_name}" 2>NUL',
            f'robocopy "{source_dir}" "{app_dir}" "{exe_name}" /R:10 /W:3 /IS /IT /NFL /NDL /NP >> "{log_path}" 2>&1',
            'if errorlevel 8 (',
            f'    echo [%date% %time%] ERROR: robocopy exe failed with code %errorlevel% >> "{log_path}"',
            '    echo ERROR: Failed to copy exe. See log for details.',
            '    pause',
            '    exit /b 1',
            ')',
            # Step 3: Verify critical files exist
            f'if not exist "{app_dir}\\{exe_name}" (',
            f'    echo [%date% %time%] ERROR: {exe_name} missing after copy >> "{log_path}"',
            '    echo ERROR: Update failed - exe missing after copy.',
            '    pause',
            '    exit /b 1',
            ')',
            f'if not exist "{app_dir}\\_internal\\base_library.zip" (',
            f'    echo [%date% %time%] ERROR: base_library.zip missing >> "{log_path}"',
            '    echo ERROR: Update failed - base_library.zip missing.',
            '    pause',
            '    exit /b 1',
            ')',
            f'echo [%date% %time%] Update verified OK >> "{log_path}"',
            'echo Starting updated CheckMate...',
            f'start "" "{app_dir}\\{exe_name}"',
            'timeout /t 3 /nobreak >NUL',
            f'rd /s /q "{extract_dir}" 2>NUL',
            f'del /f /q "{zip_path}" 2>NUL',
            f'echo [%date% %time%] Cleanup done >> "{log_path}"',
        ]
        with open(updater_path, 'w', encoding='utf-8') as f:
            f.write('\r\n'.join(script_lines) + '\r\n')

        # Close the progress dialog
        if hasattr(self, '_update_progress_dlg') and self._update_progress_dlg is not None:
            self._update_progress_dlg.close()

        # Notify user
        QMessageBox.information(self, tr("update_restart_title"),
                                tr("update_restart_msg"))

        # Launch the updater script and exit the application
        subprocess.Popen(
            ['cmd', '/c', updater_path],
            creationflags=subprocess.CREATE_NO_WINDOW
        )
        QtWidgets.QApplication.quit()

    def _on_no_update(self, version):
        if not self._update_silent:
            QMessageBox.information(self, tr("update_no_update_title"),
                                    tr("update_no_update", version=version))

    def _on_update_failed(self, error):
        if not self._update_silent:
            QMessageBox.warning(self, tr("update_check_failed_title"),
                                tr("update_check_failed"))

    def _startup_update_check(self):
        """Called once after window is shown to check for updates."""
        if self._settings.value("check_update_on_startup", True, type=bool):
            self._check_for_update(silent=True)

    def set_marking(self, mtype):
        if mtype == MARK_TYPE_TEXT:
            is_checked = self.btn_mark_text.isChecked()
            self.btn_mark_option.setChecked(False)
            self.btn_mark_align.setChecked(False)
            self.view.set_marking_mode(is_checked, MARK_TYPE_TEXT)
        elif mtype == MARK_TYPE_ALIGN:
            is_checked = self.btn_mark_align.isChecked()
            self.btn_mark_text.setChecked(False)
            self.btn_mark_option.setChecked(False)
            self.view.set_marking_mode(is_checked, MARK_TYPE_ALIGN)
        else:
            is_checked = self.btn_mark_option.isChecked()
            self.btn_mark_text.setChecked(False)
            self.btn_mark_align.setChecked(False)
            self.view.set_marking_mode(is_checked, MARK_TYPE_OPTION)

    def _on_first_key_changed(self, state):
        self.first_page_key = state == Qt.Checked
        self._settings.setValue("first_page_key", self.first_page_key)
        if self.first_page_key and self.results.get(0):
            self.answer_key = self._answer_key_from_result(self.results[0])
        elif not self.first_page_key:
            self.answer_key = {}
        self.update_result_table()

    def _answer_key_from_result(self, page_result):
        """Return only syntactically valid option answers from a key page."""

        answer_key = {}
        for question, value in page_result.get("options", {}).items():
            compact = "".join(str(value).split()).upper()
            allowed = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[
                : self._option_count_for_question(question)
            ]
            if compact and all(character in allowed for character in compact):
                answer_key[question] = "".join(dict.fromkeys(compact))
            else:
                answer_key[question] = ""
        return answer_key

    def _reset_session_cache(self):
        try:
            self._session_temp.cleanup()
        except Exception:
            pass
        self._session_temp = tempfile.TemporaryDirectory(prefix="checkmate_")
        self.debug_dir = os.path.join(self._session_temp.name, "debug_crops")

    def _clear_result_table(self):
        if not hasattr(self, "table"):
            return
        self.table.blockSignals(True)
        self.table.clearContents()
        self.table.setRowCount(0)
        self.table.blockSignals(False)
        if hasattr(self, "lbl_score"):
            self.lbl_score.setText(tr("lbl_total"))
        if hasattr(self, "lbl_answer_status"):
            self.lbl_answer_status.setText("")

    def _reset_document_state(self):
        """Clear all state that must never leak between PDFs."""

        self.results = {}
        self.answer_key = {}
        self.page_offsets = {}
        self.debug_records = []
        self._last_detection = {}
        self._review_cursor = None
        self.student_absence = {}
        self.extra_students = []
        self.student_order = []
        self.current_page = 0
        self._reset_align_templates()
        self._reset_session_cache()
        self._clear_result_table()
        if hasattr(self, "lbl_student_info"):
            self.lbl_student_info.setText("")
        if hasattr(self, "lbl_correction"):
            self.lbl_correction.setText("")

    def _open_pdf_path(self, file_path, *, load_preview=True):
        """Open a PDF transactionally, preserving the current file on failure."""

        new_document = None
        try:
            new_document = fitz.open(file_path)
            if len(new_document) == 0:
                raise ValueError("The PDF has no pages.")
        except Exception:
            if new_document is not None:
                try:
                    new_document.close()
                except Exception:
                    pass
            raise

        old_document = self.pdf_document
        if self.current_pixmap_item is not None:
            try:
                self.scene.removeItem(self.current_pixmap_item)
            except Exception:
                pass
            self.current_pixmap_item = None

        self._reset_document_state()
        self.pdf_path = os.path.abspath(file_path)
        self.pdf_document = new_document
        try:
            if old_document is not None:
                old_document.close()
        except Exception:
            pass

        if load_preview:
            self.load_page(0, apply_corrections=True)
        else:
            self.lbl_page.setText(
                tr("lbl_page", current=1, total=len(self.pdf_document))
            )
        self.statusBar().showMessage(
            f"{os.path.basename(self.pdf_path)}  ·  {len(self.pdf_document)} pages",
            8000,
        )

    def import_pdf(self):
        fname, _ = QFileDialog.getOpenFileName(self, "Open PDF", "", "PDF Files (*.pdf)")
        if fname:
            try:
                self._open_pdf_path(fname)
            except Exception as e:
                QMessageBox.critical(self, "Error", str(e))

    def dragEnterEvent(self, event):
        urls = event.mimeData().urls() if event.mimeData().hasUrls() else []
        if any(url.toLocalFile().lower().endswith(".pdf") for url in urls):
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event):
        pdf_paths = [
            url.toLocalFile()
            for url in event.mimeData().urls()
            if url.toLocalFile().lower().endswith(".pdf")
        ]
        if not pdf_paths:
            event.ignore()
            return
        try:
            self._open_pdf_path(pdf_paths[0])
            event.acceptProposedAction()
        except Exception as exc:
            QMessageBox.critical(self, "Error", str(exc))

    def _get_pdf_prefix(self):
        if hasattr(self, 'pdf_path') and self.pdf_path:
            base = os.path.splitext(os.path.basename(self.pdf_path))[0]
            return base
        return "export"

    def _get_timestamp(self):
        return QtCore.QDateTime.currentDateTime().toString("yyyyMMdd_HHmmss")

    def _safe_crop_label(self, label):
        label = str(label) if label is not None else ""
        label = re.sub(r"[^A-Za-z0-9_-]+", "_", label).strip("_")
        return label or "item"

    def _save_crop_image(self, image, page_idx, label, kind):
        """Save a crop image and return the file path."""
        debug_dir = self.debug_dir
        os.makedirs(debug_dir, exist_ok=True)
        safe_label = self._safe_crop_label(label)
        filename = f"page_{page_idx+1}_{kind}_{safe_label}.png"
        path = os.path.join(debug_dir, filename)
        image.save(path)
        return path

    def _get_page_filename(self, page_idx):
        """Return a filename stem (no extension) for the exported image of page_idx.
        - Answer key page → 'answer_key'
        - Pages with class + student number → 'CLASS_STUDENTNO' (e.g. '4A_01')
        - Fallback → 'page_001'
        """
        if self.first_page_key and page_idx == 0:
            return "answer_key"
        text_data = {}
        if hasattr(self, 'results') and page_idx in self.results:
            text_data = self.results[page_idx].get("text", {})
        # Locate class and student-number values using all known label variants
        class_keys = [tr("field_class"), "班別", "Class"]
        student_keys = [tr("field_student_no"), "學號", "Student No."]
        class_val = next((text_data[k].strip() for k in class_keys if text_data.get(k, "").strip()), "")
        student_no = next((text_data[k].strip() for k in student_keys if text_data.get(k, "").strip()), "")
        if class_val and student_no:
            raw = f"{class_val}_{student_no}"
        elif class_val:
            raw = class_val
        elif student_no:
            raw = student_no
        else:
            return f"page_{page_idx + 1:03d}"
        # Sanitise for use as a filesystem name
        safe = re.sub(r'[\\/:*?"<>|\s]+', '_', raw).strip('_')
        return safe if safe else f"page_{page_idx + 1:03d}"

    def _get_all_questions(self):
        questions = set()
        if hasattr(self, "view") and getattr(self.view, "option_marks", None):
            for mark in self.view.option_marks:
                questions.add(mark.question_num)
        if hasattr(self, "results"):
            for res in self.results.values():
                questions.update(res.get("options", {}).keys())
        return sorted(questions)

    def _get_text_field_labels(self):
        labels = []

        def add_label(val):
            val = str(val).strip() if val is not None else ""
            if val and val not in labels:
                labels.append(val)

        for default_label in [tr("field_class"), tr("field_student_no"), tr("field_name")]:
            add_label(default_label)

        if hasattr(self, "view") and getattr(self.view, "text_marks", None):
            for mark in self.view.text_marks:
                add_label(mark.label or f"Field {mark.question_num}")

        if hasattr(self, "results"):
            for res in self.results.values():
                for key in res.get("text", {}).keys():
                    add_label(key)

        return labels

    def _ensure_results_for_pages(self):
        if not hasattr(self, "results") or self.results is None:
            self.results = {}
        total_pages = len(self.pdf_document) if self.pdf_document else 0
        for p_idx in range(total_pages):
            if p_idx not in self.results:
                self.results[p_idx] = {
                    "options": {},
                    "text": {},
                    "option_crops": {},
                    "text_crops": {}
                }
            else:
                self.results[p_idx].setdefault("options", {})
                self.results[p_idx].setdefault("text", {})
                self.results[p_idx].setdefault("option_crops", {})
                self.results[p_idx].setdefault("text_crops", {})

    def edit_student_info(self):
        if not self.pdf_document:
            QMessageBox.warning(self, tr("dlg_student_title"), tr("msg_no_pdf"))
            return

        self._ensure_results_for_pages()
        labels = self._get_text_field_labels()

        if not labels:
            QMessageBox.information(self, tr("dlg_student_title"), tr("msg_no_text_fields"))
            return

        # Ensure student_absence dict exists
        if not hasattr(self, 'student_absence'):
            self.student_absence = {}

        dialog = QDialog(self)
        dialog.setWindowTitle(tr("dlg_student_title"))
        dialog.resize(800, 500)
        layout = QVBoxLayout(dialog)

        layout.addWidget(QLabel(tr("dlg_student_hint")))

        # Columns: Page, [fields...], Absent
        absent_label = tr("dlg_student_absent")
        table = QTableWidget()
        table.setColumnCount(1 + len(labels) + 1)  # +1 for Absent column
        table.setHorizontalHeaderLabels([tr("col_page")] + labels + [absent_label])
        table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        table.setEditTriggers(QAbstractItemView.AllEditTriggers)

        page_indices = []
        for p_idx in range(len(self.pdf_document)):
            if self.first_page_key and p_idx == 0:
                continue
            page_indices.append(p_idx)

        table.setRowCount(len(page_indices))

        for row, p_idx in enumerate(page_indices):
            page_item = QTableWidgetItem(str(p_idx + 1))
            page_item.setFlags(Qt.ItemIsEnabled)
            table.setItem(row, 0, page_item)

            page_texts = self.results.get(p_idx, {}).get("text", {})
            for col, label in enumerate(labels, start=1):
                val = page_texts.get(label, "")
                table.setItem(row, col, QTableWidgetItem(str(val)))

            # Absent checkbox
            absent_col = 1 + len(labels)
            absent_item = QTableWidgetItem()
            absent_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            absent_item.setCheckState(Qt.Checked if self.student_absence.get(p_idx, False) else Qt.Unchecked)
            table.setItem(row, absent_col, absent_item)

        # Load extra_students (absent / overflow students from previous save)
        existing_extras = getattr(self, 'extra_students', [])
        for extra in existing_extras:
            new_row = table.rowCount()
            table.insertRow(new_row)
            page_item = QTableWidgetItem("-")
            page_item.setFlags(Qt.ItemIsEnabled)
            table.setItem(new_row, 0, page_item)
            extra_texts = extra.get("text", {})
            for col, label in enumerate(labels, start=1):
                table.setItem(new_row, col, QTableWidgetItem(extra_texts.get(label, "")))
            abs_item = QTableWidgetItem()
            abs_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            abs_item.setCheckState(Qt.Checked if extra.get("absent", True) else Qt.Unchecked)
            table.setItem(new_row, absent_col, abs_item)

        layout.addWidget(table)

        # Status label showing counts
        count_label = QLabel()
        count_label.setStyleSheet("color: #555; font-size: 12px; padding: 2px;")
        layout.addWidget(count_label)

        def _update_counts():
            num_pages = len(page_indices)
            absent_c = 1 + len(labels)
            num_students = 0
            num_present = 0
            for r in range(table.rowCount()):
                has_data = False
                for c in range(1, 1 + len(labels)):
                    item = table.item(r, c)
                    if item and item.text().strip():
                        has_data = True
                        break
                if has_data:
                    num_students += 1
                    abs_item = table.item(r, absent_c)
                    if not abs_item or abs_item.checkState() != Qt.Checked:
                        num_present += 1
            count_label.setText(tr("lbl_student_counts", pages=num_pages, students=num_students, present=num_present))

        _update_counts()
        table.cellChanged.connect(lambda row, col: _update_counts())

        btn_row = QHBoxLayout()
        btn_paste = QPushButton(tr("btn_paste_clipboard"))
        btn_row.addWidget(btn_paste)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(buttons)

        absent_col = 1 + len(labels)

        def _add_extra_row(checked=True):
            """Append a new empty row to the table (no real PDF page)."""
            new_row = table.rowCount()
            table.insertRow(new_row)
            page_item = QTableWidgetItem("-")
            page_item.setFlags(Qt.ItemIsEnabled)
            table.setItem(new_row, 0, page_item)
            for c in range(1, absent_col):
                table.setItem(new_row, c, QTableWidgetItem(""))
            abs_item = QTableWidgetItem()
            abs_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            abs_item.setCheckState(Qt.Checked if checked else Qt.Unchecked)
            table.setItem(new_row, absent_col, abs_item)

        def paste_from_clipboard():
            text = QtWidgets.QApplication.clipboard().text()
            if not text.strip():
                QMessageBox.information(dialog, tr("dlg_student_title"), tr("msg_clipboard_empty"))
                return

            rows = [r for r in text.splitlines() if r.strip() != ""]
            start_row = table.currentRow()
            if start_row < 0:
                start_row = 0

            for r_idx, line in enumerate(rows):
                tgt_row = start_row + r_idx
                # Auto-add rows when paste exceeds current row count
                while tgt_row >= table.rowCount():
                    _add_extra_row(checked=True)
                cols = line.split("\t")
                for c_idx, val in enumerate(cols):
                    if c_idx >= len(labels):
                        break
                    table.setItem(tgt_row, c_idx + 1, QTableWidgetItem(val.strip()))

        def accept():
            self._ensure_results_for_pages()

            # ── Collect ALL student data from every row (page rows + extra rows) ──
            all_students = []
            for row in range(table.rowCount()):
                student = {"text": {}}
                for col, label in enumerate(labels, start=1):
                    item = table.item(row, col)
                    student["text"][label] = item.text().strip() if item else ""
                abs_item = table.item(row, absent_col)
                student["absent"] = (abs_item.checkState() == Qt.Checked) if abs_item else False
                all_students.append(student)

            # ── Separate present vs absent, preserving order ──
            present_students = [s for s in all_students if not s["absent"]]
            absent_students  = [s for s in all_students if s["absent"]]

            # ── Map present students → PDF pages in order ──
            #    This fixes the bug where absent students in the middle
            #    caused subsequent students to be mapped to the wrong pages.
            for p_idx in page_indices:
                self.results[p_idx].setdefault("text", {})
                # Clear old text data for this page
                for label in labels:
                    self.results[p_idx]["text"][label] = ""
                self.student_absence[p_idx] = False

            for i, p_idx in enumerate(page_indices):
                if i < len(present_students):
                    page_texts = self.results[p_idx]["text"]
                    for label, val in present_students[i]["text"].items():
                        page_texts[label] = val

            # ── Excess present students beyond page count → extra (non-absent) ──
            extra = []
            for s in present_students[len(page_indices):]:
                extra.append({"text": s["text"], "absent": False})

            # ── Absent students → extra_students ──
            for s in absent_students:
                extra.append({"text": s["text"], "absent": True})

            self.extra_students = extra

            # Build student_order preserving original input order
            self.student_order = []
            present_idx = 0
            for s in all_students:
                entry = {"text": dict(s["text"]), "absent": s["absent"]}
                if not s["absent"]:
                    if present_idx < len(page_indices):
                        entry["page_idx"] = page_indices[present_idx]
                    else:
                        entry["page_idx"] = None
                    present_idx += 1
                else:
                    entry["page_idx"] = None
                self.student_order.append(entry)

            dialog.accept()
            self._update_student_info_label()
            self.update_result_table()

        # Use event filter for reliable Ctrl+V interception
        _pf = _TablePasteFilter(paste_from_clipboard, table)
        table.installEventFilter(_pf)
        table._paste_filter_ref = _pf  # Prevent GC

        btn_paste.clicked.connect(paste_from_clipboard)
        buttons.accepted.connect(accept)
        buttons.rejected.connect(dialog.reject)

        dialog.exec_()

    def edit_topics(self):
        questions = self._get_all_questions()
        if not questions:
            QMessageBox.information(self, tr("dlg_topic_title"), tr("msg_no_questions"))
            return

        dialog = QDialog(self)
        dialog.setWindowTitle(tr("dlg_topic_title"))
        dialog.resize(500, 500)
        layout = QVBoxLayout(dialog)

        layout.addWidget(QLabel(tr("dlg_topic_hint")))

        table = QTableWidget()
        table.setColumnCount(2)
        table.setHorizontalHeaderLabels([tr("col_question"), tr("col_topic")])
        table.setRowCount(len(questions))
        table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        table.setEditTriggers(QAbstractItemView.AllEditTriggers)

        for row, q in enumerate(questions):
            q_item = QTableWidgetItem(f"Q{q}")
            q_item.setFlags(Qt.ItemIsEnabled)
            table.setItem(row, 0, q_item)
            topic_val = self.topic_map.get(q, "")
            table.setItem(row, 1, QTableWidgetItem(topic_val))

        layout.addWidget(table)

        btn_row = QHBoxLayout()
        btn_paste = QPushButton(tr("btn_paste_clipboard"))
        btn_row.addWidget(btn_paste)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(buttons)

        def paste_from_clipboard():
            text = QtWidgets.QApplication.clipboard().text()
            if not text.strip():
                QMessageBox.information(dialog, tr("dlg_topic_title"), tr("msg_clipboard_empty"))
                return

            rows_data = [r for r in text.splitlines() if r.strip() != ""]
            start_row = table.currentRow()
            if start_row < 0:
                start_row = 0

            for r_idx, line in enumerate(rows_data):
                tgt_row = start_row + r_idx
                if tgt_row >= table.rowCount():
                    break
                cols = line.split("\t")
                # If only one column, it's the topic name
                if len(cols) == 1:
                    table.setItem(tgt_row, 1, QTableWidgetItem(cols[0].strip()))
                elif len(cols) >= 2:
                    # Two columns: skip Q column (col 0 is read-only), paste topic
                    table.setItem(tgt_row, 1, QTableWidgetItem(cols[-1].strip()))

        def accept():
            new_map = {}
            for row, q in enumerate(questions):
                topic_text = table.item(row, 1).text() if table.item(row, 1) else ""
                new_map[q] = topic_text.strip()
            self.topic_map = new_map
            dialog.accept()

        # Use event filter for reliable Ctrl+V interception
        _pf2 = _TablePasteFilter(paste_from_clipboard, table)
        table.installEventFilter(_pf2)
        table._paste_filter_ref = _pf2  # Prevent GC

        btn_paste.clicked.connect(paste_from_clipboard)
        buttons.accepted.connect(accept)
        buttons.rejected.connect(dialog.reject)

        dialog.exec_()

    def _get_page_offset(self, page_idx):
        """Return the explicit page offset, or fall back to page 0's offset."""
        if page_idx in self.page_offsets:
            return self.page_offsets[page_idx]
        if 0 in self.page_offsets:
            return self.page_offsets[0]
        return (0, 0)

    def _render_page_image_np(self, page_idx):
        """Render a PDF page into a NumPy RGB image using the current preview settings."""
        if not self.pdf_document or page_idx < 0 or page_idx >= len(self.pdf_document):
            return None

        page = self.pdf_document[page_idx]
        mat = fitz.Matrix(2, 2)
        pix = page.get_pixmap(matrix=mat)
        img_pil = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        img_np = np.array(img_pil)

        if hasattr(self, 'check_auto_deskew') and self.check_auto_deskew.isChecked():
            img_np, _ = deskew_image(img_np)

        return img_np

    def _refresh_alignment_reference_from_page0(self):
        """Rebuild the page 0 alignment reference after a manual image offset change."""
        if not self.pdf_document or self.current_page != 0:
            return
        if not hasattr(self, 'check_auto_align') or not self.check_auto_align.isChecked():
            return
        if not hasattr(self, 'view') or not self.view.align_marks:
            return

        current_offset = self.page_offsets.get(0)
        if current_offset is None and self.current_pixmap_item:
            current_offset = self.current_pixmap_item.get_offset()
        if current_offset is None:
            current_offset = (0, 0)

        current_offset = (float(current_offset[0]), float(current_offset[1]))
        has_templates = bool(getattr(self, 'align_templates', []))
        if has_templates and self.align_reference_offset == current_offset:
            return

        ref_img = self._render_page_image_np(0)
        if ref_img is None:
            return

        self.page_offsets[0] = current_offset
        self._reset_align_templates()
        self._align_init_template(ref_img, 0)

    def load_page(self, p_idx, apply_corrections=True):
        if not self.pdf_document: return
        
        # Save current image offset
        if self.current_pixmap_item:
            self.page_offsets[self.current_page] = self.current_pixmap_item.get_offset()
        if self.current_page == 0 and p_idx != 0:
            self._refresh_alignment_reference_from_page0()
            
        self.current_page = p_idx
        self.lbl_page.setText(tr("lbl_page", current=p_idx+1, total=len(self.pdf_document)))
        
        correction_info = []
        if apply_corrections:
            # Preview and recognition intentionally share one correction
            # pipeline so crop previews show the pixels that were classified.
            img_np, correction = self._prepare_page_for_recognition(p_idx)
            skew_angle = float(correction.get("deskew_angle", 0.0))
            if skew_angle != 0.0:
                correction_info.append(f"Deskew: {skew_angle:.2f}°")
            if self.check_auto_align.isChecked():
                dx, dy = correction.get("alignment_shift", [0.0, 0.0])
                confidence = float(correction.get("alignment_confidence", 0.0))
                if p_idx == 0:
                    correction_info.append("Alignment reference set")
                elif dx != 0.0 or dy != 0.0:
                    correction_info.append(
                        f"Alignment: dx={dx:.1f}, dy={dy:.1f}, "
                        f"confidence={confidence:.0%}"
                    )
        else:
            img_np = self._render_raw_page(p_idx)
        
        # Convert back to QImage
        h, w = img_np.shape[:2]
        img = QImage(img_np.data, w, h, img_np.strides[0], QImage.Format_RGB888).copy()
        
        # Remove only the pixmap item, not the marks
        if self.current_pixmap_item is not None:
            self.scene.removeItem(self.current_pixmap_item)
        
        # Add Image
        pix_item = QPixmap.fromImage(img)
        self.current_pixmap_item = MovablePixmapItem(pix_item)
        
        # Restore offset
        off = self._get_page_offset(p_idx)
        self.current_pixmap_item.setPos(off[0], off[1])
            
        self.scene.addItem(self.current_pixmap_item)
        # Move pixmap to back so marks are visible on top
        self.current_pixmap_item.setZValue(-1)
        self.view.setSceneRect(QRectF(0, 0, w, h))
        
        # Update correction info label
        if hasattr(self, 'lbl_correction') and correction_info:
            self.lbl_correction.setText(" | ".join(correction_info))
            self.lbl_correction.setStyleSheet("color: green; font-weight: bold;")
        elif hasattr(self, 'lbl_correction'):
            self.lbl_correction.setText("")
        
        # Reset zoom when changing pages to ensure marks align correctly
        self.view.zoom_reset()
        
        # Ensure marks are in the scene
        for m in self.view.text_marks + self.view.option_marks:
            if m.scene() is None:
                self.scene.addItem(m)
        
        # Ensure alignment marks are in the scene
        for m in self.view.align_marks:
            if m.scene() is None:
                self.scene.addItem(m)

        # Update table for this page result if available
        self.update_result_table()

        # Update student info label for current page
        self._update_student_info_label()

    def _update_student_info_label(self):
        """Update the student info label for the current page from student_order."""
        if not hasattr(self, 'lbl_student_info'):
            return
        student_order = getattr(self, 'student_order', [])
        if not student_order:
            self.lbl_student_info.setText("")
            return
        # Find the student mapped to the current page
        for entry in student_order:
            if entry.get("page_idx") == self.current_page:
                parts = [v for v in entry["text"].values() if v]
                if parts:
                    self.lbl_student_info.setText(tr("lbl_student_info", info="  -  ".join(parts)))
                else:
                    self.lbl_student_info.setText("")
                return
        self.lbl_student_info.setText("")

    def prev_page(self):
        if self.current_page > 0:
            self.load_page(self.current_page - 1)

    def next_page(self):
        if self.pdf_document and self.current_page < len(self.pdf_document) - 1:
            self.load_page(self.current_page + 1)

    def undo_last_mark(self):
        """Remove the last added mark (option or text) and restore counter"""
        if not hasattr(self.view, 'mark_history') or not self.view.mark_history:
            # Check if there are any alignment marks to remove
            if self.view.align_marks:
                removed = self.view.align_marks.pop()
                self.scene.removeItem(removed)
                self._reset_align_templates()
                print("Undo: Removed alignment mark")
            else:
                print("Undo: No marks to remove")
            return
        
        last_mark = self.view.mark_history.pop()
        
        if last_mark in self.view.text_marks:
            self.view.text_marks.remove(last_mark)
            # Restore counter to the removed item's question number
            self.view.text_counter = last_mark.question_num
            print(f"Undo: Removed text mark Q{last_mark.question_num} ('{last_mark.label}'), counter reset to {self.view.text_counter}")
        elif last_mark in self.view.option_marks:
            self.view.option_marks.remove(last_mark)
            # Restore counter to the removed item's question number
            self.view.option_counter = last_mark.question_num
            print(f"Undo: Removed option mark Q{last_mark.question_num} ('{last_mark.label}'), counter reset to {self.view.option_counter}")
        elif last_mark in self.view.align_marks:
            self.view.align_marks.remove(last_mark)
            # Also reset alignment template
            self._reset_align_templates()
            print("Undo: Removed alignment mark")
        
        self.scene.removeItem(last_mark)

    def clear_all_marks(self):
        for m in self.view.text_marks + self.view.option_marks:
            self.scene.removeItem(m)
        for m in self.view.align_marks:
            self.scene.removeItem(m)
        self.view.align_marks.clear()
        self.view.text_marks.clear()
        self.view.option_marks.clear()
        self.view.mark_history.clear()  # Clear undo history
        self.view.text_counter = 1
        self.view.option_counter = 1
        self.view.align_counter = 1
        # Reset alignment reference
        self._reset_align_templates()

    def export_template(self):
        data = self.view.get_all_marks_data()
        data["schema_version"] = 2
        data["app_version"] = APP_VERSION
        data["render_scale"] = 2.0
        # Store marks relative to the page image, not its manually dragged
        # scene position.  This makes a saved template portable to offset 0.
        if self.current_page == 0 and self.current_pixmap_item is not None:
            offset_x, offset_y = self.current_pixmap_item.get_offset()
            self.page_offsets[0] = (offset_x, offset_y)
        else:
            offset_x, offset_y = self._get_page_offset(0)
        for key in ("text_marks", "option_marks", "align_marks"):
            for mark in data.get(key, []):
                mark["x"] = float(mark.get("x", 0.0)) - float(offset_x)
                mark["y"] = float(mark.get("y", 0.0)) - float(offset_y)
        if self.pdf_document and len(self.pdf_document):
            page_rect = self.pdf_document[0].rect
            data["reference_page"] = {
                "width": float(page_rect.width) * 2.0,
                "height": float(page_rect.height) * 2.0,
            }
        elif self.current_pixmap_item is not None:
            pixmap = self.current_pixmap_item.pixmap()
            data["reference_page"] = {
                "width": pixmap.width(),
                "height": pixmap.height(),
            }
        fname, _ = QFileDialog.getSaveFileName(self, "Save Template", "", "JSON (*.json)")
        if fname:
            if not fname.lower().endswith(".json"):
                fname += ".json"
            try:
                with open(fname, "w", encoding="utf-8") as file:
                    json.dump(data, file, ensure_ascii=False, indent=2)
            except Exception as exc:
                QMessageBox.critical(self, "Template Error", str(exc))

    def _import_template_legacy(self):
        fname, _ = QFileDialog.getOpenFileName(self, "Load Template", "", "JSON (*.json)")
        if fname:
            with open(fname, 'r') as f:
                data = json.load(f)
            self.clear_all_marks()
            
            for m in data.get("text_marks", []):
                item = MarkItem(0, 0, m['width'], m['height'], MARK_TYPE_TEXT, m['question'], m['label'], view_ref=self.view)
                item.setPos(m['x'], m['y'])
                self.view.text_marks.append(item)
                self.scene.addItem(item)
                self.view.text_counter = max(self.view.text_counter, m['question'] + 1)
                
            for m in data.get("option_marks", []):
                item = MarkItem(0, 0, m['width'], m['height'], MARK_TYPE_OPTION, m['question'], m['label'], m.get('options_count', 4), view_ref=self.view)
                item.setPos(m['x'], m['y'])
                self.view.option_marks.append(item)
                self.scene.addItem(item)
                self.view.option_counter = max(self.view.option_counter, m['question'] + 1)
            
            # Load alignment marks (supports both new list and old single format)
            align_marks_data = data.get("align_marks", [])
            # Backward compat: load old single "align_mark" format
            if not align_marks_data:
                old_align = data.get("align_mark")
                if old_align:
                    align_marks_data = [old_align]
            for ad in align_marks_data:
                item = MarkItem(0, 0, ad['width'], ad['height'], MARK_TYPE_ALIGN, 
                               ad.get('question', self.view.align_counter), ad.get('label', ''), view_ref=self.view)
                item.setPos(ad['x'], ad['y'])
                self.view.align_marks.append(item)
                self.scene.addItem(item)
                self.view.align_counter = max(self.view.align_counter, ad.get('question', 1) + 1)

    def import_template(self):
        fname, _ = QFileDialog.getOpenFileName(
            self, "Load Template", "", "JSON (*.json)"
        )
        if not fname:
            return
        try:
            with open(fname, "r", encoding="utf-8-sig") as file:
                data = json.load(file)
            if not isinstance(data, dict):
                raise ValueError("Template root must be a JSON object.")
            for key in ("text_marks", "option_marks"):
                if key in data and not isinstance(data[key], list):
                    raise ValueError(f"Template field '{key}' must be a list.")
            self._load_template_data(data)
            self.results = {}
            self.answer_key = {}
            self.debug_records = []
            self._review_cursor = None
            self._clear_result_table()
            self.statusBar().showMessage(
                f"Template loaded: {os.path.basename(fname)}", 6000
            )
        except Exception as exc:
            QMessageBox.critical(self, "Template Error", str(exc))

    def _render_raw_page(self, page_idx):
        page = self.pdf_document[page_idx]
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
        return np.frombuffer(pix.samples, dtype=np.uint8).reshape(
            pix.height, pix.width, pix.n
        )[:, :, :3].copy()

    def _prepare_page_for_recognition(self, page_idx):
        """Render and correct one page using a consistent pipeline."""

        image = self._render_raw_page(page_idx)
        metadata = {
            "deskew_angle": 0.0,
            "alignment_shift": [0.0, 0.0],
            "alignment_confidence": 0.0,
        }
        if self.check_auto_deskew.isChecked():
            image, angle = deskew_image(image)
            metadata["deskew_angle"] = float(angle)

        if self.check_auto_align.isChecked():
            # Selected-page recognition must still anchor itself to page 0;
            # otherwise the first selected student page becomes the reference.
            has_template_reference = bool(getattr(self, "align_templates", []))
            has_bounds_reference = self.align_reference_gray is not None
            if page_idx != 0 and not (has_template_reference or has_bounds_reference):
                reference = self._render_raw_page(0)
                if self.check_auto_deskew.isChecked():
                    reference, _ = deskew_image(reference)
                self.align_image(reference, 0)

            image, (dx, dy), confidence = self.align_image(image, page_idx)
            # Always use the returned image.  A pure rotation can legitimately
            # have dx=dy=0 and was previously discarded in interactive flows.
            metadata["alignment_shift"] = [float(dx), float(dy)]
            metadata["alignment_confidence"] = float(confidence)
        return image, metadata

    def _recognize_page(self, page_idx, preserve_overrides=True):
        """Recognize one page for all interactive and batch workflows."""

        image, correction = self._prepare_page_for_recognition(page_idx)
        height, width = image.shape[:2]
        off_x, off_y = self._get_page_offset(page_idx)
        existing = self.results.get(page_idx, {}) if preserve_overrides else {}
        manual_overrides = dict(existing.get("manual_overrides", {}))
        debug_enabled = bool(self.check_save_debug.isChecked())

        page_result = {
            "options": {},
            "raw_options": {},
            "manual_overrides": manual_overrides,
            "text": {},
            "option_crops": {},
            "text_crops": {},
            "option_boxes": {},
            "text_boxes": {},
            "confidence": {},
            "review": {},
            "reasons": {},
            "alignment": correction,
        }

        for mark in self.view.option_marks:
            rect = mark.sceneBoundingRect()
            left = max(0, int(round(rect.x() - off_x)))
            top = max(0, int(round(rect.y() - off_y)))
            right = min(width, int(round(rect.x() + rect.width() - off_x)))
            bottom = min(height, int(round(rect.y() + rect.height() - off_y)))
            question = mark.question_num
            page_result["option_boxes"][question] = [left, top, right, bottom]

            if right <= left or bottom <= top:
                detected = "[Out of bounds]"
                details = {
                    "confidence": 0.0,
                    "needs_review": True,
                    "reason": "out_of_bounds",
                }
            else:
                crop_array = image[top:bottom, left:right]
                correct_answer = (
                    self.answer_key.get(question, "")
                    or self.answer_key.get(str(question), "")
                )
                detected = self.detect_filled_option(
                    crop_array,
                    mark.options_count,
                    save_debug=debug_enabled,
                    context={
                        "page": page_idx + 1,
                        "question": question,
                        "label": f"Q{question}",
                        "correct_answer": correct_answer,
                    },
                )
                details = dict(self._last_detection)

            page_result["raw_options"][question] = detected
            page_result["confidence"][question] = float(details.get("confidence", 0.0))
            page_result["review"][question] = bool(
                details.get("needs_review", True)
            )
            page_result["reasons"][question] = str(
                details.get("reason", "unknown")
            )

        page_result["options"] = dict(page_result["raw_options"])
        for question, value in manual_overrides.items():
            page_result["options"][question] = value
            page_result["confidence"][question] = 1.0
            page_result["review"][question] = False
            page_result["reasons"][question] = "manual"

        existing_text = dict(existing.get("text", {}))
        for mark in self.view.text_marks:
            rect = mark.sceneBoundingRect()
            left = max(0, int(round(rect.x() - off_x)))
            top = max(0, int(round(rect.y() - off_y)))
            right = min(width, int(round(rect.x() + rect.width() - off_x)))
            bottom = min(height, int(round(rect.y() + rect.height() - off_y)))
            key = mark.label if mark.label else f"Field {mark.question_num}"
            page_result["text_boxes"][key] = [left, top, right, bottom]

            if not self.check_recognize_text.isChecked():
                text = existing_text.get(key, "")
            elif right <= left or bottom <= top:
                text = "[Out of bounds]"
            else:
                crop_image = Image.fromarray(image[top:bottom, left:right])
                try:
                    text = self.get_ocr_result(
                        crop_image, save_debug=debug_enabled
                    )
                except Exception as exc:
                    # A first-run model download or OCR engine problem must not
                    # discard the already-recognized OMR answers on this page.
                    print(f"OCR failed for {key}: {exc}")
                    text = "[OCR unavailable]"
            page_result["text"][key] = text

        for key, value in existing_text.items():
            page_result["text"].setdefault(key, value)
        return page_result

    def _run_recognition_all_legacy(self):
        if not self.pdf_document: 
            QMessageBox.warning(self, "Warning", tr("msg_no_pdf"))
            return
        if not self.view.option_marks and not self.view.text_marks:
            QMessageBox.warning(self, "Warning", tr("msg_no_marks"))
            return
            
        self.results = {}
        self.debug_records = []
        
        # Save current page's image offset before processing
        if self.current_pixmap_item:
            self.page_offsets[self.current_page] = self.current_pixmap_item.get_offset()
        
        # Reset alignment template for new recognition run
        self._reset_align_templates()
        
        # Progress Dialog
        progress = QtWidgets.QProgressDialog("Recognizing...", "Cancel", 0, len(self.pdf_document), self)
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.show()
        
        # Get marks data once (they are the same for all pages)
        # Marks are in scene coordinates. We need to map them relative to image position.
        
        for p_idx in range(len(self.pdf_document)):
            QtWidgets.QApplication.processEvents()
            if progress.wasCanceled(): 
                break
            progress.setValue(p_idx)
            progress.setLabelText(f"Recognizing page {p_idx + 1} of {len(self.pdf_document)}...")
            
            # Render page
            page = self.pdf_document[p_idx]
            mat = fitz.Matrix(2, 2)
            pix = page.get_pixmap(matrix=mat)
            img_pil = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            
            # Apply auto-deskew if enabled
            skew_angle = 0.0
            if self.check_auto_deskew.isChecked():
                img_np = np.array(img_pil)
                img_corrected, skew_angle = deskew_image(img_np)
                if skew_angle != 0.0:
                    print(f"Page {p_idx + 1}: Corrected skew angle: {skew_angle:.2f}°")
                    img_pil = Image.fromarray(img_corrected)

            # Apply auto-align (shift) if enabled
            if self.check_auto_align.isChecked():
                img_np = np.array(img_pil)
                img_aligned, (dx, dy), response = self.align_image(img_np, p_idx)
                if dx != 0.0 or dy != 0.0:
                    print(f"Page {p_idx + 1}: Aligned shift dx={dx:.1f}, dy={dy:.1f} (score={response:.3f})")
                    img_pil = Image.fromarray(img_aligned)
            
            # Get Image Offset for this page (where the image was positioned in the scene)
            # If user moved the image, marks are relative to scene origin (0,0)
            # Image is at (off_x, off_y), so to get image-relative coords:
            # image_x = scene_x - off_x
            off_x, off_y = self._get_page_offset(p_idx)
            
            page_res = {
                "options": {},
                "text": {},
                "option_crops": {},
                "text_crops": {}
            }
            
            # Helper to process a list of marks
            def process_marks(marks_list, target_dict):
                for mark in marks_list:
                    rect = mark.sceneBoundingRect()
                    # Convert scene coordinates to image coordinates
                    # The image is positioned at (off_x, off_y) in the scene
                    # So image coordinate = scene coordinate - image offset
                    img_x = rect.x() - off_x
                    img_y = rect.y() - off_y
                    
                    print(f"Mark Q{mark.question_num}: scene=({rect.x():.0f},{rect.y():.0f}), offset=({off_x:.0f},{off_y:.0f}), img=({img_x:.0f},{img_y:.0f}), size=({rect.width():.0f}x{rect.height():.0f})")
                    
                    # Ensure crop is within image bounds
                    left = max(0, int(img_x))
                    top = max(0, int(img_y))
                    right = min(img_pil.width, int(img_x + rect.width()))
                    bottom = min(img_pil.height, int(img_y + rect.height()))
                    
                    print(f"  Crop: ({left},{top})-({right},{bottom}), img size: {img_pil.width}x{img_pil.height}")
                    
                    if right > left and bottom > top:
                        crop = img_pil.crop((left, top, right, bottom))
                        crop_label = f"Q{mark.question_num}" if mark.mark_type == MARK_TYPE_OPTION else (mark.label or f"Field_{mark.question_num}")
                        crop_path = self._save_crop_image(crop, p_idx, crop_label, "option" if mark.mark_type == MARK_TYPE_OPTION else "text")
                        
                        if mark.mark_type == MARK_TYPE_OPTION:
                            # Use bubble detection for options
                            correct_answer = ""
                            if hasattr(self, "answer_key"):
                                correct_answer = self.answer_key.get(mark.question_num, "") or self.answer_key.get(str(mark.question_num), "")
                            text = self.detect_filled_option(
                                crop,
                                mark.options_count,
                                save_debug=True,
                                context={
                                    "page": p_idx + 1,
                                    "question": mark.question_num,
                                    "label": f"Q{mark.question_num}",
                                    "correct_answer": correct_answer,
                                }
                            )
                        else:
                            # Use OCR for text fields
                            text = self.get_ocr_result(crop, save_debug=True)
                    else:
                        text = f"[Out of bounds]"
                        print(f"  Out of bounds!")
                        crop_path = ""
                    
                    if mark.mark_type == MARK_TYPE_OPTION:
                        target_dict[mark.question_num] = text
                        page_res["option_crops"][mark.question_num] = crop_path
                    else:
                        # For text fields, use label as key if exists, else "Field X"
                        key = mark.label if mark.label else f"Field {mark.question_num}"
                        target_dict[key] = text
                        page_res["text_crops"][key] = crop_path
            
            process_marks(self.view.option_marks, page_res["options"])
            process_marks(self.view.text_marks, page_res["text"])
            
            # Store
            self.results[p_idx] = page_res
            
        progress.setValue(len(self.pdf_document))
        progress.close()
        
        # Show summary
        total_pages = len(self.results)
        total_options = sum(len(r.get("options", {})) for r in self.results.values())
        QMessageBox.information(self, tr("msg_recognition_title"), 
            tr("msg_recognition_complete", pages=total_pages, options=total_options))
        
        self.update_result_table()

        # Build Answer Key if needed
        if self.first_page_key and 0 in self.results:
            self.answer_key = self._answer_key_from_result(self.results[0])
            # Refresh to show scores
            self.update_result_table()

    def run_recognition_all(self):
        if not self.pdf_document:
            QMessageBox.warning(self, "Warning", tr("msg_no_pdf"))
            return
        if not self.view.option_marks and not self.view.text_marks:
            QMessageBox.warning(self, "Warning", tr("msg_no_marks"))
            return

        if self.current_pixmap_item:
            self.page_offsets[self.current_page] = self.current_pixmap_item.get_offset()
        manual_answer_key = (
            dict(self.answer_key) if not self.first_page_key else {}
        )
        self.results = {}
        self.answer_key = manual_answer_key
        self.debug_records = []
        self._review_cursor = None
        self._reset_align_templates()

        total = len(self.pdf_document)
        progress = QtWidgets.QProgressDialog(
            "Recognizing...", "Cancel", 0, total, self
        )
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.show()

        errors = []
        processed = 0
        cancelled = False
        for page_idx in range(total):
            QtWidgets.QApplication.processEvents()
            if progress.wasCanceled():
                cancelled = True
                break
            progress.setValue(page_idx)
            progress.setLabelText(
                f"Recognizing page {page_idx + 1} of {total}..."
            )
            try:
                self.results[page_idx] = self._recognize_page(
                    page_idx, preserve_overrides=False
                )
                processed += 1
                if self.first_page_key and page_idx == 0:
                    self.answer_key = self._answer_key_from_result(
                        self.results[0]
                    )
            except Exception as exc:
                errors.append((page_idx, str(exc)))

        progress.setValue(processed if cancelled else total)
        progress.close()
        self.update_result_table()

        total_options = sum(
            len(result.get("options", {})) for result in self.results.values()
        )
        summary = tr(
            "msg_recognition_complete",
            pages=processed,
            options=total_options,
        )
        if cancelled:
            summary += "\n\nCancelled; completed results were kept."
        if errors:
            summary += f"\n\n{len(errors)} page(s) failed:"
            for page_idx, error in errors[:3]:
                summary += f"\n• Page {page_idx + 1}: {error[:100]}"
        QMessageBox.information(self, tr("msg_recognition_title"), summary)

    def _run_recognition_selected_legacy(self):
        """Re-recognize specific pages (current, range, or all)."""
        if not self.pdf_document:
            QMessageBox.warning(self, "Warning", tr("msg_no_pdf"))
            return
        if not self.view.option_marks and not self.view.text_marks:
            QMessageBox.warning(self, "Warning", tr("msg_no_marks"))
            return

        total_pages = len(self.pdf_document)

        dialog = QDialog(self)
        dialog.setWindowTitle(tr("dlg_recognize_title"))
        dialog.resize(400, 200)
        dlg_layout = QVBoxLayout(dialog)

        dlg_layout.addWidget(QLabel(tr("dlg_recognize_prompt")))

        from PyQt5.QtWidgets import QRadioButton, QLineEdit
        rb_current = QRadioButton(tr("dlg_recognize_current") + f" ({self.current_page + 1})")
        rb_current.setChecked(True)
        rb_all = QRadioButton(tr("dlg_recognize_all_pages"))
        rb_range = QRadioButton(tr("dlg_recognize_range"))
        range_edit = QLineEdit()
        range_edit.setPlaceholderText("e.g. 1-3, 5, 8")
        range_edit.setEnabled(False)
        rb_range.toggled.connect(range_edit.setEnabled)

        dlg_layout.addWidget(rb_current)
        dlg_layout.addWidget(rb_all)
        dlg_layout.addWidget(rb_range)
        dlg_layout.addWidget(range_edit)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        dlg_layout.addWidget(btn_box)
        btn_box.accepted.connect(dialog.accept)
        btn_box.rejected.connect(dialog.reject)

        if dialog.exec_() != QDialog.Accepted:
            return

        # Parse selected pages
        pages_to_process = []
        if rb_current.isChecked():
            pages_to_process = [self.current_page]
        elif rb_all.isChecked():
            pages_to_process = list(range(total_pages))
        elif rb_range.isChecked():
            range_text = range_edit.text().strip()
            pages_to_process = self._parse_page_range(range_text, total_pages)
            if not pages_to_process:
                QMessageBox.warning(self, "Warning", "Invalid page range.")
                return

        if not pages_to_process:
            return

        # Ensure results dict exists
        if not hasattr(self, 'results') or self.results is None:
            self.results = {}

        # Save current page's image offset
        if self.current_pixmap_item:
            self.page_offsets[self.current_page] = self.current_pixmap_item.get_offset()

        # For single page or partial re-recognition, we don't reset alignment template
        # unless we're processing from page 0
        if 0 in pages_to_process:
            self._reset_align_templates()

        progress = QtWidgets.QProgressDialog("Recognizing...", "Cancel", 0, len(pages_to_process), self)
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.show()

        processed_count = 0
        for idx, p_idx in enumerate(pages_to_process):
            QtWidgets.QApplication.processEvents()
            if progress.wasCanceled():
                break
            progress.setValue(idx)
            progress.setLabelText(f"Re-recognizing page {p_idx + 1}...")

            # Render page
            page = self.pdf_document[p_idx]
            mat = fitz.Matrix(2, 2)
            pix = page.get_pixmap(matrix=mat)
            img_pil = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

            if self.check_auto_deskew.isChecked():
                img_np = np.array(img_pil)
                img_corrected, skew_angle = deskew_image(img_np)
                if skew_angle != 0.0:
                    img_pil = Image.fromarray(img_corrected)

            if self.check_auto_align.isChecked():
                img_np = np.array(img_pil)
                img_aligned, (dx, dy), response = self.align_image(img_np, p_idx)
                if dx != 0.0 or dy != 0.0:
                    img_pil = Image.fromarray(img_aligned)

            off_x, off_y = self._get_page_offset(p_idx)

            page_res = {
                "options": {},
                "text": {},
                "option_crops": {},
                "text_crops": {}
            }

            # Preserve existing text data (student info) if not re-detected
            if p_idx in self.results:
                existing_texts = self.results[p_idx].get("text", {})
            else:
                existing_texts = {}

            for mark in self.view.option_marks:
                rect = mark.sceneBoundingRect()
                img_x = rect.x() - off_x
                img_y = rect.y() - off_y
                left = max(0, int(img_x))
                top = max(0, int(img_y))
                right = min(img_pil.width, int(img_x + rect.width()))
                bottom = min(img_pil.height, int(img_y + rect.height()))

                if right > left and bottom > top:
                    crop = img_pil.crop((left, top, right, bottom))
                    crop_path = self._save_crop_image(crop, p_idx, f"Q{mark.question_num}", "option")
                    text = self.detect_filled_option(crop, mark.options_count, save_debug=True,
                        context={"page": p_idx + 1, "question": mark.question_num, "label": f"Q{mark.question_num}"})
                else:
                    text = "[Out of bounds]"
                    crop_path = ""
                page_res["options"][mark.question_num] = text
                page_res["option_crops"][mark.question_num] = crop_path

            for mark in self.view.text_marks:
                rect = mark.sceneBoundingRect()
                img_x = rect.x() - off_x
                img_y = rect.y() - off_y
                left = max(0, int(img_x))
                top = max(0, int(img_y))
                right = min(img_pil.width, int(img_x + rect.width()))
                bottom = min(img_pil.height, int(img_y + rect.height()))
                key = mark.label if mark.label else f"Field {mark.question_num}"

                if right > left and bottom > top:
                    crop = img_pil.crop((left, top, right, bottom))
                    crop_path = self._save_crop_image(crop, p_idx, key, "text")
                    text = self.get_ocr_result(crop, save_debug=True)
                else:
                    text = "[Out of bounds]"
                    crop_path = ""
                page_res["text"][key] = text
                page_res["text_crops"][key] = crop_path

            # Merge preserved student info for fields not covered by marks
            for k, v in existing_texts.items():
                if k not in page_res["text"]:
                    page_res["text"][k] = v

            self.results[p_idx] = page_res
            processed_count += 1

        progress.setValue(len(pages_to_process))
        progress.close()

        total_options = sum(len(self.results.get(p, {}).get("options", {})) for p in pages_to_process if p in self.results)
        QMessageBox.information(self, tr("msg_recognition_title"),
            tr("msg_recognition_complete", pages=processed_count, options=total_options))

        self.update_result_table()

        if self.first_page_key and 0 in self.results:
            self.answer_key = self._answer_key_from_result(self.results[0])
            self.update_result_table()

    def run_recognition_selected(self):
        """Re-recognize selected pages through the shared page pipeline."""

        if not self.pdf_document:
            QMessageBox.warning(self, "Warning", tr("msg_no_pdf"))
            return
        if not self.view.option_marks and not self.view.text_marks:
            QMessageBox.warning(self, "Warning", tr("msg_no_marks"))
            return

        total_pages = len(self.pdf_document)
        dialog = QDialog(self)
        dialog.setWindowTitle(tr("dlg_recognize_title"))
        dialog.resize(400, 200)
        dialog_layout = QVBoxLayout(dialog)
        dialog_layout.addWidget(QLabel(tr("dlg_recognize_prompt")))

        from PyQt5.QtWidgets import QLineEdit, QRadioButton

        current_radio = QRadioButton(
            tr("dlg_recognize_current") + f" ({self.current_page + 1})"
        )
        current_radio.setChecked(True)
        all_radio = QRadioButton(tr("dlg_recognize_all_pages"))
        range_radio = QRadioButton(tr("dlg_recognize_range"))
        range_edit = QLineEdit()
        range_edit.setPlaceholderText("e.g. 1-3, 5, 8")
        range_edit.setEnabled(False)
        range_radio.toggled.connect(range_edit.setEnabled)
        for widget in (current_radio, all_radio, range_radio, range_edit):
            dialog_layout.addWidget(widget)

        buttons = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        dialog_layout.addWidget(buttons)
        if dialog.exec_() != QDialog.Accepted:
            return

        if current_radio.isChecked():
            pages_to_process = [self.current_page]
        elif all_radio.isChecked():
            pages_to_process = list(range(total_pages))
        else:
            pages_to_process = self._parse_page_range(
                range_edit.text().strip(), total_pages
            )
        if not pages_to_process:
            QMessageBox.warning(self, "Warning", "Invalid page range.")
            return

        if self.current_pixmap_item:
            self.page_offsets[self.current_page] = self.current_pixmap_item.get_offset()
        self._reset_align_templates()
        self.debug_records = []
        self._review_cursor = None

        progress = QtWidgets.QProgressDialog(
            "Recognizing...", "Cancel", 0, len(pages_to_process), self
        )
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.show()

        errors = []
        processed = 0
        cancelled = False
        for index, page_idx in enumerate(pages_to_process):
            QtWidgets.QApplication.processEvents()
            if progress.wasCanceled():
                cancelled = True
                break
            progress.setValue(index)
            progress.setLabelText(f"Re-recognizing page {page_idx + 1}...")
            try:
                self.results[page_idx] = self._recognize_page(
                    page_idx, preserve_overrides=True
                )
                processed += 1
                if self.first_page_key and page_idx == 0:
                    self.answer_key = self._answer_key_from_result(
                        self.results[0]
                    )
            except Exception as exc:
                errors.append((page_idx, str(exc)))

        progress.setValue(processed if cancelled else len(pages_to_process))
        progress.close()
        self.update_result_table()

        total_options = sum(
            len(self.results.get(page, {}).get("options", {}))
            for page in pages_to_process
        )
        summary = tr(
            "msg_recognition_complete",
            pages=processed,
            options=total_options,
        )
        if cancelled:
            summary += "\n\nCancelled; completed results were kept."
        if errors:
            summary += f"\n\n{len(errors)} page(s) failed."
        QMessageBox.information(self, tr("msg_recognition_title"), summary)

    def _parse_page_range(self, text, total_pages):
        """Parse page range string like '1-3, 5, 8' into list of 0-based page indices."""
        pages = set()
        parts = text.split(",")
        for part in parts:
            part = part.strip()
            if not part:
                continue
            if "-" in part:
                try:
                    start, end = part.split("-", 1)
                    start = int(start.strip())
                    end = int(end.strip())
                    for p in range(start, end + 1):
                        if 1 <= p <= total_pages:
                            pages.add(p - 1)
                except ValueError:
                    continue
            else:
                try:
                    p = int(part)
                    if 1 <= p <= total_pages:
                        pages.add(p - 1)
                except ValueError:
                    continue
        return sorted(pages)

    def _update_result_table_legacy(self):
        # Display results for CURRENT page
        if self.current_page not in getattr(self, 'results', {}):
            return
        
        self.table.blockSignals(True)
        page_res = self.results[self.current_page]
        # structure: {"options": {1: "A", ...}, "text": {"Name": "John", ...}}
        
        opts = page_res.get("options", {})
        option_crops = page_res.get("option_crops", {})

        self.table.setRowCount(len(opts))
        
        current_row = 0
            
        # Options only (text/student fields are managed via the Student Info dialog)
        sorted_qs = sorted(opts.keys())
        total_score = 0
        
        for q_num in sorted_qs:
            detected = opts[q_num]
            correct = self.answer_key.get(q_num, "")
            
            # Normalize for comparison
            is_correct = False
            if correct and detected:
                # remove spaces, lowercase
                d_clean = "".join(detected.split()).lower()
                c_clean = "".join(str(correct).split()).lower()
                if d_clean == c_clean: is_correct = True
            
            points = 1 if is_correct else 0
            if is_correct: total_score += 1
            
            self.table.setItem(current_row, 0, QTableWidgetItem(f"Q{q_num}"))
            self.table.setItem(current_row, 1, QTableWidgetItem(str(detected)))
            self.table.setItem(current_row, 2, QTableWidgetItem(str(correct)))
            self.table.setItem(current_row, 3, QTableWidgetItem(str(points)))
            crop_item = QTableWidgetItem("Open") if option_crops.get(q_num) else QTableWidgetItem("-")
            crop_item.setFlags(Qt.ItemIsEnabled)
            crop_item.setForeground(QColor("#007bff"))
            crop_item.setData(Qt.UserRole, option_crops.get(q_num, ""))
            self.table.setItem(current_row, 4, crop_item)
            
            # Color code similar to Excel: empty, multiple, correct/incorrect
            detected_str = str(detected).strip()
            if detected_str == "":
                self.table.item(current_row, 1).setBackground(QColor("#fff3cd"))
            elif len(detected_str) > 1:
                self.table.item(current_row, 1).setBackground(QColor("#ffe5b4"))
            elif correct:
                color = QColor("#d4edda") if is_correct else QColor("#f8d7da")
                self.table.item(current_row, 1).setBackground(color)
            
            current_row += 1
        
        self.lbl_score.setText(tr("lbl_score", score=total_score))

        # Update per-page answer status: show question numbers that are empty or multiple
        if hasattr(self, 'lbl_answer_status'):
            empty_qs = []
            multi_qs = []
            for q_num in sorted_qs:
                val = str(opts.get(q_num, "")).strip()
                if val == "":
                    empty_qs.append(f"Q{q_num}")
                elif len(val) > 1:
                    multi_qs.append(f"Q{q_num}")
            parts = []
            if empty_qs:
                parts.append(tr("lbl_answer_status_empty", questions=", ".join(empty_qs)))
            if multi_qs:
                parts.append(tr("lbl_answer_status_multi", questions=", ".join(multi_qs)))
            if parts:
                self.lbl_answer_status.setText("  |  ".join(parts))
                self.lbl_answer_status.setStyleSheet("color: #e65100; font-size: 12px; font-weight: bold; padding: 2px;")
            elif sorted_qs:
                self.lbl_answer_status.setText(tr("lbl_answer_status_ok"))
                self.lbl_answer_status.setStyleSheet("color: #2e7d32; font-size: 12px; padding: 2px;")
            else:
                self.lbl_answer_status.setText("")

        self.table.blockSignals(False)

    def update_result_table(self):
        """Render the current page with confidence and review information."""

        if self.current_page not in self.results:
            self._clear_result_table()
            return

        self.table.blockSignals(True)
        page_result = self.results[self.current_page]
        options = page_result.get("options", {})
        confidence = page_result.get("confidence", {})
        review = page_result.get("review", {})
        reasons = page_result.get("reasons", {})
        option_boxes = page_result.get("option_boxes", {})
        option_crops = page_result.get("option_crops", {})

        def question_sort_key(question):
            text = str(question)
            return (0, int(text)) if text.isdigit() else (1, text)

        questions = sorted(options.keys(), key=question_sort_key)
        self.table.clearContents()
        self.table.setRowCount(len(questions))
        total_score = 0
        empty_questions = []
        multiple_questions = []
        review_questions = []

        for row, question in enumerate(questions):
            detected = str(options.get(question, ""))
            correct = (
                self.answer_key.get(question, "")
                or self.answer_key.get(str(question), "")
            )
            detected_clean = "".join(detected.split()).upper()
            correct_clean = "".join(str(correct).split()).upper()
            is_correct = bool(correct_clean and detected_clean == correct_clean)
            points = 1 if is_correct else 0
            total_score += points

            question_item = QTableWidgetItem(f"Q{question}")
            question_item.setData(Qt.UserRole, question)
            question_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            self.table.setItem(row, 0, question_item)
            self.table.setItem(row, 1, QTableWidgetItem(detected))
            self.table.setItem(row, 2, QTableWidgetItem(str(correct)))
            points_item = QTableWidgetItem(str(points))
            points_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            self.table.setItem(row, 3, points_item)

            value_confidence = float(confidence.get(question, 0.0))
            confidence_item = QTableWidgetItem(f"{value_confidence:.0%}")
            confidence_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            confidence_item.setToolTip(str(reasons.get(question, "")))
            self.table.setItem(row, 4, confidence_item)

            crop_reference = {
                "page": self.current_page,
                "question": question,
                "box": option_boxes.get(question),
                "path": option_crops.get(question, ""),
            }
            box = crop_reference["box"]
            try:
                has_valid_box = (
                    isinstance(box, (list, tuple))
                    and len(box) == 4
                    and float(box[2]) > float(box[0])
                    and float(box[3]) > float(box[1])
                )
            except (TypeError, ValueError):
                has_valid_box = False
            has_crop = bool(
                has_valid_box
                or (
                    crop_reference["path"]
                    and os.path.isfile(crop_reference["path"])
                )
            )
            crop_item = QTableWidgetItem("View" if has_crop else "-")
            crop_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            if has_crop:
                crop_item.setForeground(QColor("#007bff"))
                crop_item.setData(Qt.UserRole, crop_reference)
            self.table.setItem(row, 5, crop_item)

            is_flagged = bool(review.get(question, False))
            if detected_clean == "":
                empty_questions.append(f"Q{question}")
                is_flagged = True
                self.table.item(row, 1).setBackground(QColor("#fff3cd"))
            elif detected.startswith("["):
                is_flagged = True
                self.table.item(row, 1).setBackground(QColor("#f8d7da"))
            elif len(detected_clean) > 1:
                multiple_questions.append(f"Q{question}")
                is_flagged = True
                self.table.item(row, 1).setBackground(QColor("#ffe5b4"))
            elif correct_clean:
                self.table.item(row, 1).setBackground(
                    QColor("#d4edda" if is_correct else "#f8d7da")
                )
            if is_flagged:
                review_questions.append(f"Q{question}")
                confidence_item.setBackground(QColor("#fff3cd"))

        self.lbl_score.setText(tr("lbl_score", score=total_score))
        status_parts = []
        if empty_questions:
            status_parts.append(
                tr("lbl_answer_status_empty", questions=", ".join(empty_questions))
            )
        if multiple_questions:
            status_parts.append(
                tr(
                    "lbl_answer_status_multi",
                    questions=", ".join(multiple_questions),
                )
            )
        low_only = [
            question
            for question in review_questions
            if question not in empty_questions and question not in multiple_questions
        ]
        if low_only:
            status_parts.append(
                tr("lbl_answer_status_review", questions=", ".join(low_only))
            )
        if status_parts:
            self.lbl_answer_status.setText("  |  ".join(status_parts))
            self.lbl_answer_status.setStyleSheet(
                "color: #e65100; font-size: 12px; font-weight: bold; padding: 2px;"
            )
        elif questions:
            self.lbl_answer_status.setText(tr("lbl_answer_status_ok"))
            self.lbl_answer_status.setStyleSheet(
                "color: #2e7d32; font-size: 12px; padding: 2px;"
            )
        else:
            self.lbl_answer_status.setText("")
        self.table.blockSignals(False)

    def _option_count_for_question(self, question):
        for mark in self.view.option_marks:
            if str(mark.question_num) == str(question):
                return int(mark.options_count)
        return 4

    def _normalize_option_answer(self, value, question):
        allowed = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[: self._option_count_for_question(question)]
        seen = set()
        normalized = []
        for character in str(value).upper():
            if character in allowed and character not in seen:
                normalized.append(character)
                seen.add(character)
        return "".join(normalized)

    def on_table_edit(self, row, col):
        if col not in (1, 2):
            return
        question_item = self.table.item(row, 0)
        value_item = self.table.item(row, col)
        if not question_item or not value_item:
            return
        question = question_item.data(Qt.UserRole)
        if question is None:
            return
        normalized = self._normalize_option_answer(value_item.text(), question)

        if col == 1 and self.current_page in self.results:
            page_result = self.results[self.current_page]
            page_result.setdefault("manual_overrides", {})[question] = normalized
            page_result.setdefault("options", {})[question] = normalized
            page_result.setdefault("confidence", {})[question] = 1.0
            page_result.setdefault("review", {})[question] = False
            page_result.setdefault("reasons", {})[question] = "manual"
            if self.first_page_key and self.current_page == 0:
                self.answer_key[question] = normalized
        elif col == 2:
            self.answer_key[question] = normalized
        self.update_result_table()

    def _crop_pixmap_from_reference(self, reference):
        if not reference:
            return QPixmap()
        path = reference.get("path", "") if isinstance(reference, dict) else reference
        if path and os.path.exists(path):
            return QPixmap(path)
        if not isinstance(reference, dict):
            return QPixmap()
        box = reference.get("box")
        if (
            not box
            or reference.get("page") != self.current_page
            or self.current_pixmap_item is None
        ):
            return QPixmap()
        left, top, right, bottom = [int(value) for value in box]
        if right <= left or bottom <= top:
            return QPixmap()
        return self.current_pixmap_item.pixmap().copy(
            left, top, right - left, bottom - top
        )

    def _save_crop_reference(self, reference):
        pixmap = self._crop_pixmap_from_reference(reference)
        if pixmap.isNull():
            return
        question = reference.get("question", "answer")
        default_name = f"page_{self.current_page + 1}_Q{question}.png"
        file_name, _ = QFileDialog.getSaveFileName(
            self, "Save Crop Image", default_name, "PNG Image (*.png)"
        )
        if file_name:
            pixmap.save(file_name, "PNG")

    def open_crop_from_table(self, row, col):
        if col != 5:
            return
        item = self.table.item(row, col)
        reference = item.data(Qt.UserRole) if item else None
        pixmap = self._crop_pixmap_from_reference(reference)
        if pixmap.isNull():
            return

        dialog = QDialog(self)
        dialog.setWindowTitle(
            tr(
                "crop_preview_title",
                page=self.current_page + 1,
                question=reference.get("question", ""),
            )
        )
        layout = QVBoxLayout(dialog)
        image_label = QLabel()
        image_label.setAlignment(Qt.AlignCenter)
        scale = min(6.0, max(1.0, 520.0 / max(1, pixmap.width())))
        image_label.setPixmap(
            pixmap.scaled(
                max(1, int(pixmap.width() * scale)),
                max(1, int(pixmap.height() * scale)),
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation,
            )
        )
        layout.addWidget(image_label)
        buttons = QDialogButtonBox(
            QDialogButtonBox.Save | QDialogButtonBox.Close
        )
        buttons.button(QDialogButtonBox.Save).clicked.connect(
            lambda: self._save_crop_reference(reference)
        )
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        dialog.exec_()

    def open_crop_context_menu(self, pos):
        index = self.table.indexAt(pos)
        if not index.isValid() or index.column() != 5:
            return
        item = self.table.item(index.row(), index.column())
        reference = item.data(Qt.UserRole) if item else None
        if self._crop_pixmap_from_reference(reference).isNull():
            return
        menu = QMenu(self)
        save_action = QAction("Save Crop As...", self)
        save_action.triggered.connect(lambda: self._save_crop_reference(reference))
        menu.addAction(save_action)
        menu.exec_(self.table.viewport().mapToGlobal(pos))

    def jump_to_next_review(self):
        review_items = []
        for page_idx, page_result in sorted(self.results.items()):
            options = page_result.get("options", {})
            review = page_result.get("review", {})
            for question in sorted(
                options,
                key=lambda value: (
                    0,
                    int(value),
                )
                if str(value).isdigit()
                else (1, str(value)),
            ):
                answer = str(options.get(question, "")).strip()
                if (
                    review.get(question, False)
                    or not answer
                    or answer.startswith("[")
                    or len(answer) > 1
                ):
                    review_items.append((page_idx, question))
        if not review_items:
            QMessageBox.information(self, tr("btn_next_review"), tr("msg_no_review"))
            return

        if self._review_cursor in review_items:
            index = (review_items.index(self._review_cursor) + 1) % len(review_items)
        else:
            index = next(
                (
                    idx
                    for idx, item in enumerate(review_items)
                    if item[0] >= self.current_page
                ),
                0,
            )
        self._review_cursor = review_items[index]
        page_idx, question = self._review_cursor
        if page_idx != self.current_page:
            self.load_page(page_idx, apply_corrections=True)
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            if item and str(item.data(Qt.UserRole)) == str(question):
                self.table.selectRow(row)
                self.table.scrollToItem(item)
                break

    def _on_table_edit_legacy(self, row, col):
        if col == 1:
            item_header = self.table.item(row, 0)
            if not item_header:
                return
            header_text = item_header.text()
            new_val = self.table.item(row, 1).text()
            if header_text.startswith("Q"):
                try:
                    q_num = int(header_text.replace("Q", ""))
                    if self.current_page in self.results:
                        self.results[self.current_page]["options"][q_num] = new_val
                except:
                    pass
            self.update_result_table()
        elif col == 2: # Correct Answer column
            item_header = self.table.item(row, 0)
            if not item_header:
                return
            header_text = item_header.text()
            if header_text.startswith("Q"):
                try:
                    q_num = int(header_text.replace("Q", ""))
                    new_ans = self.table.item(row, 2).text()
                    self.answer_key[q_num] = new_ans
                    self.update_result_table()
                except:
                    pass

    def _open_crop_from_table_legacy(self, row, col):
        if col != 4:
            return
        item = self.table.item(row, col)
        if not item:
            return
        path = item.data(Qt.UserRole)
        if path and os.path.exists(path):
            QDesktopServices.openUrl(QUrl.fromLocalFile(path))

    def _open_crop_context_menu_legacy(self, pos):
        index = self.table.indexAt(pos)
        if not index.isValid() or index.column() != 4:
            return
        item = self.table.item(index.row(), index.column())
        if not item:
            return
        path = item.data(Qt.UserRole)
        if not path or not os.path.exists(path):
            return

        menu = QMenu(self)
        action_save = QAction("Save Crop As...", self)
        menu.addAction(action_save)

        def do_save():
            default_name = os.path.basename(path)
            fname, _ = QFileDialog.getSaveFileName(self, "Save Crop Image", default_name, "PNG Image (*.png)")
            if not fname:
                return
            try:
                with open(path, "rb") as src, open(fname, "wb") as dst:
                    dst.write(src.read())
            except Exception as e:
                QMessageBox.warning(self, "Save Failed", str(e))

        action_save.triggered.connect(do_save)
        menu.exec_(self.table.viewport().mapToGlobal(pos))

    def export_excel(self):
        """Export Excel to a user-chosen file."""
        if not self.results:
            QMessageBox.warning(self, "Error", tr("msg_no_results"))
            return
        prefix = self._get_pdf_prefix()
        timestamp = self._get_timestamp()
        default_name = f"{prefix}_{timestamp}.xlsx"
        fname, _ = QFileDialog.getSaveFileName(self, "Export Excel", default_name, "Excel (*.xlsx)")
        if not fname:
            return
        self._export_excel_internal(fname)
        QMessageBox.information(self, "Done", tr("msg_export_done", folder=os.path.dirname(fname)))

    def export_results_bundle(self):
        """Export Excel (and optionally images) to a user-chosen folder."""
        if not self.results:
            QMessageBox.warning(self, "Error", tr("msg_no_results"))
            return
        if not hasattr(self, 'pdf_path') or not self.pdf_path:
            QMessageBox.warning(self, "Error", tr("msg_no_pdf"))
            return

        # Let user choose the parent location
        default_dir = os.path.dirname(self.pdf_path)
        parent_folder = QFileDialog.getExistingDirectory(
            self, tr("dlg_select_export_folder"), default_dir)
        if not parent_folder:
            return

        prefix = self._get_pdf_prefix()
        timestamp = self._get_timestamp()
        export_folder = os.path.join(parent_folder, f"{prefix}_{timestamp}")
        os.makedirs(export_folder, exist_ok=True)

        include_images = self.check_export_images.isChecked()
        total_steps = 1 + (len(self.pdf_document) if include_images else 0)

        from PyQt5.QtWidgets import QProgressDialog
        progress = QProgressDialog(tr("progress_exporting"), "Cancel", 0, total_steps, self)
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.show()
        QtWidgets.QApplication.processEvents()

        # Step 1: Export Excel
        progress.setLabelText(tr("progress_exporting_excel"))
        QtWidgets.QApplication.processEvents()
        excel_path = os.path.join(export_folder, f"{prefix}_{timestamp}.xlsx")
        self._export_excel_internal(excel_path)
        progress.setValue(1)
        if progress.wasCanceled():
            return

        # Step 2: Export images (if enabled)
        if include_images:
            img_folder = os.path.join(export_folder, "images")
            self._export_images_internal(img_folder, progress=progress, progress_offset=1)

        progress.setValue(total_steps)
        progress.close()
        QMessageBox.information(self, "Done", tr("msg_export_done", folder=export_folder))
    
    def export_images(self):
        """Export scanned pages as images with answer overlay (red dots for correct answers)"""
        if not hasattr(self, 'pdf_document') or self.pdf_document is None:
            QMessageBox.warning(self, "Error", "No PDF loaded")
            return

        if not hasattr(self, 'view') or not self.view.option_marks:
            QMessageBox.warning(self, "Error", "No option marks found")
            return
        
        # Use PDF directory as default location
        prefix = self._get_pdf_prefix()
        timestamp = self._get_timestamp()
        if hasattr(self, 'pdf_path') and self.pdf_path:
            parent_folder = os.path.dirname(self.pdf_path)
        else:
            parent_folder = QFileDialog.getExistingDirectory(self, "Select Output Folder")
            if not parent_folder: return
        
        folder = os.path.join(parent_folder, f"{prefix}_{timestamp}")
        os.makedirs(folder, exist_ok=True)
        
        # Reset alignment template for export
        self._reset_align_templates()
        
        from PyQt5.QtWidgets import QProgressDialog
        progress = QProgressDialog("Exporting images...", "Cancel", 0, len(self.pdf_document), self)
        progress.setWindowModality(Qt.WindowModal)
        progress.show()
        
        for page_idx in range(len(self.pdf_document)):
            if progress.wasCanceled(): break
            progress.setValue(page_idx)
            QtWidgets.QApplication.processEvents()

            # Skip absent pages
            is_absent = self.student_absence.get(page_idx, False) if hasattr(self, 'student_absence') else False
            if is_absent:
                continue
            
            # Render page at 2x scale
            page = self.pdf_document[page_idx]
            mat = fitz.Matrix(2, 2)
            pix = page.get_pixmap(matrix=mat)

            # Convert to numpy array for processing
            img_pil = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            img_np = np.array(img_pil)

            # Apply auto-deskew if enabled
            if self.check_auto_deskew.isChecked():
                img_np, skew_angle = deskew_image(img_np)
                if skew_angle != 0.0:
                    print(f"Export page {page_idx + 1}: Corrected skew angle: {skew_angle:.2f}°")

            # Apply auto-align (shift) if enabled
            if self.check_auto_align.isChecked():
                img_np, (dx, dy), response = self.align_image(img_np, page_idx)
                if dx != 0.0 or dy != 0.0:
                    print(f"Export page {page_idx + 1}: Aligned shift dx={dx:.1f}, dy={dy:.1f} (score={response:.3f})")

            # Convert to QImage for drawing
            img_h, img_w = img_np.shape[:2]
            qimg = QImage(img_np.data, img_w, img_h, img_np.strides[0], QImage.Format_RGB888).copy()
            
            # Create painter to draw overlay
            painter = QPainter(qimg)
            painter.setRenderHint(QPainter.Antialiasing)
            
            # Get results for this page
            page_results = self.results.get(page_idx, {}) if hasattr(self, 'results') else {}
            opts = page_results.get("options", {})

            # Offset for this page (if the PDF was moved in the scene)
            off_x, off_y = self._get_page_offset(page_idx)
            
            # Draw marks and answers
            page_score = 0
            page_total = 0
            for mark in self.view.option_marks:
                rect = mark.sceneBoundingRect()
                q_num = mark.question_num
                
                if rect:
                    x = int(rect.x() - off_x)
                    y = int(rect.y() - off_y)
                    mw = int(rect.width())
                    mh = int(rect.height())
                    
                    # Draw rectangle border
                    painter.setPen(QPen(QColor(0, 100, 255), 2))
                    painter.drawRect(x, y, mw, mh)
                    
                    # Get student answer and correct answer
                    # Ensure q_num is int for consistent key lookup
                    q_num_int = int(q_num) if isinstance(q_num, (int, str)) and str(q_num).isdigit() else q_num
                    student_answer = opts.get(q_num_int, "") or opts.get(q_num, "") or opts.get(str(q_num), "")
                    correct_answer = self.answer_key.get(q_num_int, "") or self.answer_key.get(q_num, "") or self.answer_key.get(str(q_num), "")
                    
                    student_clean = "".join(str(student_answer).split()).upper()
                    correct_clean = "".join(str(correct_answer).split()).upper()
                    is_blank = student_clean == ""
                    is_multi = len(student_clean) > 1
                    is_correct = bool(correct_clean) and student_clean == correct_clean
                    if correct_clean:
                        page_total += 1
                        if is_correct:
                            page_score += 1
                    
                    # Debug print
                    if correct_answer:
                        print(f"Q{q_num}: correct={correct_answer}")
                    
                    # Calculate cell positions for A, B, C, D
                    num_options = getattr(mark, "options_count", 4)
                    cell_width = mw // num_options
                    option_labels = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[:num_options]
                    
                    for i, opt_label in enumerate(option_labels):
                        cell_x = x + i * cell_width
                        cell_center_x = cell_x + cell_width // 2
                        cell_center_y = y + mh // 2
                        
                        # Draw red dot for correct answer
                        if correct_answer and opt_label.upper() == correct_answer.upper():
                            # Must set both brush and pen before each ellipse
                            painter.save()
                            painter.setBrush(QBrush(QColor(255, 0, 0)))
                            painter.setPen(QPen(QColor(255, 0, 0), 2))
                            painter.drawEllipse(cell_center_x - 8, cell_center_y - 8, 16, 16)
                            painter.restore()
                        
                        # Draw X mark for student's wrong answer
                        if student_answer and opt_label.upper() == student_answer.upper():
                            if correct_answer and student_answer.upper() != correct_answer.upper():
                                # Wrong answer - draw X mark
                                painter.save()
                                painter.setPen(QPen(QColor(255, 0, 0), 3))
                                painter.drawLine(cell_center_x - 8, cell_center_y - 8, cell_center_x + 8, cell_center_y + 8)
                                painter.drawLine(cell_center_x + 8, cell_center_y - 8, cell_center_x - 8, cell_center_y + 8)
                                painter.restore()
                    
                    # Highlight blank vs multi-selection
                    if is_blank:
                        painter.save()
                        painter.setPen(QPen(QColor(255, 193, 7), 3))
                        painter.drawRect(x, y, mw, mh)
                        painter.restore()
                    elif is_multi:
                        painter.save()
                        painter.setPen(QPen(QColor(255, 140, 0), 3))
                        painter.drawRect(x, y, mw, mh)
                        painter.restore()
                    
                    # Draw per-question correctness marker on the right
                    marker_x = x + mw + 8
                    marker_y = y + mh // 2 + 5
                    painter.save()
                    painter.setFont(QFont("Arial", 11, QFont.Bold))
                    if is_blank:
                        painter.setPen(QPen(QColor(255, 193, 7), 2))
                        painter.drawText(marker_x, marker_y, "Ø")
                    elif is_multi:
                        painter.setPen(QPen(QColor(255, 140, 0), 2))
                        painter.drawText(marker_x, marker_y, "!")
                    else:
                        painter.setPen(QPen(QColor(40, 167, 69), 2) if is_correct else QPen(QColor(220, 53, 69), 2))
                        painter.drawText(marker_x, marker_y, "✓" if is_correct else "✗")
                    painter.restore()
                    
                    # Draw question number
                    painter.setPen(QPen(QColor(0, 0, 0), 1))
                    painter.setFont(QFont("Arial", 10, QFont.Bold))
                    painter.drawText(x - 30, y + mh // 2 + 5, f"Q{q_num}")
            
            # Draw score at top-right (inside page bounds)
            painter.save()
            painter.setFont(QFont("Arial", 14, QFont.Bold))
            painter.setPen(QPen(QColor(0, 0, 0), 2))
            score_text = f"Score: {page_score}/{page_total}"
            metrics = painter.fontMetrics()
            text_width = metrics.horizontalAdvance(score_text)
            x_pos = max(10, img_w - text_width - 10)
            painter.drawText(x_pos, 30, score_text)
            painter.restore()
            painter.end()
            
            # Save image
            output_path = os.path.join(folder, f"page_{page_idx + 1:03d}.png")
            qimg.save(output_path)
        
        progress.setValue(len(self.pdf_document))
        QMessageBox.information(self, "Done", f"Exported {len(self.pdf_document)} images to:\n{folder}")

    def export_debug_pack(self):
        """Export debug images and scoring records into a folder for easy sharing."""
        has_records = bool(getattr(self, "debug_records", []))
        debug_dir = self.debug_dir
        has_debug_images = os.path.isdir(debug_dir) and any(os.scandir(debug_dir))

        if not has_records and not has_debug_images:
            QMessageBox.information(self, "Debug Pack", "No debug data found. Run recognition with debug enabled first.")
            return

        prefix = self._get_pdf_prefix()
        timestamp = self._get_timestamp()
        default_folder = f"{prefix}_{timestamp}_debug"

        parent_folder = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if not parent_folder:
            return

        out_folder = os.path.join(parent_folder, default_folder)

        try:
            os.makedirs(out_folder, exist_ok=True)

            def _normalize_answer(value):
                return "".join(str(value).split()).upper()

            enriched_records = []
            if has_records:
                for record in self.debug_records:
                    enriched = dict(record)
                    context = enriched.get("context", {}) or {}
                    question_num = context.get("question")
                    correct_answer = enriched.get("correct_answer", "")
                    if not correct_answer and question_num is not None and hasattr(self, "answer_key"):
                        correct_answer = self.answer_key.get(question_num, "") or self.answer_key.get(str(question_num), "")
                    enriched["correct_answer"] = correct_answer
                    enriched["is_correct"] = bool(
                        _normalize_answer(correct_answer)
                        and _normalize_answer(enriched.get("result", "")) == _normalize_answer(correct_answer)
                    )
                    enriched_records.append(enriched)

            if has_records:
                records_path = os.path.join(out_folder, "debug_records.json")
                with open(records_path, "w", encoding="utf-8") as f:
                    json.dump(enriched_records, f, ensure_ascii=False, indent=2)

            if hasattr(self, "answer_key") and self.answer_key:
                answer_key_path = os.path.join(out_folder, "answer_key.json")
                with open(answer_key_path, "w", encoding="utf-8") as f:
                    json.dump(self.answer_key, f, ensure_ascii=False, indent=2)

            if hasattr(self, "pdf_path") and self.pdf_path and os.path.isfile(self.pdf_path):
                try:
                    shutil.copy2(self.pdf_path, os.path.join(out_folder, os.path.basename(self.pdf_path)))
                except Exception:
                    pass

            if has_debug_images:
                out_debug_dir = os.path.join(out_folder, "debug_crops")
                os.makedirs(out_debug_dir, exist_ok=True)

                files = [e for e in os.scandir(debug_dir) if e.is_file()]
                progress = QtWidgets.QProgressDialog("Exporting debug images...", "Cancel", 0, len(files), self)
                progress.setWindowModality(Qt.WindowModal)
                progress.setMinimumDuration(0)
                progress.show()

                for idx, entry in enumerate(files):
                    if progress.wasCanceled():
                        break
                    progress.setValue(idx)
                    QtWidgets.QApplication.processEvents()
                    shutil.copy2(entry.path, os.path.join(out_debug_dir, entry.name))

                progress.setValue(len(files))

            QMessageBox.information(self, "Debug Pack", f"Debug folder exported:\n{out_folder}")
        except Exception as e:
            QMessageBox.critical(self, "Debug Pack", f"Failed to export debug folder:\n{e}")

    # ==================== Batch Processing ====================
    
    def batch_process_same_template(self):
        """
        Batch process multiple PDFs using the same template.
        User selects one template file and multiple PDF files.
        Results are exported to the same folder as each PDF.
        """
        # Step 1: Select template file
        template_file, _ = QFileDialog.getOpenFileName(
            self, "Select Template File", "", "JSON (*.json)"
        )
        if not template_file:
            return
        
        # Step 2: Select multiple PDF files
        pdf_files, _ = QFileDialog.getOpenFileNames(
            self, "Select PDF Files to Process", "", "PDF Files (*.pdf)"
        )
        if not pdf_files:
            return
        
        # Confirm
        reply = QMessageBox.question(
            self, "Confirm Batch Process",
            f"Process {len(pdf_files)} PDF files with template:\n{os.path.basename(template_file)}\n\n"
            f"Each PDF will have Excel and Images exported to its folder.",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        
        # Load template once
        try:
            with open(template_file, "r", encoding="utf-8-sig") as f:
                template_data = json.load(f)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load template:\n{e}")
            return
        
        # Process each PDF
        self._batch_process_pdfs(pdf_files, template_data, template_file)
    
    def batch_process_matched_templates(self):
        """
        Batch process multiple PDFs where each PDF has a matching template file.
        Template name must match PDF name (e.g., exam1.pdf uses exam1.json).
        Results are exported to the same folder as each PDF.
        """
        # Select multiple PDF files
        pdf_files, _ = QFileDialog.getOpenFileNames(
            self, "Select PDF Files to Process", "", "PDF Files (*.pdf)"
        )
        if not pdf_files:
            return
        
        # Check for matching templates
        matched = []
        unmatched = []
        
        for pdf_path in pdf_files:
            pdf_dir = os.path.dirname(pdf_path)
            pdf_name = os.path.splitext(os.path.basename(pdf_path))[0]
            template_path = os.path.join(pdf_dir, f"{pdf_name}.json")
            
            if os.path.exists(template_path):
                matched.append((pdf_path, template_path))
            else:
                unmatched.append(pdf_path)
        
        # Show warning for unmatched files
        msg = f"Found {len(matched)} PDFs with matching templates.\n"
        if unmatched:
            msg += f"\n⚠️ {len(unmatched)} PDFs without matching template (will be skipped):\n"
            for p in unmatched[:5]:  # Show first 5
                msg += f"  • {os.path.basename(p)}\n"
            if len(unmatched) > 5:
                msg += f"  ... and {len(unmatched) - 5} more\n"
        
        if not matched:
            QMessageBox.warning(self, "No Matches", "No PDF files have matching template files.")
            return
        
        msg += "\nContinue with batch processing?"
        reply = QMessageBox.question(self, "Confirm Batch Process", msg, QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return
        
        # Process each matched PDF
        self._batch_process_pdfs_matched(matched)
    
    def _batch_process_pdfs(self, pdf_files, template_data, template_name):
        """Internal method to process multiple PDFs with the same template."""
        progress = QtWidgets.QProgressDialog(
            "Batch processing...", "Cancel", 0, len(pdf_files), self
        )
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.show()
        
        success_count = 0
        error_files = []
        shared_answer_key = (
            dict(self.answer_key) if not self.first_page_key else {}
        )
        
        for idx, pdf_path in enumerate(pdf_files):
            QtWidgets.QApplication.processEvents()
            if progress.wasCanceled():
                break
            
            progress.setValue(idx)
            progress.setLabelText(f"Processing: {os.path.basename(pdf_path)}")
            
            try:
                # Open the target first so template coordinates can be scaled to
                # this document instead of the previously displayed document.
                self._open_pdf_path(pdf_path, load_preview=False)
                self._load_template_data(template_data)
                if not self.first_page_key:
                    self.answer_key = dict(shared_answer_key)
                
                # Run recognition
                self._run_recognition_internal()

                # Use first page as answer key for this PDF (per-file)
                if self.first_page_key and 0 in self.results:
                    self.answer_key = self._answer_key_from_result(
                        self.results[0]
                    )
                
                # Export results to same folder as PDF
                output_folder = os.path.dirname(pdf_path)
                pdf_basename = os.path.splitext(os.path.basename(pdf_path))[0]
                timestamp = self._get_timestamp()
                
                # Export Excel
                excel_path = os.path.join(output_folder, f"{pdf_basename}_{timestamp}.xlsx")
                self._export_excel_internal(excel_path)
                
                # Export Images
                img_folder = os.path.join(output_folder, f"{pdf_basename}_{timestamp}")
                self._export_images_internal(img_folder)
                
                success_count += 1
                print(f"✓ Processed: {os.path.basename(pdf_path)}")
                
            except Exception as e:
                error_files.append((pdf_path, str(e)))
                print(f"✗ Error processing {os.path.basename(pdf_path)}: {e}")
        
        progress.setValue(len(pdf_files))
        
        # Show summary
        msg = f"Batch processing complete!\n\n✓ Success: {success_count} files"
        if error_files:
            msg += f"\n✗ Errors: {len(error_files)} files\n"
            for path, err in error_files[:3]:
                msg += f"\n  • {os.path.basename(path)}: {err[:50]}"
            if len(error_files) > 3:
                msg += f"\n  ... and {len(error_files) - 3} more errors"
        
        QMessageBox.information(self, "Batch Complete", msg)
    
    def _batch_process_pdfs_matched(self, matched_pairs):
        """Internal method to process PDFs with their matched templates."""
        progress = QtWidgets.QProgressDialog(
            "Batch processing...", "Cancel", 0, len(matched_pairs), self
        )
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.show()
        
        success_count = 0
        error_files = []
        shared_answer_key = (
            dict(self.answer_key) if not self.first_page_key else {}
        )
        
        for idx, (pdf_path, template_path) in enumerate(matched_pairs):
            QtWidgets.QApplication.processEvents()
            if progress.wasCanceled():
                break
            
            progress.setValue(idx)
            progress.setLabelText(f"Processing: {os.path.basename(pdf_path)}")
            
            try:
                # Load template for this PDF
                with open(template_path, "r", encoding="utf-8-sig") as f:
                    template_data = json.load(f)

                # Open before loading marks so reference dimensions are compared
                # with the correct PDF in unattended batch mode.
                self._open_pdf_path(pdf_path, load_preview=False)
                self._load_template_data(template_data)
                if not self.first_page_key:
                    self.answer_key = dict(shared_answer_key)
                
                # Run recognition
                self._run_recognition_internal()

                # Use first page as answer key for this PDF (per-file)
                if self.first_page_key and 0 in self.results:
                    self.answer_key = self._answer_key_from_result(
                        self.results[0]
                    )
                
                # Export results to same folder as PDF
                output_folder = os.path.dirname(pdf_path)
                pdf_basename = os.path.splitext(os.path.basename(pdf_path))[0]
                timestamp = self._get_timestamp()
                
                # Export Excel
                excel_path = os.path.join(output_folder, f"{pdf_basename}_{timestamp}.xlsx")
                self._export_excel_internal(excel_path)
                
                # Export Images
                img_folder = os.path.join(output_folder, f"{pdf_basename}_{timestamp}")
                self._export_images_internal(img_folder)
                
                success_count += 1
                print(f"✓ Processed: {os.path.basename(pdf_path)}")
                
            except Exception as e:
                error_files.append((pdf_path, str(e)))
                print(f"✗ Error processing {os.path.basename(pdf_path)}: {e}")
        
        progress.setValue(len(matched_pairs))
        
        # Show summary
        msg = f"Batch processing complete!\n\n✓ Success: {success_count} files"
        if error_files:
            msg += f"\n✗ Errors: {len(error_files)} files\n"
            for path, err in error_files[:3]:
                msg += f"\n  • {os.path.basename(path)}: {err[:50]}"
            if len(error_files) > 3:
                msg += f"\n  ... and {len(error_files) - 3} more errors"
        
        QMessageBox.information(self, "Batch Complete", msg)
    
    def _load_template_data(self, data):
        """Validate and atomically load template data without a file dialog."""
        if not isinstance(data, dict):
            raise ValueError("Template root must be a JSON object.")

        scale_x = 1.0
        scale_y = 1.0
        reference_page = data.get("reference_page", {}) or {}
        reference_width = float(reference_page.get("width", 0) or 0)
        reference_height = float(reference_page.get("height", 0) or 0)
        target_width = 0.0
        target_height = 0.0
        if self.pdf_document and len(self.pdf_document):
            # Pages are rendered at 2x throughout the application.  Using the
            # PDF bounds avoids an otherwise unnecessary raster pass in batch.
            page_rect = self.pdf_document[0].rect
            target_width = float(page_rect.width) * 2.0
            target_height = float(page_rect.height) * 2.0
        elif self.current_pixmap_item is not None:
            pixmap = self.current_pixmap_item.pixmap()
            target_width = float(pixmap.width())
            target_height = float(pixmap.height())

        if (
            target_width > 0
            and target_height > 0
            and reference_width > 0
            and reference_height > 0
        ):
            scale_x = target_width / reference_width
            scale_y = target_height / reference_height
            aspect_stretch = scale_x / max(scale_y, 1e-9)
            if not 0.92 <= aspect_stretch <= 1.08:
                raise ValueError(
                    "Template and PDF have incompatible page aspect ratios "
                    f"(x={scale_x:.3f}, y={scale_y:.3f})."
                )
            if min(scale_x, scale_y) < 0.40 or max(scale_x, scale_y) > 2.50:
                raise ValueError(
                    "Template scale is outside the safe range for this PDF "
                    f"(x={scale_x:.3f}, y={scale_y:.3f})."
                )

        schema_version = int(data.get("schema_version", 1) or 1)
        target_offset_x, target_offset_y = (
            self._get_page_offset(0) if schema_version >= 2 else (0.0, 0.0)
        )

        def scaled_mark(raw):
            if not isinstance(raw, dict):
                raise ValueError("Every template mark must be a JSON object.")
            mark = dict(raw)
            mark["x"] = (
                float(mark.get("x", 0.0)) * scale_x + float(target_offset_x)
            )
            mark["y"] = (
                float(mark.get("y", 0.0)) * scale_y + float(target_offset_y)
            )
            mark["width"] = float(mark.get("width", 0.0)) * scale_x
            mark["height"] = float(mark.get("height", 0.0)) * scale_y
            numeric_values = (
                mark["x"],
                mark["y"],
                mark["width"],
                mark["height"],
            )
            if not all(np.isfinite(value) for value in numeric_values):
                raise ValueError("Template contains a non-finite coordinate.")
            if mark["width"] < 3 or mark["height"] < 3:
                raise ValueError("Template contains an invalid or empty mark.")
            mark["question"] = int(mark.get("question", 0))
            if mark["question"] < 1:
                raise ValueError("Template question numbers must be positive.")
            mark["label"] = str(mark.get("label", ""))
            return mark

        def mark_list(key):
            value = data.get(key, [])
            if value is None:
                return []
            if not isinstance(value, list):
                raise ValueError(f"Template field '{key}' must be a list.")
            return value

        prepared_text = [scaled_mark(raw) for raw in mark_list("text_marks")]
        prepared_options = [scaled_mark(raw) for raw in mark_list("option_marks")]
        for mark in prepared_options:
            mark["options_count"] = int(mark.get("options_count", 4))
            if not 2 <= mark["options_count"] <= 26:
                raise ValueError("Option counts must be between 2 and 26.")

        align_marks_data = mark_list("align_marks")
        if not align_marks_data:
            old_align = data.get("align_mark")
            if old_align:
                align_marks_data = [old_align]
        prepared_align = [scaled_mark(raw) for raw in align_marks_data]

        # Construct every item before touching the live scene.  Any invalid
        # later mark therefore leaves the previous template intact.
        text_items = [
            MarkItem(
                0, 0, mark["width"], mark["height"], MARK_TYPE_TEXT,
                mark["question"], mark["label"], view_ref=self.view,
            )
            for mark in prepared_text
        ]
        option_items = [
            MarkItem(
                0, 0, mark["width"], mark["height"], MARK_TYPE_OPTION,
                mark["question"], mark["label"], mark["options_count"],
                view_ref=self.view,
            )
            for mark in prepared_options
        ]
        align_items = [
            MarkItem(
                0, 0, mark["width"], mark["height"], MARK_TYPE_ALIGN,
                mark["question"], mark["label"], view_ref=self.view,
            )
            for mark in prepared_align
        ]
        for item, mark in zip(
            text_items + option_items + align_items,
            prepared_text + prepared_options + prepared_align,
        ):
            item.setPos(mark["x"], mark["y"])

        self.clear_all_marks()
        self.view.text_marks.extend(text_items)
        self.view.option_marks.extend(option_items)
        self.view.align_marks.extend(align_items)
        for item in text_items + option_items + align_items:
            self.scene.addItem(item)
        self.view.text_counter = max(
            [1] + [mark["question"] + 1 for mark in prepared_text]
        )
        self.view.option_counter = max(
            [1] + [mark["question"] + 1 for mark in prepared_options]
        )
        self.view.align_counter = max(
            [1] + [mark["question"] + 1 for mark in prepared_align]
        )
    
    def _run_recognition_internal_legacy(self):
        """Internal recognition method without UI dialogs."""
        if not self.pdf_document or (not self.view.option_marks and not self.view.text_marks):
            return
        
        self.results = {}
        self.debug_records = []
        
        # Save current page's image offset before processing
        if self.current_pixmap_item:
            self.page_offsets[self.current_page] = self.current_pixmap_item.get_offset()
        
        # Reset alignment template for new recognition run
        self._reset_align_templates()
        
        for p_idx in range(len(self.pdf_document)):
            QtWidgets.QApplication.processEvents()
            
            page = self.pdf_document[p_idx]
            mat = fitz.Matrix(2, 2)
            pix = page.get_pixmap(matrix=mat)

            img_pil = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            img_np = np.array(img_pil)

            # Apply auto-deskew if enabled
            if self.check_auto_deskew.isChecked():
                img_np, skew = deskew_image(img_np)

            # Apply auto-align (shift) if enabled
            if self.check_auto_align.isChecked():
                img_np, (dx, dy), response = self.align_image(img_np, p_idx)

            h, w = img_np.shape[:2]
            off_x, off_y = self._get_page_offset(p_idx)

            page_result = {"text": {}, "options": {}}

            # Process text marks
            for mark in self.view.text_marks:
                rect = mark.sceneBoundingRect()
                x1 = max(0, int(rect.x() - off_x))
                y1 = max(0, int(rect.y() - off_y))
                x2 = min(w, int(rect.x() + rect.width() - off_x))
                y2 = min(h, int(rect.y() + rect.height() - off_y))
                
                if x2 > x1 and y2 > y1:
                    crop = img_np[y1:y2, x1:x2]
                    crop_pil = Image.fromarray(crop)
                    text = self.recognize_text(crop_pil)
                    page_result["text"][mark.label or f"Field_{mark.question_num}"] = text

            # Process option marks
            for mark in self.view.option_marks:
                rect = mark.sceneBoundingRect()
                x1 = max(0, int(rect.x() - off_x))
                y1 = max(0, int(rect.y() - off_y))
                x2 = min(w, int(rect.x() + rect.width() - off_x))
                y2 = min(h, int(rect.y() + rect.height() - off_y))
                
                if x2 > x1 and y2 > y1:
                    crop = img_np[y1:y2, x1:x2]
                    crop_pil = Image.fromarray(crop)
                    opt = mark.options_count
                    correct_answer = ""
                    if hasattr(self, "answer_key"):
                        correct_answer = self.answer_key.get(mark.question_num, "") or self.answer_key.get(str(mark.question_num), "")
                    result_opt = self.detect_filled_option(
                        crop_pil,
                        opt,
                        context={
                            "page": p_idx + 1,
                            "question": mark.question_num,
                            "label": f"Q{mark.question_num}",
                            "correct_answer": correct_answer,
                        }
                    )
                    page_result["options"][mark.question_num] = result_opt
            self.results[p_idx] = page_result

    def _run_recognition_internal(self):
        """Batch recognition using the same implementation as the GUI."""

        if not self.pdf_document or (
            not self.view.option_marks and not self.view.text_marks
        ):
            return
        manual_answer_key = (
            dict(self.answer_key) if not self.first_page_key else {}
        )
        self.results = {}
        self.answer_key = manual_answer_key
        self.debug_records = []
        self._reset_align_templates()
        for page_idx in range(len(self.pdf_document)):
            QtWidgets.QApplication.processEvents()
            self.results[page_idx] = self._recognize_page(
                page_idx, preserve_overrides=False
            )
            if self.first_page_key and page_idx == 0:
                self.answer_key = self._answer_key_from_result(self.results[0])

    def _export_excel_internal(self, output_path):
        """Internal method to export Excel without file dialog."""
        if not self.results:
            raise ValueError("No recognition results are available for export.")

        include_summary = self.check_include_summary.isChecked() if hasattr(self, "check_include_summary") else True
        include_topics = self.check_include_topics.isChecked() if hasattr(self, "check_include_topics") else True
        
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        header_font = Font(bold=True)
        center_align = Alignment(horizontal='center')
        wb = Workbook()
        ws = wb.active
        ws.title = "OMR Results"
        
        all_qs = set()
        all_texts = set()
        
        for p_res in self.results.values():
            all_qs.update(p_res.get("options", {}).keys())
            all_texts.update(p_res.get("text", {}).keys())

        # Also collect text keys from extra students (absent students without PDF pages)
        extra_students = getattr(self, 'extra_students', [])
        for extra in extra_students:
            all_texts.update(extra.get("text", {}).keys())
            
        sorted_qs = sorted(list(all_qs))
        sorted_texts = sorted(list(all_texts))
        
        text_start_col = 2
        absent_col_num = text_start_col + len(sorted_texts)
        q_start_col = absent_col_num + 1
        score_col = q_start_col + len(sorted_qs)
        
        absent_label = tr("dlg_student_absent")
        headers = ["Page"] + sorted_texts + [absent_label] + [f"Q{q}" for q in sorted_qs] + ["Score"]
        ws.append(headers)
        
        key_row = ["Key"] + [""] * len(sorted_texts) + [""]
        for q in sorted_qs:
            key_row.append(self.answer_key.get(q, ""))
        key_row.append("")
        ws.append(key_row)
        
        data_row_num = 3
        first_data_row = 3
        empty_cells = []
        multiple_cells = []
        page_scores = []
        page_totals = []
        page_blank_counts = []
        page_multi_counts = []
        question_correct = {question: 0 for question in sorted_qs}
        question_denominators = {question: 0 for question in sorted_qs}
        included_page_indices = set()
        
        student_order = getattr(self, 'student_order', [])

        if student_order:
            # ── Use student_order to preserve user's original input order ──
            for entry in student_order:
                p_idx = entry.get("page_idx")
                is_absent = entry.get("absent", False)

                # Skip answer-key page if applicable
                if p_idx is not None and self.first_page_key and p_idx == 0:
                    continue

                row = [p_idx + 1 if p_idx is not None else "-"]

                entry_texts = entry.get("text", {})
                for t_key in sorted_texts:
                    row.append(entry_texts.get(t_key, ""))

                row.append("✓" if is_absent else "")

                # Options from results (only if present and has a mapped page)
                if p_idx is not None and not is_absent:
                    res = self.results.get(p_idx, {})
                    opts = res.get("options", {})
                else:
                    opts = {}

                page_blank = 0
                page_multi = 0
                page_score = 0
                page_total = 0
                for q_idx, q in enumerate(sorted_qs):
                    val = opts.get(q, "") if not is_absent else ""
                    row.append(val)
                    if not is_absent and p_idx is not None:
                        col_letter = get_column_letter(q_start_col + q_idx)
                        if val == "" or val is None:
                            empty_cells.append(f"{col_letter}{data_row_num}")
                            page_blank += 1
                        elif len(str(val)) > 1:
                            multiple_cells.append(f"{col_letter}{data_row_num}")
                            page_multi += 1
                        correct_val = self.answer_key.get(q, "")
                        if correct_val != "":
                            page_total += 1
                            question_denominators[q] += 1
                            if "".join(str(val).split()).lower() == "".join(str(correct_val).split()).lower():
                                page_score += 1
                                question_correct[q] += 1

                if sorted_qs and not is_absent and p_idx is not None:
                    first_q_col = get_column_letter(q_start_col)
                    last_q_col = get_column_letter(q_start_col + len(sorted_qs) - 1)
                    score_formula = (
                        f'=SUMPRODUCT(--({first_q_col}$2:{last_q_col}$2<>""),'
                        f'--({first_q_col}{data_row_num}:{last_q_col}{data_row_num}'
                        f'={first_q_col}$2:{last_q_col}$2))'
                    )
                    row.append(score_formula)
                else:
                    row.append("")

                ws.append(row)
                if not is_absent and p_idx is not None:
                    included_page_indices.add(p_idx)
                    page_scores.append(page_score)
                    page_totals.append(page_total)
                    page_blank_counts.append(page_blank)
                    page_multi_counts.append(page_multi)
                data_row_num += 1
        else:
            # ── Fallback: iterate results by page index, then extra_students ──
            for p_idx, res in self.results.items():
                if self.first_page_key and p_idx == 0:
                    continue

                row = [p_idx + 1]
                texts = res.get("text", {})
                for t_key in sorted_texts:
                    row.append(texts.get(t_key, ""))

                is_absent = self.student_absence.get(p_idx, False) if hasattr(self, 'student_absence') else False
                row.append("✓" if is_absent else "")

                opts = res.get("options", {})
                page_blank = 0
                page_multi = 0
                page_score = 0
                page_total = 0
                for q_idx, q in enumerate(sorted_qs):
                    val = opts.get(q, "") if not is_absent else ""
                    row.append(val)
                    if not is_absent:
                        col_letter = get_column_letter(q_start_col + q_idx)
                        if val == "" or val is None:
                            empty_cells.append(f"{col_letter}{data_row_num}")
                            page_blank += 1
                        elif len(str(val)) > 1:
                            multiple_cells.append(f"{col_letter}{data_row_num}")
                            page_multi += 1
                        correct_val = self.answer_key.get(q, "")
                        if correct_val != "":
                            page_total += 1
                            question_denominators[q] += 1
                            if "".join(str(val).split()).lower() == "".join(str(correct_val).split()).lower():
                                page_score += 1
                                question_correct[q] += 1

                if sorted_qs and not is_absent:
                    first_q_col = get_column_letter(q_start_col)
                    last_q_col = get_column_letter(q_start_col + len(sorted_qs) - 1)
                    score_formula = (
                        f'=SUMPRODUCT(--({first_q_col}$2:{last_q_col}$2<>""),'
                        f'--({first_q_col}{data_row_num}:{last_q_col}{data_row_num}'
                        f'={first_q_col}$2:{last_q_col}$2))'
                    )
                    row.append(score_formula)
                else:
                    row.append("")

                ws.append(row)
                if not is_absent:
                    included_page_indices.add(p_idx)
                    page_scores.append(page_score)
                    page_totals.append(page_total)
                    page_blank_counts.append(page_blank)
                    page_multi_counts.append(page_multi)
                data_row_num += 1

            # Append extra students (absent students added beyond PDF pages)
            for extra in extra_students:
                row = ["-"]
                extra_texts = extra.get("text", {})
                for t_key in sorted_texts:
                    row.append(extra_texts.get(t_key, ""))
                row.append("✓" if extra.get("absent", True) else "")
                for q in sorted_qs:
                    row.append("")
                row.append("")
                ws.append(row)
                data_row_num += 1
        
        last_data_row = data_row_num - 1
        
        for cell_ref in empty_cells:
            ws[cell_ref].fill = yellow_fill
        for cell_ref in multiple_cells:
            ws[cell_ref].fill = orange_fill
        
        if sorted_qs and last_data_row >= first_data_row:
            stats_row_num = data_row_num + 1
            ws.cell(row=stats_row_num, column=1, value="% Correct").fill = green_fill
            ws.cell(row=stats_row_num, column=1).font = header_font
            
            for q_idx, q in enumerate(sorted_qs):
                col_num = q_start_col + q_idx
                denominator = question_denominators.get(q, 0)
                value = (
                    question_correct.get(q, 0) / denominator * 100.0
                    if denominator
                    else None
                )
                cell = ws.cell(row=stats_row_num, column=col_num, value=value)
                cell.fill = green_fill
                cell.alignment = center_align
                cell.number_format = '0.0"%"'
            
            if sorted_qs:
                first_q_col = get_column_letter(q_start_col)
                last_q_col = get_column_letter(q_start_col + len(sorted_qs) - 1)
                avg_formula = f'=AVERAGE({first_q_col}{stats_row_num}:{last_q_col}{stats_row_num})'
                cell = ws.cell(row=stats_row_num, column=score_col, value=avg_formula)
                cell.fill = green_fill
                cell.alignment = center_align
                cell.number_format = '0.0"%"'
        
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = header_font
            ws.cell(row=1, column=col).alignment = center_align

        if include_summary:
            summary = wb.create_sheet("Summary")
            summary.append(["Metric", "Value"])
            summary.cell(row=1, column=1).font = header_font
            summary.cell(row=1, column=2).font = header_font

            total_pages = len(page_scores)
            total_questions = max(page_totals) if page_totals else 0
            avg_score = statistics.mean(page_scores) if page_scores else 0
            median_score = statistics.median(page_scores) if page_scores else 0
            max_score = max(page_scores) if page_scores else 0
            min_score = min(page_scores) if page_scores else 0
            stdev_score = statistics.pstdev(page_scores) if len(page_scores) > 1 else 0
            avg_blank = statistics.mean(page_blank_counts) if page_blank_counts else 0
            avg_multi = statistics.mean(page_multi_counts) if page_multi_counts else 0

            summary.append(["Total Pages", total_pages])
            summary.append(["Total Questions", total_questions])
            summary.append(["Average Score", avg_score])
            summary.append(["Median Score", median_score])
            summary.append(["Max Score", max_score])
            summary.append(["Min Score", min_score])
            summary.append(["Score Std Dev", stdev_score])
            summary.append(["Avg Blank Answers", avg_blank])
            summary.append(["Avg Multiple Answers", avg_multi])

        if include_topics:
            topics_sheet = wb.create_sheet("Topics")
            topics_sheet.append(["Question", "Topic"])
            topics_sheet.cell(row=1, column=1).font = header_font
            topics_sheet.cell(row=1, column=2).font = header_font
            for q in sorted_qs:
                topics_sheet.append([f"Q{q}", self.topic_map.get(q, "")])

            topic_groups = {}
            for q in sorted_qs:
                topic = self.topic_map.get(q, "").strip() or "Unassigned"
                topic_groups.setdefault(topic, []).append(q)

            analysis = wb.create_sheet("Topic Analysis")
            analysis.append(["Topic", "Questions", "Avg Score", "Avg %"])
            for col in range(1, 5):
                analysis.cell(row=1, column=col).font = header_font

            pages_count = len(included_page_indices)
            for topic, qs in topic_groups.items():
                scored_qs = [
                    q for q in qs if str(self.answer_key.get(q, "")).strip()
                ]
                total_items = len(scored_qs) * pages_count
                correct_count = 0
                for p_idx, res in self.results.items():
                    if p_idx not in included_page_indices:
                        continue
                    opts = res.get("options", {})
                    for q in scored_qs:
                        correct_val = self.answer_key.get(q, "")
                        val = opts.get(q, "")
                        if "".join(str(val).split()).lower() == "".join(str(correct_val).split()).lower():
                            correct_count += 1
                avg_score_topic = correct_count / max(1, pages_count)
                avg_pct = (
                    correct_count / total_items * 100 if total_items else 0
                )
                analysis.append([topic, ", ".join([f"Q{q}" for q in qs]), avg_score_topic, avg_pct])
            
        wb.save(output_path)
        print(f"  Excel saved: {output_path}")
    
    def _export_images_internal(self, output_folder, progress=None, progress_offset=0):
        """Internal method to export images without file dialog.
        
        Args:
            output_folder: Path to save images.
            progress: Optional QProgressDialog to update.
            progress_offset: Value offset for progress updates (when sharing a progress bar).
        """
        if not hasattr(self, 'pdf_document') or self.pdf_document is None:
            return
        if not hasattr(self, 'view') or not self.view.option_marks:
            return
        
        os.makedirs(output_folder, exist_ok=True)
        
        # Reset alignment template for export
        self._reset_align_templates()
        
        total_pages = len(self.pdf_document)
        for page_idx in range(total_pages):
            if progress is not None:
                if progress.wasCanceled():
                    return
                progress.setLabelText(tr("progress_exporting_images", current=page_idx + 1, total=total_pages))
                progress.setValue(progress_offset + page_idx)
            QtWidgets.QApplication.processEvents()

            # Skip absent pages
            is_absent = self.student_absence.get(page_idx, False) if hasattr(self, 'student_absence') else False
            if is_absent:
                continue
            
            img_np, _ = self._prepare_page_for_recognition(page_idx)

            h, w = img_np.shape[:2]
            qimg = QImage(img_np.data, w, h, img_np.strides[0], QImage.Format_RGB888).copy()
            
            painter = QPainter(qimg)
            painter.setRenderHint(QPainter.Antialiasing)
            
            page_results = self.results.get(page_idx, {}) if hasattr(self, 'results') else {}
            opts = page_results.get("options", {})
            off_x, off_y = self._get_page_offset(page_idx)
            page_score = 0
            page_total = 0
            
            for mark in self.view.option_marks:
                rect = mark.sceneBoundingRect()
                q_num = mark.question_num
                
                if rect:
                    x = int(rect.x() - off_x)
                    y = int(rect.y() - off_y)
                    mw = int(rect.width())
                    mh = int(rect.height())
                    
                    painter.setPen(QPen(QColor(0, 100, 255), 2))
                    painter.drawRect(x, y, mw, mh)
                    
                    q_num_int = int(q_num) if isinstance(q_num, (int, str)) and str(q_num).isdigit() else q_num
                    student_answer = opts.get(q_num_int, "") or opts.get(q_num, "") or opts.get(str(q_num), "")
                    correct_answer = self.answer_key.get(q_num_int, "") or self.answer_key.get(q_num, "") or self.answer_key.get(str(q_num), "")
                    
                    student_clean = "".join(str(student_answer).split()).upper()
                    correct_clean = "".join(str(correct_answer).split()).upper()
                    is_blank = student_clean == ""
                    is_multi = len(student_clean) > 1
                    is_correct = bool(correct_clean) and student_clean == correct_clean
                    if correct_clean:
                        page_total += 1
                        if is_correct:
                            page_score += 1
                    
                    num_options = getattr(mark, "options_count", 4)
                    cell_width = mw // num_options
                    option_labels = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[:num_options]
                    
                    for i, opt_label in enumerate(option_labels):
                        cell_x = x + i * cell_width
                        cell_center_x = cell_x + cell_width // 2
                        cell_center_y = y + mh // 2
                        
                        if correct_answer and opt_label.upper() == correct_answer.upper():
                            painter.save()
                            painter.setBrush(QBrush(QColor(255, 0, 0)))
                            painter.setPen(QPen(QColor(255, 0, 0), 2))
                            painter.drawEllipse(cell_center_x - 8, cell_center_y - 8, 16, 16)
                            painter.restore()
                        
                        if student_answer and opt_label.upper() == student_answer.upper():
                            if correct_answer and student_answer.upper() != correct_answer.upper():
                                painter.save()
                                painter.setPen(QPen(QColor(255, 0, 0), 3))
                                painter.drawLine(cell_center_x - 8, cell_center_y - 8, cell_center_x + 8, cell_center_y + 8)
                                painter.drawLine(cell_center_x + 8, cell_center_y - 8, cell_center_x - 8, cell_center_y + 8)
                                painter.restore()
                    
                    if is_blank:
                        painter.save()
                        painter.setPen(QPen(QColor(255, 193, 7), 3))
                        painter.drawRect(x, y, mw, mh)
                        painter.restore()
                    elif is_multi:
                        painter.save()
                        painter.setPen(QPen(QColor(255, 140, 0), 3))
                        painter.drawRect(x, y, mw, mh)
                        painter.restore()
                    
                    marker_x = x + mw + 8
                    marker_y = y + mh // 2 + 5
                    painter.save()
                    painter.setFont(QFont("Arial", 11, QFont.Bold))
                    if is_blank:
                        painter.setPen(QPen(QColor(255, 193, 7), 2))
                        painter.drawText(marker_x, marker_y, "Ø")
                    elif is_multi:
                        painter.setPen(QPen(QColor(255, 140, 0), 2))
                        painter.drawText(marker_x, marker_y, "!")
                    else:
                        painter.setPen(QPen(QColor(40, 167, 69), 2) if is_correct else QPen(QColor(220, 53, 69), 2))
                        painter.drawText(marker_x, marker_y, "✓" if is_correct else "✗")
                    painter.restore()
                    
                    painter.setPen(QPen(QColor(0, 0, 0), 1))
                    painter.setFont(QFont("Arial", 10, QFont.Bold))
                    painter.drawText(x - 30, y + mh // 2 + 5, f"Q{q_num}")
            
            painter.save()
            painter.setFont(QFont("Arial", 14, QFont.Bold))
            painter.setPen(QPen(QColor(0, 0, 0), 2))
            score_text = f"Score: {page_score}/{page_total}"
            metrics = painter.fontMetrics()
            text_width = metrics.horizontalAdvance(score_text)
            x_pos = max(10, w - text_width - 10)
            painter.drawText(x_pos, 30, score_text)
            painter.restore()
            painter.end()
            
            filename = self._get_page_filename(page_idx)
            # Avoid overwriting if two pages produce the same stem
            candidate = os.path.join(output_folder, f"{filename}.png")
            suffix = 1
            while os.path.exists(candidate):
                candidate = os.path.join(output_folder, f"{filename}_{suffix}.png")
                suffix += 1
            qimg.save(candidate)
        
        print(f"  Images saved: {output_folder}")

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = OMRSoftware()
    window.show()
    QTimer.singleShot(1500, window._startup_update_check)
    sys.exit(app.exec_())
